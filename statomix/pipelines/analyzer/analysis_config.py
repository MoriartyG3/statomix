import shutil
import pandas as pd

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from fileverse.formats.excel import BaseExcel

from statomix.analytics.datatypes.survival import MultiClassSurv
from statomix.analytics.datatypes.survival.thresholds import MinimumPValue

SURVIVAL_MODULES = [MultiClassSurv, MinimumPValue]

class AnalysisConfig:
    def __init__(self):
        pass

    @staticmethod
    def create_analysis_config(path, datatype_map_df):
        writer = pd.ExcelWriter(path=path, engine='openpyxl')
        
        datatype_map_df.to_excel(excel_writer=writer, sheet_name="Datatype Map", index=False)
        
        writer.close()

        BaseExcel.format_cell_length(path=path)
        
        AnalysisConfig.add_survival_modules(path=path)
        
        shutil.copy(src=path, dst=path.name)

    @staticmethod
    def add_survival_modules(path):
        writer = pd.ExcelWriter(path=path, engine='openpyxl', mode="a")
        for survival_module in SURVIVAL_MODULES:
            survival_module.get_config_df().to_excel(excel_writer=writer, sheet_name=survival_module.MODULE_NAME, index=False)
        writer.close()
        
        for survival_module in SURVIVAL_MODULES:
            AnalysisConfig._add_validation_to_analysis_config_file(path=path, sheet_name=survival_module.MODULE_NAME)

    @staticmethod
    def _add_validation_to_analysis_config_file(path, sheet_name, max_row=500):
        
        center_align = Alignment(horizontal="center", vertical="center")
        
        workbook = load_workbook(filename=path)
        datatype_col_map = BaseExcel.get_worksheet_col_map(workbook['Datatype Map'])
    
        worksheet = workbook[sheet_name]
        module_col_map = BaseExcel.get_worksheet_col_map(worksheet)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)
    
        for key, value in module_col_map.items():
            cell_coordinate = datatype_col_map[key]
        
            #Add Validation
            col_cells = workbook['Datatype Map'][cell_coordinate]
            populated_rows = [c.row for c in col_cells if c.value not in (None, "")]
            last_row = max(populated_rows) if populated_rows else 1
            
            validation = DataValidation(
                type="list",
                formula1=f"'Datatype Map'!${cell_coordinate}$2:${cell_coordinate}${last_row}",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Entry",
                error="You must select from the provided drop-down menu.",
            )
            
            worksheet.add_data_validation(validation)
            validation.add(f"{value}2:{value}{max_row}")

            # #Center Values
            header_cell = worksheet[f"{value}1"]
            col_idx = header_cell.column
            for row in range(1, max_row + 1):
                cell = worksheet.cell(row=row, column=col_idx)
                if not isinstance(cell, MergedCell):
                    cell.alignment = center_align


            #Format Cell Length
            max_length = max(
                (len(str(c.value)) for c in col_cells if c.value not in (None, "")),
                default=0
            )
    
            # header length on the module sheet (not Datatype Map)
            header_cell = worksheet[f"{value}1"]
            header_length = len(str(header_cell.value)) if header_cell.value else 0
    
            adjusted_width = max(max_length, header_length) + 2  # padding
            worksheet.column_dimensions[value].width = adjusted_width
        
        workbook.save(filename=path)