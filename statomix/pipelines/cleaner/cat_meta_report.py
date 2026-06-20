import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from fileverse.formats.excel import BaseExcel

from statomix.analytics.datatypes.categorical.base import BaseCategorical

from .col.col_semantic_rules import DataTypes

@dataclass
class CatEdit:
    col_name:str
    category:str
    rename_to:str
    remove:bool

    def to_dict(self):
        return {
            "col_name":self.col_name,
            "category":self.category,
            "rename_to":self.rename_to,
            "remove":self.remove
        }

    @staticmethod
    def from_dict(data:dict):
        return CatEdit(
            col_name=data["col_name"],
            category=data["category"],
            rename_to=data.get("rename_to") if pd.notna(data.get("rename_to")) else None,
            remove=bool(data.get("remove", False))
        )

@dataclass
class CatMetaEditSchema:
    cat_edits: dict[str, dict[str, CatEdit]]
    #survival_meta: dict[str, SurvivalMeta]

    def save(self, path: Path):
        categorical_rows = []
        for col_name, categories in self.cat_edits.items():
            for category, categorical_edit in categories.items():
                categorical_rows.append(categorical_edit.to_dict())

        pd.DataFrame(categorical_rows).to_parquet(path=path)


    @staticmethod
    def load(path: Path) -> "CatMetaEditSchema":
    
        cat_df = pd.read_parquet(path=path)
    
        cat_edits: dict[str, dict[str, CatEdit]] = defaultdict(dict)
    
        for _, row in cat_df.iterrows():
            edit = CatEdit.from_dict(row.to_dict())
    
            cat_edits[edit.col_name][edit.category] = edit
    
        return CatMetaEditSchema(
            cat_edits=dict(cat_edits)
        )


class CatMetaReport:
    def __init__(self):
        pass

    def create_meta_report(self, df, col_profiles, rename_mapping, report_path):
        meta_dfs = self._get_meta_dfs(
            df=df, rename_mapping=rename_mapping, col_profiles=col_profiles
        )
        self._save_meta_report(meta_dfs=meta_dfs, report_path=report_path)
        self._add_categorical_validation(report_path=report_path, worksheet_name="Categorical")
        BaseExcel.protect_cols(path=report_path, protected_col_names=["col_name", "category", "note"], lock=True, password = "statomix")

    def _get_meta_dfs(self, df, rename_mapping, col_profiles):
        datatype_map = defaultdict(list)

        for profile in col_profiles.values():
            datatype_map[profile.col_type].append(profile.col_name)

        metadata_dfs = {}
        for datatype, col_names in datatype_map.items():
            if datatype == DataTypes.CATEGORICAL:
                metadata_dfs["Categorical"] = self._get_cat_meta_df(
                    df=df, col_names=col_names, rename_mapping=rename_mapping
                )
                
        return metadata_dfs

    def _save_meta_report(self, meta_dfs, report_path):

        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for df_name, df_meta in meta_dfs.items():

                #index = df_name in {"Categorical", "Survival Distribution"}
                index = df_name in {"Categorical"}
                df_meta.to_excel(excel_writer=writer, sheet_name=df_name, index=index)

            validation_df = self._get_validation_df()
            validation_df.to_excel(
                excel_writer=writer, sheet_name="__ValidationRanges__", index=False
            )

            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

        BaseExcel.format_cell_length(path=report_path)


    @staticmethod
    def _get_cat_meta_df(df, col_names, rename_mapping):
        distribution_dfs = []

        for col_name in col_names:

            target_name = (
                rename_mapping.get(col_name, col_name) if rename_mapping else col_name
            )

            series = df[target_name]

            distribution_df = BaseCategorical.get_distribution_df(series)

            distribution_df["col_name"] = col_name
            distribution_df["rename_to"] = ""
            distribution_df["remove"] = ""

            distribution_dfs.append(distribution_df)

        final_distribution_df = pd.concat(
            distribution_dfs,
            ignore_index=True,
        )

        other_cols = [
            c
            for c in final_distribution_df.columns
            if c not in ["col_name", "category", "rename_to", "remove"]
        ]
        ordered_cols = ["col_name", "category", "rename_to", "remove"] + other_cols
        final_distribution_df = final_distribution_df[ordered_cols]

        final_distribution_df = final_distribution_df.set_index(
            ["col_name", "category"]
        )

        return final_distribution_df

    def _get_validation_df(self):
        boolean = ["True", "False"]

        validation_df = pd.DataFrame(
            data=
            {
                "Booleans": boolean
            }
        )

        return validation_df
        
        
    def _add_categorical_validation(self, report_path, worksheet_name):
        workbook = load_workbook(filename=report_path)
        worksheet = workbook[worksheet_name]

        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

        validation_remove = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Input",
            error="Please choose only 'True' or 'False' from the dropdown list.",
        )

        worksheet.add_data_validation(validation_remove)

        validation_remove.add(
            f"{col_map["remove"]}2:{col_map["remove"]}{worksheet.max_row}"
        )

        workbook.save(filename=report_path)

    @staticmethod
    def _get_cat_edits(cat_meta_df):
        edits: dict[str, CatEdit] = defaultdict(dict)
        for (col_name, category) , row in cat_meta_df.iterrows():
            rename_to=None
            if pd.notna(row['rename_to']):
                new_name = str(row['rename_to']).strip()
                if new_name:
                    rename_to = new_name
            
            remove=False
            if pd.notna(row['remove']):
                remove=bool(row['remove'])
            
            if not (remove or rename_to is not None):
                continue
        
            edits[col_name][category] = CatEdit(
                col_name=col_name,
                category=category,
                rename_to=rename_to,
                remove=remove
            )
    
        return edits

    def get_meta_edit_schema(self, curated_meta_report):
    
        cat_meta_df = curated_meta_report.parse(sheet_name='Categorical', index_col=[0, 1])
        cat_edits = self._get_cat_edits(cat_meta_df=cat_meta_df)
    
        return CatMetaEditSchema(cat_edits=cat_edits)

    