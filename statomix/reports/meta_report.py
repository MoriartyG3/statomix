import pandas as pd
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from fileverse.formats.excel import BaseExcel

from statomix.semantic_rules import DataTypes
from statomix.analytics.datatypes.base.numerical import BaseNumerical
from statomix.analytics.datatypes.base.categorical import BaseCategorical


class MetaReport:
    def __init__(self):
        pass

    def create_meta_report(self, df, col_profiles, rename_mapping, report_path):
        meta_dfs = self._get_meta_dfs(
            df=df, rename_mapping=rename_mapping, col_profiles=col_profiles
        )
        self._save_meta_report(meta_dfs=meta_dfs, report_path=report_path)
        self._add_categorical_validation(report_path=report_path)
        self._add_survival_validation(report_path=report_path)

    def _get_meta_dfs(self, df, rename_mapping, col_profiles):
        datatype_map = defaultdict(list)

        for profile in col_profiles.values():
            datatype_map[profile.col_type].append(profile.col_name)

        metadata_dfs = {}
        for datatype, col_names in datatype_map.items():
            if datatype == DataTypes.CATEGORICAL:
                metadata_dfs["Categorical"] = self.get_categorical_meta_df(
                    df=df, col_names=col_names, rename_mapping=rename_mapping
                )

            elif datatype == DataTypes.SURVIVAL:
                metadata_dfs["Survival"] = self.get_survival_meta_df(
                    col_names=col_names
                )

        return metadata_dfs

    def _save_meta_report(self, meta_dfs, report_path):

        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for df_name, df_meta in meta_dfs.items():

                index = df_name == "Categorical"
                df_meta.to_excel(excel_writer=writer, sheet_name=df_name, index=index)

            validation_df = self._get_validation_df()
            validation_df.to_excel(
                excel_writer=writer, sheet_name="__ValidationRanges__", index=False
            )

            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

        BaseExcel.format_cell_length(report_path=report_path)

    @staticmethod
    def get_survival_meta_df(col_names):

        df = pd.DataFrame(
            {
                "col_name": col_names,
                "event": pd.NA,
                "time": pd.NA,
                "km_label": pd.NA,
            }
        )

        return df

    @staticmethod
    def get_categorical_meta_df(df, col_names, rename_mapping):
        distribution_dfs = []

        for col_name in col_names:

            target_name = (
                rename_mapping.get(col_name, col_name) if rename_mapping else col_name
            )

            series = df[target_name]

            distribution_df = BaseCategorical.get_distribution_df(series)

            distribution_df["col_name"] = col_name
            distribution_df["rename_to"] = ""
            distribution_df["remove"] = ""

            distribution_dfs.append(distribution_df)

        final_distribution_df = pd.concat(
            distribution_dfs,
            ignore_index=True,
        )

        other_cols = [
            c
            for c in final_distribution_df.columns
            if c not in ["col_name", "category", "rename_to", "remove"]
        ]
        ordered_cols = ["col_name", "category", "rename_to", "remove"] + other_cols
        final_distribution_df = final_distribution_df[ordered_cols]

        final_distribution_df = final_distribution_df.set_index(
            ["col_name", "category"]
        )

        return final_distribution_df

    def _get_validation_df(self):
        # dropdown_options = [datatype.value for datatype in DataTypes]

        # max_len = max(len(dropdown_options), 2)
        # datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        boolean = ["True", "False"]  # + [""] * (max_len - 2)

        validation_df = pd.DataFrame(data={"Booleans": boolean})

        return validation_df

    def _add_survival_validation(self, report_path):

        workbook = load_workbook(filename=report_path)
        worksheet = workbook["Survival"]
        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)
        event_col = col_map["event"]
        time_col = col_map["time"]

        validation_event = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Input",
            error="Please choose only 'True' or 'False' from the dropdown list.",
        )

        validation_time = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Input",
            error="Please choose only 'True' or 'False' from the dropdown list.",
        )

        worksheet.add_data_validation(validation_event)
        worksheet.add_data_validation(validation_time)

        validation_event.add(f"{event_col}2:{event_col}{worksheet.max_row}")

        validation_time.add(f"{time_col}2:{time_col}{worksheet.max_row}")

        # Cross-column validation:
        # Event and Time cannot both be True in the same row.
        mutual_exclusion = DataValidation(
            type="custom",
            formula1=(f'=OR({event_col}2<>"True",' f'{time_col}2<>"True")'),
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Selection",
            error="Event and Time cannot both be True for the same row.",
        )

        worksheet.add_data_validation(mutual_exclusion)

        mutual_exclusion.add(f"{event_col}2:{time_col}{worksheet.max_row}")

        workbook.save(filename=report_path)

    def _add_categorical_validation(self, report_path):
        workbook = load_workbook(filename=report_path)
        worksheet = workbook["Categorical"]

        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

        validation_remove = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Input",
            error="Please choose only 'True' or 'False' from the dropdown list.",
        )

        worksheet.add_data_validation(validation_remove)

        validation_remove.add(
            f"{col_map["remove"]}2:{col_map["remove"]}{worksheet.max_row}"
        )

        workbook.save(filename=report_path)
