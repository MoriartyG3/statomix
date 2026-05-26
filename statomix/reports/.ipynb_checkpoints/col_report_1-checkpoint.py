#Long Term Solution to Create Excel Reports

import pandas as pd
from pathlib import Path

from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.semantic_rules import DataTypes
from statomix.col_profiler import ColProfiler, ColProfile


class ColReport:
    def __init__(self, df: pd.DataFrame, df_name: str):
        self.df = df
        self.df_name = df_name

        self.col_profiler = ColProfiler(cat_unique_thresh=4, num_conversion_thresh=95)

    def create_col_report_default(self, report_path: Path):
        assert report_path.suffix == ".xlsx", "report_path should be a .xlsx path."

        self._create_col_profiles()
        self._create_col_report(report_path=report_path)
        self._format_cell_length(report_path=report_path)
        self._add_validation_datatype(report_path=report_path)
        self._add_validation_categories(
            report_path=report_path, datatype=DataTypes.CATEGORICAL.value
        )
        self._add_validation_categories(
            report_path=report_path, datatype=DataTypes.SURVIVAL.value
        )

    def _create_col_profiles(self):
        col_profiles: dict[str, ColProfile] = {}
        for col_name in self.df.columns:
            col_series = self.df[col_name]

            col_profile = self.col_profiler.get_col_profile(
                col_name=col_name,
                col_series=col_series,
            )

            col_profiles[col_name] = col_profile

        self.col_profiles = col_profiles

    def _get_col_names_by_type(self, datatype):
        return [
            profile.col_name
            for profile in self.col_profiles.values()
            if profile.col_type == datatype
        ]

    def _create_col_report(self, report_path):
        """
        Creates a raw column report without any formatting and initializes 
        a hidden sheet for scalable data validation dropdowns.
        """
        sheet_map = {
            datatype: [
                profile.col_name
                for profile in self.col_profiles.values()
                if profile.col_type == datatype
            ]
            for datatype in DataTypes
        }

        profiled_cols_n = sum(len(col_names) for col_names in sheet_map.values())
        assert profiled_cols_n == len(self.df.columns)

        # Create the .xlsx file
        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for datatype, col_names in sheet_map.items():
                rows = []

                for col_name in col_names:
                    profile = self.col_profiles[col_name]

                    if datatype == DataTypes.DATETIME:
                        rows.append(
                            {
                                "col_name": profile.col_name,
                                "change_col_name": "", 
                                "inferred_datatype": profile.col_type.value,
                                "change_datatype": "", 
                                "format": "",
                                "remove": "",
                            }
                        )
                    else:
                        rows.append(
                            {
                                "col_name": profile.col_name,
                                "change_col_name": "", 
                                "inferred_datatype": profile.col_type.value,
                                "change_datatype": "", 
                                "remove": "",
                            }
                        )

                df_sheet = pd.DataFrame(data=rows)
                sheet_name = datatype.value[:31]

                df_sheet.to_excel(
                    excel_writer=writer,
                    sheet_name=sheet_name,
                    index=False,
                )

            # --- NEW SCALABLE ARCHITECTURE: Initialize Validation Metadata Sheet ---
            dropdown_options = [datatype.value for datatype in DataTypes]
            
            # Ensure columns are equal length to build a valid DataFrame
            max_len = max(len(dropdown_options), 2)
            dt_list = dropdown_options + [""] * (max_len - len(dropdown_options))
            bool_list = ["True", "False"] + [""] * (max_len - 2)

            df_validation_lists = pd.DataFrame(data={
                "DataTypes": dt_list,
                "Booleans": bool_list
            })
            
            df_validation_lists.to_excel(
                excel_writer=writer, 
                sheet_name="__ValidationRanges__", 
                index=False
            )

    def _format_cell_length(self, report_path):
        workbook = load_workbook(filename=report_path)

        for worksheet in workbook.worksheets:
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )

                adjusted_width = max_length + 4
                column_letter = get_column_letter(column_cells[0].column)

                worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename=report_path)

    def _add_validation_datatype(self, report_path):
        workbook = load_workbook(filename=report_path)
        total_datatypes = len(DataTypes)

        for worksheet in workbook.worksheets:
            # Skip applying validation to our hidden metadata sheet
            if worksheet.title == "__ValidationRanges__":
                continue

            max_row = worksheet.max_row
            if max_row < 2:
                continue

            # 1. Add Data Type Dropdown (Referencing Column A of hidden sheet)
            dv_datatype = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!$A$2:$A${total_datatypes + 1}",
                allow_blank=True,
            )
            worksheet.add_data_validation(data_validation=dv_datatype)
            dv_datatype.add(f"D2:D{max_row}") # Note: .add() remains positional-only per openpyxl specs

            # 2. Add Boolean Dropdown (Referencing Column B of hidden sheet)
            dv_boolean = DataValidation(
                type="list",
                formula1="=__ValidationRanges__!$B$2:$B$3",
                allow_blank=True,
            )
            worksheet.add_data_validation(data_validation=dv_boolean)
            dv_boolean.add(f"E2:E{max_row}")

        workbook.save(filename=report_path)

    def _add_validation_categories(self, report_path, datatype):
        workbook = load_workbook(filename=report_path)

        if datatype not in workbook.sheetnames:
            return

        worksheet = workbook[datatype]
        val_sheet = workbook["__ValidationRanges__"]

        for cell in worksheet["A"][1:]:
            if not cell.value:
                continue

            categories = list(self.df[cell.value].dropna().unique())

            if len(categories) >= 20 or len(categories) == 0:
                continue

            # Dynamically find the next available column on the validation sheet
            next_col_idx = val_sheet.max_column + 1
            next_col_letter = get_column_letter(next_col_idx)

            # Write the category options down this new column
            val_sheet.cell(row=1, column=next_col_idx, value=f"{cell.value}_cats")
            for idx, cat_val in enumerate(iterable=categories, start=2):
                val_sheet.cell(row=idx, column=next_col_idx, value=str(cat_val))

            # Validate against the new dynamic column range
            dv_category = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!${next_col_letter}$2:${next_col_letter}${len(categories) + 1}",
                allow_blank=True,
            )

            worksheet.add_data_validation(data_validation=dv_category)
            dv_category.add(cell.coordinate) # Positional-only parameter

        # Securely hide the structural data from the end-user
        val_sheet.sheet_state = "hidden"

        workbook.save(filename=report_path)