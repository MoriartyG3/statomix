import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Protection, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.semantic_rules import DataTypes
from statomix.col_profiler import ColProfiler, ColProfile

from fileverse.logger import Logger

logger = Logger(name="col_report").get_logger()

SHEET_CELL_MAP = {
    DataTypes.IDENTIFIER.value: {
        "col_name": "A",
        "change_col_name": "B",
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
    },
    DataTypes.NUMERICAL.value: {
        "col_name": "A",
        "change_col_name": "B",
        # "units": 'C',
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
    },
    DataTypes.CATEGORICAL.value: {
        "col_name": "A",
        "change_col_name": "B",
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
    },
    DataTypes.SURVIVAL.value: {
        "col_name": "A",
        "change_col_name": "B",
        # "units": 'C',
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
    },
    DataTypes.DATETIME.value: {
        "col_name": "A",
        "change_col_name": "B",
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
        # "format": 'F',
    },
    DataTypes.FREE_TEXT.value: {
        "col_name": "A",
        "change_col_name": "B",
        "inferred_datatype": "C",
        "change_datatype": "D",
        "remove": "E",
    },
}

# EDITABLE_COL_NAMES = {
#     "change_col_name",
#     "change_datatype",
#     "remove",
#     "format",
# }

PROTECTED_COL_NAMES = [
    "col_name",
    "inferred_datatype",
]


@dataclass
class ColEdit:
    """
    User modifications for a single column.
    """

    col_name: str
    remove: bool = False
    change_col_name: str | None = None
    change_datatype: DataTypes | None = None

    def to_dict(self) -> dict:
        """Converts the instance to a dictionary for saving."""
        return {
            "col_name": self.col_name,
            "remove": self.remove,
            "change_col_name": self.change_col_name,
            # We store the value of the Enum so it can be saved as a string/primitive
            "change_datatype": (
                self.change_datatype.value if self.change_datatype else None
            ),
        }

    @staticmethod
    def from_dict(data: dict) -> "ColEdit":
        """Creates an instance from a dictionary loaded from a file."""
        return ColEdit(
            col_name=data["col_name"],
            remove=bool(data.get("remove", False)),
            change_col_name=(
                data.get("change_col_name")
                if pd.notna(data.get("change_col_name"))
                else None
            ),
            # We convert the string value back into the DataTypes Enum
            change_datatype=(
                DataTypes(data["change_datatype"])
                if pd.notna(data.get("change_datatype"))
                else None
            ),
        )


@dataclass
class ColEditSchema:
    """
    Contains only columns that have been modified by the user.
    """

    edits: dict[str, ColEdit]

    def save(
        self, path: Path
    ) -> None:  # schema: ColEditSchema, save_path: Path) -> None:
        """
        Converts the ColEditSchema to a DataFrame and saves it as a CSV.
        """
        # Convert dictionary of ColEdit objects to a list of dictionaries
        rows = [edit.to_dict() for edit in self.edits.values()]

        df = pd.DataFrame(data=rows)
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> "ColEditSchema":
        """
        Loads the CSV and reconstructs the ColEditSchema object.
        """
        df = pd.read_parquet(path=path)

        edits: dict[str, ColEdit] = {}
        for _, row in df.iterrows():
            # Use the static method you defined in the ColEdit class
            edit = ColEdit.from_dict(row.to_dict())
            edits[edit.col_name] = edit

        return cls(edits=edits)


class ColReport:
    def __init__(self):

        self.col_profiler = ColProfiler(cat_unique_thresh=4, num_conversion_thresh=95)

    def create_col_report(
        self,
        df: pd.DataFrame,
        report_path: Path,
        profiles_path: Path,
        password,
        lock,
        replace,
        rename_mapping=None,
    ):
        assert report_path.suffix == ".xlsx", "report_path should be a .xlsx path."
        if report_path.exists() and not replace:
            logger.warning(f"Column report already exists at:\n{report_path}\nSet replace=True to replace.")
            return
            
        col_profiles = self.load_col_profiles(profiles_path=profiles_path)

        self._save_col_report(
            report_path=report_path,
            col_profiles=col_profiles,
            rename_mapping=rename_mapping,
        )
        self._format_cell_length(report_path=report_path)
        self._add_validation_datatype(report_path=report_path)

        for datatype in DataTypes:
            self._add_validation_categories(
                df=df,
                report_path=report_path,
                datatype=datatype.value,
                rename_mapping=rename_mapping,
            )

        self._protect_cols(report_path=report_path, password=password, lock=lock)

    def create_col_profiles(self, df, profiles_path, replace):
        if profiles_path.exists() and not replace:
            logger.warning(f"Column profiles exists at:\n{profiles_path}\nSet replace=True to replace.")
            return

        col_profiles: dict[str, ColProfile] = {}
        for col_name in df.columns:
            col_series = df[col_name]

            col_profile = self.col_profiler.get_col_profile(
                col_name=col_name,
                col_series=col_series,
            )

            col_profiles[col_name] = col_profile

        rows = [profile.to_dict() for profile in col_profiles.values()]
        pd.DataFrame(rows).to_parquet(profiles_path)


    def get_curated_col_profiles(self, col_profiles, col_edit_schema:ColEditSchema):
        for col_name, col_edit in col_edit_schema.edits.items():
            if col_edit.remove:
                if col_name in col_profiles:
                    del col_profiles[col_name]
                    continue
                    
            if col_edit.change_col_name is not None:
                col_profiles[col_name].col_name = col_edit.change_col_name
        
            if col_edit.change_datatype is not None:
                col_profiles[col_name].col_type = col_edit.change_datatype
    
        return col_profiles

    def load_col_profiles(
        self,
        profiles_path: Path,
    ) -> dict[str, ColProfile]:

        df = pd.read_parquet(profiles_path)

        col_profiles: dict[str, ColProfile] = {}

        for _, row in df.iterrows():
            profile = ColProfile.from_dict(row)
            col_profiles[profile.col_name] = profile

        return col_profiles

    def _format_cell_length(self, report_path):
        workbook = load_workbook(filename=report_path)

        # Create a reusable alignment object (Memory efficient!)
        center_align = Alignment(horizontal="center", vertical="center")

        for worksheet in workbook.worksheets:
            # 1. Bold the header row
            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            # 2. Loop through every column exactly once
            for column_cells in worksheet.columns:
                max_length = 0

                # Look at every single cell in this specific column
                for cell in column_cells:
                    # Apply the centering alignment
                    cell.alignment = center_align

                    # Calculate the text length for the width adjuster
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))

                # Apply the final adjusted width to the column letter
                adjusted_width = max_length + 4
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename=report_path)

    def _get_validation_df(self):
        dropdown_options = [datatype.value for datatype in DataTypes]

        max_len = max(len(dropdown_options), 2)
        datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        boolean = ["True", "False"] + [""] * (max_len - 2)

        validation_df = pd.DataFrame(data={"DataTypes": datatypes, "Booleans": boolean})

        return validation_df

    def _save_col_report(self, report_path, col_profiles, rename_mapping=None):
        """
        Creates the raw col_report without validation or formatting
        """
        # datatype_map = {
        #     datatype: [
        #         profile.col_name
        #         for profile in col_profiles.values()
        #         if profile.col_type == datatype
        #     ]
        #     for datatype in DataTypes
        # }

        datatype_map = defaultdict(list)

        for profile in col_profiles.values():
            datatype_map[profile.col_type].append(profile.col_name)

        # profiled_cols_n = sum(len(col_names) for col_names in sheet_map.values())
        # assert profiled_cols_n == len(df.columns)

        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for datatype, col_names in datatype_map.items():
                if not col_names:
                    continue
                schema = SHEET_CELL_MAP[datatype.value]
                rows = []
                for col_name in col_names:
                    target_name = (
                        rename_mapping.get(col_name, col_name)
                        if rename_mapping
                        else col_name
                    )
                    profile = col_profiles[target_name]

                    row_data = {}
                    for col_header in schema.keys():
                        if col_header == "col_name":
                            row_data[col_header] = profile.col_name
                        elif col_header == "inferred_datatype":
                            row_data[col_header] = profile.col_type.value
                        else:
                            row_data[col_header] = ""  # Leave blank for user input

                    rows.append(row_data)
                df_sheet = pd.DataFrame(data=rows)
                sheet_name = datatype.value[:31]
                df_sheet.to_excel(
                    excel_writer=writer, sheet_name=sheet_name, index=False
                )

            validation_df = self._get_validation_df()

            validation_df.to_excel(
                excel_writer=writer, sheet_name="__ValidationRanges__", index=False
            )

            # writer.sheets["__ValidationRanges__"].sheet_state = "hidden"
            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

    def _add_validation_datatype(self, report_path):
        workbook = load_workbook(filename=report_path)
        total_datatypes = len(DataTypes)

        for worksheet in workbook.worksheets:
            if worksheet.title == "__ValidationRanges__":
                continue
            max_row = worksheet.max_row
            if max_row < 2:
                continue
            validation_datatype = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!$A$2:$A${total_datatypes + 1}",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Datatype",
                error="You must select a valid datatype from the provided drop-down menu.",
            )

            worksheet.add_data_validation(validation_datatype)

            cell_coordinate_datatype = SHEET_CELL_MAP[worksheet.title][
                "change_datatype"
            ]
            validation_datatype.add(
                f"{cell_coordinate_datatype}2:{cell_coordinate_datatype}{max_row}"
            )

            validation_boolean = DataValidation(
                type="list",
                formula1="=__ValidationRanges__!$B$2:$B$3",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Input",
                error="Please choose only 'True' or 'False' from the dropdown list.",
            )
            worksheet.add_data_validation(validation_boolean)
            cell_coordinate_boolean = SHEET_CELL_MAP[worksheet.title]["remove"]
            validation_boolean.add(
                f"{cell_coordinate_boolean}2:{cell_coordinate_boolean}{max_row}"
            )

        workbook.save(filename=report_path)

    def _add_validation_categories(
        self, df, report_path, datatype, rename_mapping=None
    ):

        # mapping = rename_mapping or {}
        workbook = load_workbook(filename=report_path)

        if datatype not in workbook.sheetnames:
            return

        worksheet = workbook[datatype]
        val_sheet = workbook["__ValidationRanges__"]

        for cell in worksheet["A"][1:]:
            if not cell.value:
                continue

            # target_name = mapping.get(cell.value, cell.value)
            target_name = (
                rename_mapping.get(cell.value, cell.value)
                if rename_mapping
                else cell.value
            )
            categories = list(df[target_name].dropna().unique())
            categories.sort()

            # if len(categories) >= 20 or len(categories) == 0:
            categories = categories[:20]

            # Dynamically find the next available column on the validation sheet
            next_col_idx = val_sheet.max_column + 1
            next_col_letter = get_column_letter(next_col_idx)

            # Write the category options down this new column
            val_sheet.cell(row=1, column=next_col_idx, value=f"{cell.value}_cats")
            for idx, cat_val in enumerate(iterable=categories, start=2):
                val_sheet.cell(row=idx, column=next_col_idx, value=str(cat_val))

            # Informational reference dropdown on Column A (No error restriction popup)
            validation_category = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!${next_col_letter}$2:${next_col_letter}${len(categories) + 1}",
                allow_blank=False,
                showErrorMessage=False,
            )

            worksheet.add_data_validation(validation_category)
            validation_category.add(cell.coordinate)

        workbook.save(filename=report_path)

    def _protect_cols(self, report_path: Path, lock: bool, password: str | None = None):
        workbook = load_workbook(filename=report_path)

        for worksheet in workbook.worksheets:

            if worksheet.title == "__ValidationRanges__":
                continue

            # Unlock all cells first
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)

            # Lock only protected columns
            for col_header in PROTECTED_COL_NAMES:

                col_letter = SHEET_CELL_MAP[worksheet.title][col_header]

                for cell in worksheet[col_letter]:
                    cell.protection = Protection(locked=True)

            # Enable sorting and filtering
            worksheet.auto_filter.ref = worksheet.dimensions
            # worksheet.sheet_view.showGridLines = True

            worksheet.protection.sheet = lock
            worksheet.protection.autoFilter = True
            # worksheet.protection.sort = True

            if password:
                worksheet.protection.password = password

            # Optional: allow selecting only editable cells
            worksheet.protection.enableSelection = "unlockedCells"

            # worksheet.protection.selectLockedCells = True
            # worksheet.protection.selectUnlockedCells = True

        workbook.save(filename=report_path)

    def get_col_edit_schema(self, curated_col_report):
        rename_mapping = {}
        edits: dict[str, ColEdit] = {}

        for sheet_name in curated_col_report.sheet_names:
            if sheet_name == "__ValidationRanges__":
                continue

            datatype_df = curated_col_report.parse(sheet_name=sheet_name)

            for _, row in datatype_df.iterrows():
                col_name = row["col_name"]

                remove = False
                if pd.notna(row["remove"]):
                    remove = bool(row["remove"])

                change_col_name = None
                if pd.notna(row["change_col_name"]):
                    new_name = str(row["change_col_name"]).strip()
                    if new_name:
                        change_col_name = new_name
                        rename_mapping[new_name] = col_name

                change_datatype = None
                if pd.notna(row["change_datatype"]):
                    change_datatype = DataTypes(row["change_datatype"])

                if not (
                    remove or change_col_name is not None or change_datatype is not None
                ):
                    continue

                edits[col_name] = ColEdit(
                    col_name=col_name,
                    remove=remove,
                    change_col_name=change_col_name,
                    change_datatype=change_datatype,
                )

        return rename_mapping, ColEditSchema(edits=edits)
