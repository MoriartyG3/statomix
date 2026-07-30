import re
import pandas as pd

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def sanitize(name: str) -> str:
    name = re.sub(r"[^0-9A-Za-z]+", "_", str(name).strip())
    return "N_" + name if not name or name[0].isdigit() else name


def sanitize_formula(cell_ref: str) -> str:
    expr = cell_ref
    for ch in [" ", "-", "(", ")", ".", "/", ","]:
        expr = f'SUBSTITUTE({expr},"{ch}","_")'
    return expr


# ---------------------------------------------------------------------
# STEP 1 - datatype_map is already a DataFrame (Identifier/Numerical/...
# columns, each a list of column names) -> pd.melt turns it straight into
# long rows. No manual per-category looping needed.
# ---------------------------------------------------------------------
def build_long_format_table(project, version, config_version) -> pd.DataFrame:
    frames = []
    for dataset_name, dataset in project.datasets.items():
        group_analyzer = dataset.analyzer._get_group_analyzer(
            version=version, config_version=config_version
        )
        datatype_map = group_analyzer._get_datatype_map_df()

        melted = datatype_map.melt(var_name="Datatype", value_name="Column Name").dropna()
        melted.insert(0, "Dataset Name", dataset_name)
        frames.append(melted)

    long_df = pd.concat(frames, ignore_index=True)
    # Sort so each (Dataset Name, Datatype) combo is a contiguous block
    long_df = long_df.sort_values(["Dataset Name", "Datatype"], kind="stable").reset_index(drop=True)
    long_df.insert(0, "Index", range(1, len(long_df) + 1))
    return long_df


# ---------------------------------------------------------------------
# STEP 2 - write the table, then use groupby (not a manual row-walk) to
# find each block's start/end row and register its named range.
# ---------------------------------------------------------------------
def write_raw_data_sheet(wb, long_df: pd.DataFrame):
    ws = wb.active
    ws.title = "Raw Data"

    for col_idx, h in enumerate(["Index", "Dataset Name", "Datatype", "Column Name"], start=1):
        ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)
    for r, row in enumerate(long_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)

    # Since long_df is sorted, groupby(sort=False) yields each combo's rows
    # already contiguous - min/max of the group's positional index gives
    # the block's start/end row directly (+2 = header row + 0-based offset).
    for (ds_name, dt_name), group in long_df.groupby(["Dataset Name", "Datatype"], sort=False):
        start_row, end_row = group.index.min() + 2, group.index.max() + 2
        safe_name = f"NR_{sanitize(ds_name)}_{sanitize(dt_name)}"
        wb.defined_names[safe_name] = DefinedName(
            safe_name, attr_text=f"'Raw Data'!$D${start_row}:$D${end_row}"
        )

    # Fixed lists for the two parent dropdowns
    for col, header, values, range_name in [
        (6, "DatasetNames", long_df["Dataset Name"].unique(), "DatasetList"),
        (7, "Datatypes", long_df["Datatype"].unique(), "CategoryList"),
    ]:
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=v)
        wb.defined_names[range_name] = DefinedName(
            range_name, attr_text=f"'Raw Data'!${letter}$2:${letter}${1 + len(values)}"
        )

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------
# STEP 3 - Input sheet: Dataset -> Datatype -> Column Name
# ---------------------------------------------------------------------
def write_input_sheet(wb, n_rows=50):
    ws = wb.create_sheet("Input", 0)
    # UID is column A: empty, freetext - user types a group name/ID to tag
    # rows that belong together (e.g. all rows sharing a UID become one
    # box plot's groups downstream). Not a dropdown, so it gets no
    # DataValidation and no yellow fill - just header + column width.
    headers = ["UID", "Dataset", "Datatype", "Column Name"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        #cell.font, cell.fill = Font(bold=True, color="FFFFFF"), PatternFill("solid", fgColor="4C72B0")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    last_row = 1 + n_rows

    # Dataset -> column B, Datatype -> column C, Column Name -> column D
    dv = DataValidation(type="list", formula1="=DatasetList", allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"B2:B{last_row}")

    dv = DataValidation(type="list", formula1="=CategoryList", allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"C2:C{last_row}")

    formula = f'=INDIRECT("NR_"&{sanitize_formula("$B2")}&"_"&{sanitize_formula("$C2")})'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"D2:D{last_row}")

    # Yellow fill only on dropdown columns (B, C, D) - UID (A) stays
    # plain since it's freetext, not a pick-from-list cell.
    # for r in range(2, last_row + 1):
    #     for c in range(2, 5):
    #         ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FFFF00")


# ---------------------------------------------------------------------
def _create_analysis_config(project, version, config_version, path):
    long_df = build_long_format_table(project, version, config_version)
    wb = openpyxl.Workbook()
    write_raw_data_sheet(wb, long_df)
    write_input_sheet(wb)
    wb.active = wb.sheetnames.index("Input")
    wb.save(path)
    #print(f"Saved: {output_path}  ({len(long_df)} rows)")
    #return long_df
