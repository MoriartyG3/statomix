import pandas as pd
from pathlib import Path
from dataclasses import dataclass
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from fileverse.formats.excel import BaseExcel

from statomix.semantic_rules import DataTypes
from statomix.analytics.datatypes.base.numerical import BaseNumerical
from statomix.analytics.datatypes.base.categorical import BaseCategorical

@dataclass
class CategoricalEdit:
    col_name:str
    category:str
    rename_to:str
    remove:bool

    def to_dict(self):
        return {
            "col_name":self.col_name,
            "category":self.category,
            "rename_to":self.rename_to,
            "remove":self.remove
        }

    @staticmethod
    def from_dict(data:dict):
        return CategoricalEdit(
            col_name=data["col_name"],
            category=data["category"],
            rename_to=data.get("rename_to") if pd.notna(data.get("rename_to")) else None,
            remove=bool(data.get("remove", False))
        )

# @dataclass
# class SurvivalMeta:
#     label:str
#     event:str
#     time:str

#     def to_dict(self):
#         return {
#             "label":self.label,
#             "event":self.event,
#             "time":self.time,
#         }

#     def from_dict(data):
#         return SurvivalMeta(
#             label=data["label"],
#             event=data["event"],
#             time=data["time"],
#         )

@dataclass
class CatMetaEditSchema:
    categorical_edits: dict[str, dict[str, CategoricalEdit]]
    #survival_meta: dict[str, SurvivalMeta]

    def save(self, path: Path):
        categorical_rows = []
        for col_name, categories in self.categorical_edits.items():
            for category, categorical_edit in categories.items():
                categorical_rows.append(categorical_edit.to_dict())

        # survival_rows = []
        # for label, survival_meta_object in self.survival_meta.items():
        #     survival_rows.append(survival_meta_object.to_dict())

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            pd.DataFrame(categorical_rows).to_excel(
                writer,
                sheet_name="CategoricalEdits",
                index=False,
            )
    
            # pd.DataFrame(survival_rows).to_excel(
            #     writer,
            #     sheet_name="SurvivalMeta",
            #     index=False,
            # )

    @staticmethod
    def load(path: Path) -> "CatMetaEditSchema":
    
        categorical_df = pd.read_excel(
            path,
            sheet_name="CategoricalEdits"
        )
    
        # survival_df = pd.read_excel(
        #     path,
        #     sheet_name="SurvivalMeta"
        # )
    
        categorical_edits: dict[str, dict[str, CategoricalEdit]] = defaultdict(dict)
    
        for _, row in categorical_df.iterrows():
            edit = CategoricalEdit.from_dict(row.to_dict())
    
            categorical_edits[edit.col_name][edit.category] = edit
    
        # survival_meta: dict[str, SurvivalMeta] = {}
    
        # for _, row in survival_df.iterrows():
        #     meta = SurvivalMeta.from_dict(row.to_dict())
    
        #     survival_meta[meta.label] = meta
    
        return CatMetaEditSchema(
            categorical_edits=dict(categorical_edits),
            #survival_meta=survival_meta,
        )


class CatMetaReport:
    def __init__(self):
        pass

    def create_meta_report(self, df, col_profiles, rename_mapping, report_path):
        meta_dfs = self._get_meta_dfs(
            df=df, rename_mapping=rename_mapping, col_profiles=col_profiles
        )
        self._save_meta_report(meta_dfs=meta_dfs, report_path=report_path)
        #self._add_categorical_validation(report_path=report_path)
        self._add_categorical_validation(report_path=report_path, worksheet_name="Categorical")
        #self._add_categorical_validation(report_path=report_path, worksheet_name="Survival Distribution")
        #self._add_survival_validation(report_path=report_path)
        BaseExcel.protect_cols(file_path=report_path, protected_col_names=["col_name", "category", "note"], lock=True, password = "statomix")

    def _get_meta_dfs(self, df, rename_mapping, col_profiles):
        datatype_map = defaultdict(list)

        for profile in col_profiles.values():
            datatype_map[profile.col_type].append(profile.col_name)

        metadata_dfs = {}
        for datatype, col_names in datatype_map.items():
            if datatype == DataTypes.CATEGORICAL:
                metadata_dfs["Categorical"] = self._get_categorical_meta_df(
                    df=df, col_names=col_names, rename_mapping=rename_mapping
                )

            # elif datatype == DataTypes.SURVIVAL:
            #     metadata_dfs["Survival"] = self._get_survival_meta_df(
            #         col_names=col_names
            #     )

            #     metadata_dfs["Survival Distribution"] = self._get_survival_distribution_df(
            #         df=df, 
            #         col_names=col_names, 
            #         col_profiles = col_profiles, 
            #         rename_mapping=rename_mapping
            #     )

        return metadata_dfs

    def _save_meta_report(self, meta_dfs, report_path):

        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for df_name, df_meta in meta_dfs.items():

                #index = df_name in {"Categorical", "Survival Distribution"}
                index = df_name in {"Categorical"}
                df_meta.to_excel(excel_writer=writer, sheet_name=df_name, index=index)

            validation_df = self._get_validation_df()
            validation_df.to_excel(
                excel_writer=writer, sheet_name="__ValidationRanges__", index=False
            )

            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

        BaseExcel.format_cell_length(file_path=report_path)

    # @staticmethod
    # def _get_survival_meta_df(col_names):

    #     df = pd.DataFrame(
    #         {
    #             "col_name": col_names,
    #             "type": pd.NA,
    #             #"time": pd.NA,
    #             "label": pd.NA,
    #         }
    #     )

    #     return df

    # @staticmethod
    # def _get_survival_distribution_df(df, col_names, col_profiles, rename_mapping):
    #     distribution_dfs = []
    #     rows = []
    #     for col_name in col_names:
    #         col_profile = col_profiles[col_name]
    #         if col_profile.unique_n<10 and col_profile.num_conversion_pct<10:
    #             target_name = (
    #                 rename_mapping.get(col_name, col_name) if rename_mapping else col_name
    #             )
    #             series = df[target_name]
    #             distribution_df = BaseCategorical.get_distribution_df(series)
    #             distribution_df["col_name"] = col_name
    #             distribution_df["rename_to"] = pd.NA
    #             distribution_df["remove"] = pd.NA
    #             distribution_dfs.append(distribution_df)
    #         else:
    #             rows.append({
    #                 'col_name':col_profile.col_name,
    #                 'note': "Classified as Time"
    #             })
    #     final_distribution_df = pd.concat(
    #         distribution_dfs,
    #         ignore_index=True,
    #     )
    #     final_distribution_df = pd.concat([final_distribution_df, pd.DataFrame(rows)])
    
    #     other_cols = [
    #         c
    #         for c in final_distribution_df.columns
    #         if c not in ["col_name", "category", "rename_to", "remove"]
    #     ]
    #     ordered_cols = ["col_name", "category", "rename_to", "remove"] + other_cols
    #     final_distribution_df = final_distribution_df[ordered_cols]
        
    #     final_distribution_df = final_distribution_df.set_index(
    #         ["col_name", "category"]
    #     )
        
    #     return final_distribution_df

    @staticmethod
    def _get_categorical_meta_df(df, col_names, rename_mapping):
        distribution_dfs = []

        for col_name in col_names:

            target_name = (
                rename_mapping.get(col_name, col_name) if rename_mapping else col_name
            )

            series = df[target_name]

            distribution_df = BaseCategorical.get_distribution_df(series)

            distribution_df["col_name"] = col_name
            distribution_df["rename_to"] = ""
            distribution_df["remove"] = ""

            distribution_dfs.append(distribution_df)

        final_distribution_df = pd.concat(
            distribution_dfs,
            ignore_index=True,
        )

        other_cols = [
            c
            for c in final_distribution_df.columns
            if c not in ["col_name", "category", "rename_to", "remove"]
        ]
        ordered_cols = ["col_name", "category", "rename_to", "remove"] + other_cols
        final_distribution_df = final_distribution_df[ordered_cols]

        final_distribution_df = final_distribution_df.set_index(
            ["col_name", "category"]
        )

        return final_distribution_df

    def _get_validation_df(self):
        # dropdown_options = [datatype.value for datatype in DataTypes]

        # max_len = max(len(dropdown_options), 2)
        # datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        boolean = ["True", "False"]  # + [""] * (max_len - 2)
        #survival_type = ["Event", "Time"]

        validation_df = pd.DataFrame(data=
                                     {
                                         "Booleans": boolean, 
                                         #"Survival Type": survival_type,
                                     }
                                    )

        return validation_df
        
    # def _add_survival_validation(self, report_path):
        
    #     workbook = load_workbook(filename=report_path)
    #     worksheet = workbook["Survival"]
    #     col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

    #     type_col = col_map["type"]
        
    #     validation_type = DataValidation(
    #         type="list",
    #         formula1="=__ValidationRanges__!$B$2:$B$3",
    #         allow_blank=True,
    #         showErrorMessage=True,
    #         errorStyle="stop",
    #         errorTitle="Invalid Input",
    #         error="Please choose only 'Time' or 'Event' from the dropdown list.",
    #     )

    #     worksheet.add_data_validation(validation_type)

    #     validation_type.add(f"{type_col}2:{type_col}{worksheet.max_row}")

    #     workbook.save(filename=report_path)
        
    def _add_categorical_validation(self, report_path, worksheet_name):
        workbook = load_workbook(filename=report_path)
        worksheet = workbook[worksheet_name]

        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

        validation_remove = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$A$2:$A$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Input",
            error="Please choose only 'True' or 'False' from the dropdown list.",
        )

        worksheet.add_data_validation(validation_remove)

        validation_remove.add(
            f"{col_map["remove"]}2:{col_map["remove"]}{worksheet.max_row}"
        )

        workbook.save(filename=report_path)

    @staticmethod
    def _get_categorical_edits(categorical_meta_df):
        edits: dict[str, CategoricalEdit] = defaultdict(dict)
        for (col_name, category) , row in categorical_meta_df.iterrows():
            rename_to=None
            if pd.notna(row['rename_to']):
                new_name = str(row['rename_to']).strip()
                if new_name:
                    rename_to = new_name
            
            remove=False
            if pd.notna(row['remove']):
                remove=bool(row['remove'])
            
            if not (remove or rename_to is not None):
                continue
        
            edits[col_name][category] = CategoricalEdit(
                col_name=col_name,
                category=category,
                rename_to=rename_to,
                remove=remove
            )
    
        return edits
    
    # @staticmethod
    # def _get_survival_meta(survival_meta_df):
    
    #     survival_meta :dict[str, SurvivalMeta] = {}
    #     for label, group in survival_meta_df.groupby('label'):
    #         time = group[group['type'] == 'Time']['col_name'].item()
    #         event = group[group['type'] == 'Event']['col_name'].item()
        
    #         survival_meta[label] =  SurvivalMeta(
    #             label=label,
    #             time=time,
    #             event=event
    #         )
    
    #     return survival_meta

    def get_meta_edit_schema(self, curated_meta_report):
    
        categorical_meta_df = curated_meta_report.parse(sheet_name='Categorical', index_col=[0, 1])
        categorical_edits = self._get_categorical_edits(categorical_meta_df=categorical_meta_df)
    
        # survival_meta_df = curated_meta_report.parse(sheet_name='Survival')
        # survival_meta = self._get_survival_meta(survival_meta_df=survival_meta_df)
    
        #return MetaEditSchema(categorical_edits=categorical_edits, survival_meta=survival_meta)
        return CatMetaEditSchema(categorical_edits=categorical_edits)

    
