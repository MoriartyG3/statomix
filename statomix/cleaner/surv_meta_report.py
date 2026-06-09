import pandas as pd
from enum import Enum

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from fileverse.formats.excel import BaseExcel

class SurvivalDataTypes(Enum):
    CATEGORICAL = "Categorical"
    NUMERICAL = "Numerical"
    DATETIME_FROM = "DateTime From"
    DATETIME_TO = "DateTime To"

class SurvMetaReport:
    def __init__(self):
        pass

    @staticmethod
    def _get_validation_df():
        dropdown_options = [datatype.value for datatype in SurvivalDataTypes]

        max_len = max(len(dropdown_options), 2)
        datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        survival_type = ["Event", "Time"] + [""] * (max_len - 2)
        boolean = ["True", "False"] + [""] * (max_len - 2)

        validation_df = pd.DataFrame(
            data={
                "DataTypes": datatypes,
                "Survival Type": survival_type,
                "Booleans": boolean,
            }
        ) 

        return validation_df

    @staticmethod
    def _get_overview_df(col_names):
        
        overview_df = pd.DataFrame(
            {
                'col_name':col_names,
                'change_datatype':pd.NA,
                'type':pd.NA
            }
        )

        return overview_df

    @staticmethod
    def _add_overview_validation(report_path):
        workbook = load_workbook(filename=report_path)
        worksheet = workbook["Overview"]
        
        total_datatypes = len(SurvivalDataTypes)
        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)
        
        max_row = worksheet.max_row
        if max_row < 2:
            print(f"No data to add validation")
            return
        
        validation_datatype = DataValidation(
            type="list",
            formula1=f"=__ValidationRanges__!$A$2:$A${total_datatypes + 1}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Datatype",
            error="You must select a valid datatype from the provided drop-down menu.",
        )
        
        validation_type =  DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$B$2:$B$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Datatype",
            error="You must select a valid type from the provided drop-down menu.",
        )
        
        
        worksheet.add_data_validation(validation_datatype)
        worksheet.add_data_validation(validation_type)
        
        validation_datatype.add(
            f"{col_map["change_datatype"]}2:{col_map["change_datatype"]}{worksheet.max_row}"
        )
        
        validation_type.add(
            f"{col_map["type"]}2:{col_map["type"]}{worksheet.max_row}"
        )
        
        workbook.save(filename=report_path)