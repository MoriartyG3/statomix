"""Cross-dataset analysis configuration workbook renderer."""

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.storage.atomic import atomic_output_path

KEY_SEP = "|"


def sanitize(name: str) -> str:
    """One character in -> one character out.

    Deliberately does NOT collapse runs of specials and does NOT strip, so the
    mapping is total and reproducible.  No leading-digit prefix is needed
    because every generated name is prefixed with 'NR_' by the caller.
    """
    return "".join(ch if ch.isalnum() else "_" for ch in str(name))


# ---------------------------------------------------------------------
# STEP 1 - long-format table
# ---------------------------------------------------------------------
def build_long_format_table(project, version, config_version) -> pd.DataFrame:
    frames = []
    for dataset_name, dataset in project.datasets.items():
        if dataset.dataset_role == "reference":
            continue
        group_analyzer = dataset.analyzer._get_group_analyzer(
            version=version, config_version=config_version
        )
        datatype_map = group_analyzer._get_datatype_map_df()

        melted = datatype_map.melt(
            var_name="Datatype", value_name="Column Name"
        ).dropna(subset=["Column Name"])
        melted["Column Name"] = melted["Column Name"].astype(str).str.strip()
        melted = melted[melted["Column Name"] != ""]
        melted.insert(0, "Dataset Name", dataset_name)
        frames.append(melted)

    if not frames:
        raise ValueError("No analysis datasets are available - nothing to write.")

    long_df = pd.concat(frames, ignore_index=True)
    long_df = long_df.drop_duplicates(
        subset=["Dataset Name", "Datatype", "Column Name"]
    )

    if long_df.empty:
        raise ValueError("No (dataset, datatype, column) rows were produced.")

    # Sort so each (Dataset Name, Datatype) combo is one contiguous block.
    long_df = long_df.sort_values(
        ["Dataset Name", "Datatype"], kind="stable"
    ).reset_index(drop=True)
    long_df.insert(0, "Index", range(1, len(long_df) + 1))
    return long_df


# ---------------------------------------------------------------------
# STEP 2 - hidden Raw Data sheet
#   A: Index   B: Dataset Name   C: Datatype   D: Column Name   E: Key
#   G: DatasetNames   H: Datatypes
# ---------------------------------------------------------------------
def write_raw_data_sheet(wb, long_df: pd.DataFrame) -> None:
    ws = wb.active
    ws.title = "Raw Data"

    headers = ["Index", "Dataset Name", "Datatype", "Column Name", "Key"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h).font = Font(bold=True)

    for r, row in enumerate(long_df.itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val)
        # Key column: what the Input sheet matches against.
        ws.cell(row=r, column=5, value=f"{row[1]}{KEY_SEP}{row[2]}")

    last_row = 1 + len(long_df)
    wb.defined_names["KeyList"] = DefinedName(
        "KeyList", attr_text=f"'Raw Data'!$E$2:$E${last_row}"
    )

    # Named ranges per block. Not used by the dropdowns any more (see
    # write_input_sheet) but handy for other formulas / debugging.
    seen = {}
    for (ds_name, dt_name), group in long_df.groupby(
        ["Dataset Name", "Datatype"], sort=False
    ):
        start_row, end_row = group.index.min() + 2, group.index.max() + 2
        if end_row - start_row + 1 != len(group):
            raise RuntimeError("Long-format dataset/datatype block is not contiguous.")

        safe_name = f"NR_{sanitize(ds_name)}_{sanitize(dt_name)}"[:255]
        if safe_name in seen:
            raise ValueError(
                f"Defined-name collision: {(ds_name, dt_name)} and {seen[safe_name]} "
                f"both sanitize to {safe_name!r}."
            )
        seen[safe_name] = (ds_name, dt_name)
        wb.defined_names[safe_name] = DefinedName(
            safe_name, attr_text=f"'Raw Data'!$D${start_row}:$D${end_row}"
        )

    # Fixed lists for the two parent dropdowns.
    for col, header, values, range_name in [
        (7, "DatasetNames", long_df["Dataset Name"].unique(), "DatasetList"),
        (8, "Datatypes", long_df["Datatype"].unique(), "CategoryList"),
    ]:
        letter = get_column_letter(col)
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)
        for i, v in enumerate(values, start=2):
            ws.cell(row=i, column=col, value=v)
        wb.defined_names[range_name] = DefinedName(
            range_name, attr_text=f"'Raw Data'!${letter}$2:${letter}${1 + len(values)}"
        )

    ws.sheet_state = "hidden"


# ---------------------------------------------------------------------
# STEP 3 - Input sheet
# ---------------------------------------------------------------------
def write_input_sheet(
    wb,
    n_rows: int = 50,
    n_col_selectors: int = 10,
) -> None:
    ws = wb.create_sheet("Input", 0)

    headers = ["Title", "Subtitle", "Dataset", "Datatype"] + [
        f"Column Name {i}" for i in range(1, n_col_selectors + 1)
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = 24

    # Column positions are derived from `headers`, never hard-coded, so
    # inserting a column can't silently desync the validations.
    ds_col = get_column_letter(headers.index("Dataset") + 1)
    dt_col = get_column_letter(headers.index("Datatype") + 1)
    first_sel = headers.index("Column Name 1") + 1

    last_row = 1 + n_rows

    dv_ds = DataValidation(type="list", formula1="=DatasetList", allow_blank=True)
    ws.add_data_validation(dv_ds)
    dv_ds.add(f"{ds_col}2:{ds_col}{last_row}")

    dv_dt = DataValidation(type="list", formula1="=CategoryList", allow_blank=True)
    ws.add_data_validation(dv_dt)
    dv_dt.add(f"{dt_col}2:{dt_col}{last_row}")

    # Column-name selectors, all driven by the SAME row's Dataset + Datatype.
    # OFFSET/MATCH against the Key column instead of INDIRECT+SUBSTITUTE:
    # no name sanitising to keep in sync, and it stays well under Excel's
    # 255-character limit for a data-validation formula.
    key = f'${ds_col}2&"{KEY_SEP}"&${dt_col}2'
    formula = (
        f"=OFFSET('Raw Data'!$D$1,MATCH({key},KeyList,0),0,COUNTIF(KeyList,{key}),1)"
    )
    if len(formula) > 255:
        raise ValueError(f"Data-validation formula is too long ({len(formula)} chars).")

    dv_col = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv_col)
    for i in range(n_col_selectors):
        letter = get_column_letter(first_sel + i)
        dv_col.add(f"{letter}2:{letter}{last_row}")

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------
def _create_analysis_config(
    project,
    version: int,
    config_version: int,
    path,
    n_rows: int = 50,
    n_col_selectors: int = 10,
) -> pd.DataFrame:
    long_df = build_long_format_table(project, version, config_version)
    wb = openpyxl.Workbook()
    write_raw_data_sheet(wb, long_df)
    write_input_sheet(wb, n_rows=n_rows, n_col_selectors=n_col_selectors)
    wb.active = wb.sheetnames.index("Input")
    with atomic_output_path(destination=path) as temporary_path:
        wb.save(temporary_path)
    return long_df
