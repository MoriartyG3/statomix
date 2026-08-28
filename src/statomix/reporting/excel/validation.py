"""Reusable Excel list-validation rendering."""

from __future__ import annotations

from pathlib import Path

from fileverse.formats.excel import BaseExcel
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation


def add_datatype_list_validations(
    *,
    path: Path,
    sheet_name: str,
    max_row: int = 500,
    style_sheet: bool = False,
) -> None:
    """Add non-empty datatype dropdowns to one analysis worksheet."""

    if max_row < 2:
        raise ValueError("max_row must be at least 2.")

    workbook = load_workbook(filename=path)
    datatype_sheet = workbook["Datatype Map"]
    datatype_col_map = BaseExcel.get_worksheet_col_map(datatype_sheet)
    worksheet = workbook[sheet_name]
    module_col_map = BaseExcel.get_worksheet_col_map(worksheet)

    if style_sheet:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)

    center_alignment = Alignment(horizontal="center", vertical="center")
    for key, target_coordinate in module_col_map.items():
        source_coordinate = datatype_col_map[key]
        source_cells = datatype_sheet[source_coordinate]
        populated_rows = [
            cell.row
            for cell in source_cells
            if cell.row >= 2 and cell.value not in (None, "")
        ]

        if populated_rows:
            validation = DataValidation(
                type="list",
                formula1=(
                    f"'Datatype Map'!${source_coordinate}$2:"
                    f"${source_coordinate}${max(populated_rows)}"
                ),
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Entry",
                error="You must select from the provided drop-down menu.",
            )
            worksheet.add_data_validation(validation)
            validation.add(f"{target_coordinate}2:{target_coordinate}{max_row}")

        if not style_sheet:
            continue

        header_cell = worksheet[f"{target_coordinate}1"]
        for row in range(1, max_row + 1):
            cell = worksheet.cell(row=row, column=header_cell.column)
            if not isinstance(cell, MergedCell):
                cell.alignment = center_alignment

        content_width = max(
            (
                len(str(cell.value))
                for cell in source_cells
                if cell.value not in (None, "")
            ),
            default=0,
        )
        header_width = len(str(header_cell.value)) if header_cell.value else 0
        worksheet.column_dimensions[target_coordinate].width = (
            max(content_width, header_width) + 2
        )

    workbook.save(filename=path)
