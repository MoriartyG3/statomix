from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import ClassVar

import pandas as pd
from fileverse.formats.excel import BaseExcel
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.pipelines.artifacts import frame_from_rows
from statomix.pipelines.cleaner.cat_meta_report import CatMetaReport

from .surv_profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
    get_survival_sematic_col_profile,
)


@dataclass
class SurvCatEdit:
    col_name: str
    category: str
    rename_to: str | None
    remove: bool

    def to_dict(self):
        return {
            "col_name": self.col_name,
            "category": self.category,
            "rename_to": self.rename_to,
            "remove": self.remove,
        }

    @staticmethod
    def from_dict(data: dict):
        return SurvCatEdit(
            col_name=data["col_name"],
            category=data["category"],
            rename_to=(
                data.get("rename_to") if pd.notna(data.get("rename_to")) else None
            ),
            remove=bool(data.get("remove", False)),
        )


@dataclass
class SurvCatMetaEditSchema:
    cat_edits: dict[str, dict[str, SurvCatEdit]]

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "col_name": "object",
        "category": "object",
        "rename_to": "object",
        "remove": "boolean",
    }

    @classmethod
    def empty(cls) -> "SurvCatMetaEditSchema":
        return cls(cat_edits={})

    @property
    def edit_count(self) -> int:
        return sum(len(edits) for edits in self.cat_edits.values())

    @property
    def is_empty(self) -> bool:
        return self.edit_count == 0

    def save(self, path: Path):
        rows = []
        for categories in self.cat_edits.values():
            for categorical_edit in categories.values():
                rows.append(categorical_edit.to_dict())

        categorical_df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        categorical_df.to_parquet(path=path, index=False)

    @staticmethod
    def load(path: Path) -> "SurvCatMetaEditSchema":
        categorical_df = pd.read_parquet(path=path)

        categorical_edits: dict[str, dict[str, SurvCatEdit]] = defaultdict(dict)

        for _, row in categorical_df.iterrows():
            edit = SurvCatEdit.from_dict(row.to_dict())
            categorical_edits[edit.col_name][edit.category] = edit

        return SurvCatMetaEditSchema(
            cat_edits=dict(categorical_edits),
        )


@dataclass
class SurvEdit:
    col_name: str
    change_datatype: SurvivalDataTypes | None = None
    remove: bool = False

    def to_dict(self):
        return {
            "col_name": self.col_name,
            "change_datatype": (
                self.change_datatype.value if self.change_datatype is not None else None
            ),
            "remove": self.remove,
        }

    @staticmethod
    def from_dict(data: dict):
        return SurvEdit(
            col_name=data["col_name"],
            change_datatype=(
                SurvivalDataTypes(data["change_datatype"])
                if pd.notna(data.get("change_datatype"))
                else None
            ),
            remove=data["remove"],
        )


@dataclass
class SurvEditSchema:
    edits: dict[str, SurvEdit]

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "col_name": "object",
        "change_datatype": "object",
        "remove": "boolean",
    }

    @classmethod
    def empty(cls) -> "SurvEditSchema":
        return cls(edits={})

    @property
    def is_empty(self) -> bool:
        return not self.edits

    def save(self, path: Path) -> None:
        """Save this edit schema as a parquet artifact."""
        rows = [edit.to_dict() for edit in self.edits.values()]

        df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> "SurvEditSchema":
        """Load a parquet artifact and reconstruct the edit schema."""
        df = pd.read_parquet(path=path)

        edits: dict[str, SurvEdit] = {}
        for _, row in df.iterrows():
            edit = SurvEdit.from_dict(row.to_dict())
            edits[edit.col_name] = edit

        return cls(edits=edits)


@dataclass
class SurvPair:
    surv_label: str
    event_profile: SurvivalSemanticProfile
    time_profile: SurvivalSemanticProfile

    def to_dict(self):
        return {
            "surv_label": self.surv_label,
            "event_profile": self.event_profile.to_dict(),
            "time_profile": self.time_profile.to_dict(),
        }

    @classmethod
    def from_dict(cls, data):
        return cls(
            surv_label=data["surv_label"],
            event_profile=SurvivalSemanticProfile.from_dict(data["event_profile"]),
            time_profile=SurvivalSemanticProfile.from_dict(data["time_profile"]),
        )


@dataclass
class SurvPairs:
    pairs: dict[str, SurvPair]

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "surv_label": "object",
        "event_profile": "object",
        "time_profile": "object",
    }

    @classmethod
    def empty(cls) -> "SurvPairs":
        return cls(pairs={})

    @property
    def is_empty(self) -> bool:
        return not self.pairs

    def save(self, path: Path) -> None:
        rows = [pair.to_dict() for pair in self.pairs.values()]

        df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> "SurvPairs":
        df = pd.read_parquet(path)

        pairs: dict[str, SurvPair] = {}

        for _, row in df.iterrows():
            pair = SurvPair.from_dict(row.to_dict())
            pairs[pair.surv_label] = pair

        return cls(pairs=pairs)


class SurvMetaReport:
    def __init__(self):
        pass

    def create_surv_report(self, col_names, report_path, profiles_path):

        self.create_semantic_profiles(col_names=col_names, path=profiles_path)
        semantic_profiles = self.load_semantic_profiles(path=profiles_path)

        self.save_col_report(
            path=report_path,
            col_names=col_names,
            semantic_profiles=semantic_profiles,
        )

        BaseExcel.format_cell_length(path=report_path)
        self._add_validation_datatype(report_path=report_path)
        BaseExcel.protect_cols(
            path=report_path,
            protected_col_names=["col_name", "inferred_datatype"],
            password="statomix",
            lock=True,
        )

    def create_semantic_profiles(self, col_names, path):
        semantic_profiles: dict[str, SurvivalSemanticProfile] = {}
        for col_name in col_names:
            semantic_profiles[col_name] = get_survival_sematic_col_profile(col_name)

        self.save_semantic_profiles(semantic_profiles=semantic_profiles, path=path)

    def save_semantic_profiles(self, semantic_profiles, path):
        rows = [profile.to_dict() for profile in semantic_profiles.values()]
        profiles_df = frame_from_rows(
            rows=rows,
            schema=SurvivalSemanticProfile.PARQUET_SCHEMA,
        )
        profiles_df.to_parquet(path=path, index=False)

    @staticmethod
    def load_semantic_profiles(path: Path):
        df = pd.read_parquet(path)

        semantic_profiles: dict[str, SurvivalSemanticProfile] = {}

        for _, row in df.iterrows():
            profile = SurvivalSemanticProfile.from_dict(row)
            semantic_profiles[profile.col_name] = profile

        return semantic_profiles

    @staticmethod
    def _get_validation_df():
        dropdown_options = [datatype.value for datatype in SurvivalDataTypes]

        max_len = max(len(dropdown_options), 2)
        datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        survival_type = ["Event", "Time"] + [""] * (max_len - 2)
        boolean = ["True", "False"] + [""] * (max_len - 2)

        validation_df = pd.DataFrame(
            data={
                "DataTypes": datatypes,
                "Survival Type": survival_type,
                "Booleans": boolean,
            }
        )

        return validation_df

    @staticmethod
    def save_col_report(path, col_names, semantic_profiles):
        rows = []
        for col_name in col_names:
            rows.append(
                {
                    "col_name": col_name,
                    "inferred_datatype": semantic_profiles[col_name].col_type.value,
                    "change_datatype": pd.NA,
                    "survival_label": pd.NA,
                    "remove": pd.NA,
                }
            )

        writer = pd.ExcelWriter(path=path, engine="openpyxl")

        report_columns = [
            "col_name",
            "inferred_datatype",
            "change_datatype",
            "survival_label",
            "remove",
        ]
        pd.DataFrame(data=rows, columns=report_columns).to_excel(
            excel_writer=writer, index=False, sheet_name="SurvMeta"
        )

        validation_df = SurvMetaReport._get_validation_df()
        validation_df.to_excel(
            excel_writer=writer, sheet_name="__ValidationRanges__", index=False
        )
        writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

        writer.close()

    @staticmethod
    def _add_validation_datatype(report_path):

        workbook = load_workbook(filename=report_path)
        total_datatypes = len(SurvivalDataTypes)

        worksheet = workbook["SurvMeta"]

        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

        max_row = worksheet.max_row
        if max_row < 2:
            print("No data to add validation")
            return

        validation_datatype = DataValidation(
            type="list",
            formula1=f"=__ValidationRanges__!$A$2:$A${total_datatypes + 1}",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Datatype",
            error="You must select a valid datatype from the provided drop-down menu.",
        )
        worksheet.add_data_validation(validation_datatype)
        validation_datatype.add(
            f"{col_map["change_datatype"]}2:{col_map["change_datatype"]}{worksheet.max_row}"
        )

        validation_remove = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$C$2:$C$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Datatype",
            error="You must select an option from the provided drop-down menu.",
        )
        worksheet.add_data_validation(validation_remove)
        validation_remove.add(
            f"{col_map["remove"]}2:{col_map["remove"]}{worksheet.max_row}"
        )

        workbook.save(filename=report_path)

    @staticmethod
    def _get_surv_edits(surv_meta_df):

        edits: dict[str, SurvEdit] = {}

        for _, row in surv_meta_df.iterrows():
            col_name = row["col_name"]

            change_datatype = None
            if pd.notna(row["change_datatype"]):
                change_datatype = SurvivalDataTypes(row["change_datatype"])

            remove = False
            if pd.notna(row["remove"]):
                remove = bool(row["remove"])

            if not (remove or change_datatype is not None):
                continue

            edits[col_name] = SurvEdit(
                col_name=col_name, remove=remove, change_datatype=change_datatype
            )

        return edits

    def get_surv_edit_schema(self, curated_meta_report):
        surv_meta_df = curated_meta_report.parse("SurvMeta")
        edits = self._get_surv_edits(surv_meta_df=surv_meta_df)

        return SurvEditSchema(edits=edits)

    @staticmethod
    def get_curated_surv_profiles(meta_edit_schema, surv_profiles):

        for col_name, surv_edit in meta_edit_schema.edits.items():
            if surv_edit.remove and col_name in surv_profiles:
                del surv_profiles[col_name]
                continue

            if surv_edit.change_datatype is not None:
                surv_profiles[col_name].col_type = surv_edit.change_datatype

        return surv_profiles

    @staticmethod
    def get_surv_pairs(surv_meta_df, surv_profiles):

        pairs: dict[str, SurvPair] = {}

        required_types = {
            SurvivalDataTypes.EVENT.value,
            SurvivalDataTypes.TIME.value,
        }

        for surv_label, surv_group in surv_meta_df.groupby("survival_label"):
            if not surv_label:
                continue

            if len(surv_group) != 2:
                raise ValueError(
                    f"Survival label '{surv_label}' must have exactly 2 rows, "
                    f"found {len(surv_group)}."
                )

            found_types = set(surv_group["inferred_datatype"])

            if found_types != required_types:
                raise ValueError(
                    f"Survival label '{surv_label}' must contain exactly one "
                    f"'{SurvivalDataTypes.EVENT.value}' and one "
                    f"'{SurvivalDataTypes.TIME.value}' row. "
                    f"Found: {sorted(found_types)}."
                )

            group_by_type = surv_group.set_index("inferred_datatype")

            event_col = group_by_type.at[
                SurvivalDataTypes.EVENT.value,
                "col_name",
            ]

            time_col = group_by_type.at[
                SurvivalDataTypes.TIME.value,
                "col_name",
            ]

            pairs[surv_label] = SurvPair(
                surv_label=surv_label,
                event_profile=surv_profiles[event_col],
                time_profile=surv_profiles[time_col],
            )

        surv_pairs = SurvPairs(pairs=pairs)

        return surv_pairs

    @staticmethod
    def save_cat_meta_report(df, rename_mapping, report_path, profiles_path):
        semantic_profiles = SurvMetaReport.load_semantic_profiles(path=profiles_path)
        cat_col_names = []
        for col_name, semantic_profile in semantic_profiles.items():
            if semantic_profile.col_type == SurvivalDataTypes.EVENT:
                cat_col_names.append(col_name)

        surv_cat_meta_df = CatMetaReport._get_cat_meta_df(
            df=df, col_names=cat_col_names, rename_mapping=rename_mapping
        )

        writer = pd.ExcelWriter(path=report_path, engine="openpyxl")

        surv_cat_meta_df.to_excel(excel_writer=writer, sheet_name="SurvCatMeta")

        validation_df = SurvMetaReport._get_validation_df()
        validation_df.to_excel(
            excel_writer=writer, sheet_name="__ValidationRanges__", index=False
        )
        writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"
        writer.close()

    @staticmethod
    def _add_surv_cat_validation(report_path):
        workbook = load_workbook(filename=report_path)
        worksheet = workbook["SurvCatMeta"]

        col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

        validation_remove = DataValidation(
            type="list",
            formula1="=__ValidationRanges__!$C$2:$C$3",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="stop",
            errorTitle="Invalid Datatype",
            error="You must select an option from the provided drop-down menu.",
        )
        worksheet.add_data_validation(validation_remove)
        validation_remove.add(
            f"{col_map["remove"]}2:{col_map["remove"]}{worksheet.max_row}"
        )
        workbook.save(filename=report_path)

    def create_cat_meta_report(self, df, rename_mapping, profiles_path, report_path):
        self.save_cat_meta_report(
            df=df,
            rename_mapping=rename_mapping,
            profiles_path=profiles_path,
            report_path=report_path,
        )

        BaseExcel.format_cell_length(path=report_path)
        self._add_surv_cat_validation(report_path=report_path)
        BaseExcel.protect_cols(
            path=report_path,
            protected_col_names=["col_name", "category", "count", "percentage"],
            password="statomix",
            lock=True,
        )

    @staticmethod
    def _get_surv_cat_edits(surv_cat_meta_df):
        edits: dict[str, SurvCatEdit] = defaultdict(dict)
        for (col_name, category), row in surv_cat_meta_df.iterrows():
            rename_to = None
            if pd.notna(row["rename_to"]):
                new_name = str(row["rename_to"]).strip()
                if new_name:
                    rename_to = new_name

            remove = False
            if pd.notna(row["remove"]):
                remove = bool(row["remove"])

            if not (remove or rename_to is not None):
                continue

            edits[col_name][category] = SurvCatEdit(
                col_name=col_name, category=category, rename_to=rename_to, remove=remove
            )

        return edits

    def get_surv_cat_meta_edit_schema(self, curated_meta_report):
        surv_cat_meta_df = curated_meta_report.parse(
            sheet_name="SurvCatMeta", index_col=[0, 1]
        )
        surv_cat_edits = self._get_surv_cat_edits(surv_cat_meta_df=surv_cat_meta_df)

        return SurvCatMetaEditSchema(cat_edits=surv_cat_edits)
