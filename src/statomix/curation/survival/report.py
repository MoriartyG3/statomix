"""Survival edit schemas, endpoint pairs, and Excel curation reports."""

from __future__ import annotations

from collections import Counter, defaultdict
from collections.abc import Mapping
from dataclasses import dataclass, replace
from pathlib import Path
from types import MappingProxyType
from typing import ClassVar

import pandas as pd
from fileverse.formats.excel import BaseExcel
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.core.tabular import frame_from_rows
from statomix.logging import get_logger

from .declarations import (
    DEFAULT_EVENT_STRUCTURE,
    DEFAULT_OBSERVATION_SCHEME,
    EVENT_STRUCTURES,
    OBSERVATION_SCHEMES,
    has_declaration_fields,
    read_survival_declaration,
    require_supported_survival,
)
from .events import (
    decode_category_scalar,
    encode_category_scalar,
    is_blank_cell,
    is_canonical_event_value,
    is_missing_scalar,
    parse_optional_event_observed,
    parse_remove_instruction,
)
from .profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
    get_survival_semantic_col_profile,
)

logger = get_logger(name="survival_report")


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvCatEdit:
    """One explicit mapping or removal of a source event category."""

    col_name: str
    category: object
    category_encoding: str
    event_observed: bool | None
    remove: bool

    def __post_init__(self) -> None:
        if not isinstance(self.col_name, str) or not self.col_name.strip():
            raise ValueError("col_name must be a non-empty string.")

        expected_encoding = encode_category_scalar(self.category)
        if self.category_encoding != expected_encoding:
            raise ValueError(
                f"Inconsistent category identity for {self.col_name!r}, "
                f"category {self.category!r}."
            )

        normalized_event = parse_optional_event_observed(
            self.event_observed,
            col_name=self.col_name,
            category=self.category,
        )
        normalized_remove = parse_remove_instruction(
            self.remove,
            col_name=self.col_name,
            category=self.category,
        )

        if normalized_remove and normalized_event is not None:
            raise ValueError(
                f"Category {self.category!r} in {self.col_name!r} cannot "
                "be both mapped and removed."
            )
        if not normalized_remove and normalized_event is None:
            raise ValueError(
                f"Category {self.category!r} in {self.col_name!r} has no action."
            )

        object.__setattr__(self, "event_observed", normalized_event)
        object.__setattr__(self, "remove", normalized_remove)

    def to_dict(self) -> dict[str, object]:
        return {
            "col_name": self.col_name,
            "category": str(self.category),
            "category_encoding": self.category_encoding,
            "event_observed": self.event_observed,
            "remove": self.remove,
        }

    @classmethod
    def from_dict(cls, data: Mapping[str, object]) -> SurvCatEdit:
        stored_encoding = data.get("category_encoding")

        if is_blank_cell(stored_encoding):
            # Legacy schemas stored the source category directly.
            category = data["category"]
        else:
            category = decode_category_scalar(stored_encoding)

        category_encoding = encode_category_scalar(category)

        # Safe legacy Boolean/0/1 rename instructions remain readable.
        event_field = "event_observed" if "event_observed" in data else "rename_to"
        event_observed = parse_optional_event_observed(
            data.get(event_field),
            col_name=data["col_name"],
            category=category,
        )
        remove = parse_remove_instruction(
            data.get("remove"),
            col_name=data["col_name"],
            category=category,
        )

        return cls(
            col_name=data["col_name"],
            category=category,
            category_encoding=category_encoding,
            event_observed=event_observed,
            remove=remove,
        )


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvCatMetaEditSchema:
    """Immutable, typed survival-event curation instructions."""

    cat_edits: Mapping[str, Mapping[str, SurvCatEdit]]

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "col_name": "object",
        "category": "object",
        "category_encoding": "object",
        "event_observed": "boolean",
        "remove": "boolean",
    }

    def __post_init__(self) -> None:
        nested = {}

        for col_name, column_edits in self.cat_edits.items():
            normalized_edits = {}

            for edit in column_edits.values():
                if edit.col_name != col_name:
                    raise ValueError(
                        f"Schema column {col_name!r} contains an edit for "
                        f"{edit.col_name!r}."
                    )

                identity = edit.category_encoding
                if identity in normalized_edits:
                    raise ValueError(
                        f"Duplicate survival-event edit for {col_name!r}, "
                        f"category {edit.category!r}."
                    )

                normalized_edits[identity] = edit

            nested[col_name] = MappingProxyType(normalized_edits)

        object.__setattr__(self, "cat_edits", MappingProxyType(nested))

    @classmethod
    def empty(cls) -> SurvCatMetaEditSchema:
        return cls(cat_edits={})

    @property
    def edit_count(self) -> int:
        return sum(len(edits) for edits in self.cat_edits.values())

    @property
    def is_empty(self) -> bool:
        return self.edit_count == 0

    def save(self, path: Path) -> None:
        rows = [
            edit.to_dict()
            for column_edits in self.cat_edits.values()
            for edit in column_edits.values()
        ]

        frame = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        frame = frame.astype(
            {
                "event_observed": "boolean",
                "remove": "boolean",
            }
        )
        frame.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> SurvCatMetaEditSchema:
        frame = pd.read_parquet(path=path)

        required_columns = {"col_name", "category", "remove"}
        missing_columns = required_columns - set(frame.columns)
        if missing_columns:
            raise ValueError(
                "Survival-event schema is missing columns: "
                f"{sorted(missing_columns)!r}."
            )

        if not {"event_observed", "rename_to"}.intersection(frame.columns):
            raise ValueError(
                "Survival-event schema has no event_observed or legacy "
                "rename_to field."
            )

        edits: dict[str, dict[str, SurvCatEdit]] = defaultdict(dict)

        for record in frame.to_dict(orient="records"):
            edit = SurvCatEdit.from_dict(record)
            column_edits = edits[edit.col_name]

            if edit.category_encoding in column_edits:
                raise ValueError(
                    f"Duplicate persisted event edit for {edit.col_name!r}, "
                    f"category {edit.category!r}."
                )

            column_edits[edit.category_encoding] = edit

        return cls(cat_edits=dict(edits))


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvEdit:
    col_name: str
    change_datatype: SurvivalDataTypes | None = None
    remove: bool = False

    def to_dict(self):
        return {
            "col_name": self.col_name,
            "change_datatype": (
                self.change_datatype.value if self.change_datatype is not None else None
            ),
            "remove": self.remove,
        }

    @staticmethod
    def from_dict(data: dict):
        return SurvEdit(
            col_name=data["col_name"],
            change_datatype=(
                SurvivalDataTypes(data["change_datatype"])
                if pd.notna(data.get("change_datatype"))
                else None
            ),
            remove=data["remove"],
        )


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvEditSchema:
    edits: Mapping[str, SurvEdit]

    def __post_init__(self) -> None:
        object.__setattr__(self, "edits", MappingProxyType(dict(self.edits)))

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "col_name": "object",
        "change_datatype": "object",
        "remove": "boolean",
    }

    @classmethod
    def empty(cls) -> SurvEditSchema:
        return cls(edits={})

    @property
    def is_empty(self) -> bool:
        return not self.edits

    def save(self, path: Path) -> None:
        """Save this edit schema as a parquet artifact."""
        rows = [edit.to_dict() for edit in self.edits.values()]

        df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> SurvEditSchema:
        """Load a parquet artifact and reconstruct the edit schema."""
        df = pd.read_parquet(path=path)

        edits: dict[str, SurvEdit] = {}
        for _, row in df.iterrows():
            edit = SurvEdit.from_dict(row.to_dict())
            edits[edit.col_name] = edit

        return cls(edits=edits)


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvPair:
    surv_label: str
    event_profile: SurvivalSemanticProfile
    time_profile: SurvivalSemanticProfile
    event_structure: str = DEFAULT_EVENT_STRUCTURE
    observation_scheme: str = DEFAULT_OBSERVATION_SCHEME

    def __post_init__(self) -> None:
        event_structure, observation_scheme = read_survival_declaration(
            {
                "event_structure": self.event_structure,
                "observation_scheme": self.observation_scheme,
            }
        )
        object.__setattr__(self, "event_structure", event_structure)
        object.__setattr__(self, "observation_scheme", observation_scheme)

    def to_dict(self):
        return {
            "surv_label": self.surv_label,
            "event_profile": self.event_profile.to_dict(),
            "time_profile": self.time_profile.to_dict(),
            "event_structure": self.event_structure,
            "observation_scheme": self.observation_scheme,
        }

    @classmethod
    def from_dict(cls, data):
        event_structure, observation_scheme = read_survival_declaration(data)

        return cls(
            surv_label=data["surv_label"],
            event_profile=SurvivalSemanticProfile.from_dict(data["event_profile"]),
            time_profile=SurvivalSemanticProfile.from_dict(data["time_profile"]),
            event_structure=event_structure,
            observation_scheme=observation_scheme,
        )

    def require_supported(self, *, operation: str) -> None:
        require_supported_survival(
            surv_label=self.surv_label,
            event_structure=self.event_structure,
            observation_scheme=self.observation_scheme,
            operation=operation,
        )


@dataclass(frozen=True, slots=True, kw_only=True)
class SurvPairs:
    pairs: Mapping[str, SurvPair]

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "surv_label": "object",
        "event_profile": "object",
        "time_profile": "object",
        "event_structure": "object",
        "observation_scheme": "object",
    }

    def __post_init__(self):
        object.__setattr__(
            self,
            "pairs",
            MappingProxyType(dict(self.pairs)),
        )

    @classmethod
    def empty(cls):
        return cls(pairs={})

    @property
    def is_empty(self):
        return not self.pairs

    def require_supported(self, *, operation: str) -> None:
        for pair in self.pairs.values():
            pair.require_supported(operation=operation)

    def save(self, path):
        rows = [pair.to_dict() for pair in self.pairs.values()]
        df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path):
        df = pd.read_parquet(path)

        # Validate the schema even when there are no endpoint rows.
        has_declaration_fields(df.columns)

        pairs = {}
        for row in df.to_dict(orient="records"):
            pair = SurvPair.from_dict(row)

            if pair.surv_label in pairs:
                raise ValueError(f"Duplicate survival endpoint: {pair.surv_label!r}.")

            pairs[pair.surv_label] = pair

        return cls(pairs=pairs)


class SurvMetaReport:
    def __init__(self):
        pass

    def create_surv_report(self, col_names, report_path, profiles_path):

        self.create_semantic_profiles(col_names=col_names, path=profiles_path)
        semantic_profiles = self.load_semantic_profiles(path=profiles_path)

        self.create_surv_report_from_profiles(
            col_names=col_names,
            semantic_profiles=semantic_profiles,
            report_path=report_path,
        )

    def create_surv_report_from_profiles(
        self,
        *,
        col_names,
        semantic_profiles,
        report_path,
        survival_labels=None,
    ):
        """Render a report from already-resolved survival semantics."""

        self.save_col_report(
            path=report_path,
            col_names=col_names,
            semantic_profiles=semantic_profiles,
            survival_labels=survival_labels,
        )

        BaseExcel.format_cell_length(path=report_path)
        self._add_validation_datatype(report_path=report_path)
        BaseExcel.protect_cols(
            path=report_path,
            protected_col_names=["col_name", "inferred_datatype"],
            password="statomix",
            lock=True,
        )

    def create_semantic_profiles(self, col_names, path):
        semantic_profiles: dict[str, SurvivalSemanticProfile] = {}
        for col_name in col_names:
            semantic_profiles[col_name] = get_survival_semantic_col_profile(
                col_name=col_name
            )

        self.save_semantic_profiles(semantic_profiles=semantic_profiles, path=path)

    def save_semantic_profiles(self, semantic_profiles, path):
        rows = [profile.to_dict() for profile in semantic_profiles.values()]
        profiles_df = frame_from_rows(
            rows=rows,
            schema=SurvivalSemanticProfile.PARQUET_SCHEMA,
        )
        profiles_df.to_parquet(path=path, index=False)

    @staticmethod
    def load_semantic_profiles(path: Path):
        df = pd.read_parquet(path)

        semantic_profiles: dict[str, SurvivalSemanticProfile] = {}

        for _, row in df.iterrows():
            profile = SurvivalSemanticProfile.from_dict(row)
            semantic_profiles[profile.col_name] = profile

        return semantic_profiles

    @staticmethod
    def _get_validation_df():
        options = {
            "DataTypes": [datatype.value for datatype in SurvivalDataTypes],
            "Survival Type": ["Event", "Time"],
            "Booleans": ["True", "False"],
            "Event Structure": list(EVENT_STRUCTURES),
            "Observation Scheme": list(OBSERVATION_SCHEMES),
        }

        max_length = max(len(values) for values in options.values())

        return pd.DataFrame(
            {
                name: values + [""] * (max_length - len(values))
                for name, values in options.items()
            }
        )

    @staticmethod
    def save_col_report(
        path,
        col_names,
        semantic_profiles,
        survival_labels=None,
    ):
        survival_labels = survival_labels or {}
        rows = []

        for col_name in col_names:
            rows.append(
                {
                    "col_name": col_name,
                    "inferred_datatype": (semantic_profiles[col_name].col_type.value),
                    "change_datatype": pd.NA,
                    "survival_label": survival_labels.get(col_name, pd.NA),
                    "event_structure": DEFAULT_EVENT_STRUCTURE,
                    "observation_scheme": DEFAULT_OBSERVATION_SCHEME,
                    "remove": pd.NA,
                }
            )

        report_columns = [
            "col_name",
            "inferred_datatype",
            "change_datatype",
            "survival_label",
            "event_structure",
            "observation_scheme",
            "remove",
        ]

        with pd.ExcelWriter(path=path, engine="openpyxl") as writer:
            pd.DataFrame(
                rows,
                columns=report_columns,
            ).to_excel(
                excel_writer=writer,
                index=False,
                sheet_name="SurvMeta",
            )

            SurvMetaReport._get_validation_df().to_excel(
                excel_writer=writer,
                index=False,
                sheet_name="__ValidationRanges__",
            )
            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

    @staticmethod
    def _add_validation_datatype(report_path):
        workbook = load_workbook(filename=report_path)

        try:
            worksheet = workbook["SurvMeta"]
            col_map = BaseExcel.get_worksheet_col_map(worksheet=worksheet)

            if worksheet.max_row < 2:
                return

            rules = {
                "change_datatype": (
                    f"=__ValidationRanges__!$A$2:" f"$A${len(SurvivalDataTypes) + 1}",
                    True,
                ),
                "remove": (
                    "=__ValidationRanges__!$C$2:$C$3",
                    True,
                ),
                "event_structure": (
                    f"=__ValidationRanges__!$D$2:" f"$D${len(EVENT_STRUCTURES) + 1}",
                    False,
                ),
                "observation_scheme": (
                    f"=__ValidationRanges__!$E$2:" f"$E${len(OBSERVATION_SCHEMES) + 1}",
                    False,
                ),
            }

            for field_name, (formula, allow_blank) in rules.items():
                validation = DataValidation(
                    type="list",
                    formula1=formula,
                    allow_blank=allow_blank,
                    showDropDown=False,
                    showInputMessage=False,
                    showErrorMessage=True,
                    errorStyle="stop",
                    errorTitle="Invalid survival metadata",
                    error="Select a value from the dropdown.",
                )
                worksheet.add_data_validation(validation)

                column_letter = col_map[field_name]
                validation.add(
                    f"{column_letter}2:" f"{column_letter}{worksheet.max_row}"
                )

            workbook.save(filename=report_path)
        finally:
            workbook.close()

    @staticmethod
    def _get_surv_edits(surv_meta_df):

        edits: dict[str, SurvEdit] = {}

        for _, row in surv_meta_df.iterrows():
            col_name = row["col_name"]

            change_datatype = None
            if pd.notna(row["change_datatype"]):
                change_datatype = SurvivalDataTypes(row["change_datatype"])

            remove = False
            if pd.notna(row["remove"]):
                remove = bool(row["remove"])

            if not (remove or change_datatype is not None):
                continue

            edits[col_name] = SurvEdit(
                col_name=col_name, remove=remove, change_datatype=change_datatype
            )

        return edits

    def get_surv_edit_schema(self, curated_meta_report):
        surv_meta_df = curated_meta_report.parse("SurvMeta")
        edits = self._get_surv_edits(surv_meta_df=surv_meta_df)

        return SurvEditSchema(edits=edits)

    @staticmethod
    def get_curated_surv_profiles(meta_edit_schema, surv_profiles):
        curated_profiles = dict(surv_profiles)
        for col_name, surv_edit in meta_edit_schema.edits.items():
            if surv_edit.remove and col_name in curated_profiles:
                del curated_profiles[col_name]
                continue

            profile = curated_profiles.get(col_name)
            if profile is not None and surv_edit.change_datatype is not None:
                curated_profiles[col_name] = replace(
                    profile,
                    col_type=surv_edit.change_datatype,
                )

        return curated_profiles

    @staticmethod
    def get_surv_pairs(surv_meta_df, surv_profiles):
        if not surv_meta_df.columns.is_unique:
            raise ValueError("SurvMeta contains duplicate column headers.")

        required_columns = {"col_name", "survival_label"}
        missing = required_columns - set(surv_meta_df.columns)
        if missing:
            raise ValueError(f"SurvMeta is missing columns: {sorted(missing)!r}.")

        has_declaration_fields(surv_meta_df.columns)

        grouped = defaultdict(list)
        seen_columns = set()

        for row in surv_meta_df.to_dict(orient="records"):
            col_name = row["col_name"]

            # Profiles already exclude columns removed by the edit schema.
            if col_name not in surv_profiles:
                continue

            if col_name in seen_columns:
                raise ValueError(f"Duplicate survival column row: {col_name!r}.")
            seen_columns.add(col_name)

            declaration = read_survival_declaration(row)
            surv_label = row["survival_label"]

            if is_blank_cell(surv_label):
                if declaration != (
                    DEFAULT_EVENT_STRUCTURE,
                    DEFAULT_OBSERVATION_SCHEME,
                ):
                    raise ValueError(
                        f"Column {col_name!r} declares a non-default "
                        "survival type but has no survival_label."
                    )
                continue

            if not isinstance(surv_label, str):
                raise ValueError("survival_label must be text.")

            surv_label = surv_label.strip()

            grouped[surv_label].append((surv_profiles[col_name], declaration))

        pairs = {}

        for surv_label, entries in grouped.items():
            if len(entries) != 2:
                raise ValueError(
                    f"Survival label {surv_label!r} must have exactly "
                    f"2 rows, found {len(entries)}."
                )

            declarations = {declaration for _, declaration in entries}
            if len(declarations) != 1:
                raise ValueError(
                    f"Survival label {surv_label!r} has inconsistent "
                    "event_structure or observation_scheme values. "
                    "Use the same choices on its Event and Time rows."
                )

            # Use the curated roles, including change_datatype decisions.
            profiles_by_type = {profile.col_type: profile for profile, _ in entries}
            required_types = {
                SurvivalDataTypes.EVENT,
                SurvivalDataTypes.TIME,
            }

            if set(profiles_by_type) != required_types:
                raise ValueError(
                    f"Survival label {surv_label!r} must contain "
                    "exactly one Event row and one Time row."
                )

            event_structure, observation_scheme = declarations.pop()

            pairs[surv_label] = SurvPair(
                surv_label=surv_label,
                event_profile=profiles_by_type[SurvivalDataTypes.EVENT],
                time_profile=profiles_by_type[SurvivalDataTypes.TIME],
                event_structure=event_structure,
                observation_scheme=observation_scheme,
            )

        return SurvPairs(pairs=pairs)

    @staticmethod
    def _get_surv_cat_meta_df(df, col_names, rename_mapping):
        """Build event-category rows without losing source scalar types."""

        if not df.columns.is_unique:
            raise ValueError("Event reports require unique source column names.")

        rows = []

        for col_name in dict.fromkeys(col_names):
            source_name = (
                rename_mapping.get(col_name, col_name) if rename_mapping else col_name
            )
            if source_name not in df.columns:
                raise KeyError(
                    f"Event column {col_name!r} resolves to missing source "
                    f"column {source_name!r}."
                )

            category_counts: Counter[str] = Counter()
            for value in df[source_name].array:
                if not is_missing_scalar(value):
                    identity = encode_category_scalar(value)
                    category_counts[identity] += 1

            total_non_missing = sum(category_counts.values())

            for identity in sorted(category_counts):
                category = decode_category_scalar(identity)
                count = category_counts[identity]

                rows.append(
                    {
                        "col_name": col_name,
                        "category": str(category),
                        "event_observed": "",
                        "remove": "",
                        "count": count,
                        "percentage": round(
                            100 * count / total_non_missing,
                            2,
                        ),
                        "category_encoding": identity,
                    }
                )

        return pd.DataFrame(
            rows,
            columns=[
                "col_name",
                "category",
                "event_observed",
                "remove",
                "count",
                "percentage",
                "category_encoding",
            ],
        )

    @staticmethod
    def save_cat_meta_report(df, rename_mapping, report_path, profiles_path):
        semantic_profiles = SurvMetaReport.load_semantic_profiles(path=profiles_path)
        event_col_names = [
            profile.col_name
            for profile in semantic_profiles.values()
            if profile.col_type == SurvivalDataTypes.EVENT
        ]

        report_df = SurvMetaReport._get_surv_cat_meta_df(
            df=df,
            col_names=event_col_names,
            rename_mapping=rename_mapping,
        )

        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            report_df.to_excel(
                excel_writer=writer,
                sheet_name="SurvCatMeta",
                index=False,
            )

            # Store source labels as literal Excel text, including strings
            # beginning with "=". The encoding preserves their actual type.
            worksheet = writer.sheets["SurvCatMeta"]
            for cells in worksheet.iter_rows(min_row=2):
                for column_index in (0, 1, 6):
                    cells[column_index].data_type = "s"

            validation_df = SurvMetaReport._get_validation_df()
            validation_df.to_excel(
                excel_writer=writer,
                sheet_name="__ValidationRanges__",
                index=False,
            )
            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

    @staticmethod
    def _add_surv_cat_validation(report_path):
        """Add event and removal dropdowns without input popups."""

        workbook = load_workbook(filename=report_path)

        try:
            worksheet = workbook["SurvCatMeta"]
            column_numbers = {cell.value: cell.column for cell in worksheet[1]}

            required_columns = {
                "category_encoding",
                "event_observed",
                "remove",
            }
            missing_columns = required_columns - set(column_numbers)
            if missing_columns:
                raise ValueError(
                    "SurvCatMeta is missing required columns: "
                    f"{sorted(missing_columns)!r}."
                )

            encoding_letter = get_column_letter(column_numbers["category_encoding"])
            worksheet.column_dimensions[encoding_letter].hidden = True

            instructions = {
                "event_observed": (
                    "True = event observed; False = right-censored. "
                    "Blank does not request removal."
                ),
                "remove": (
                    "True = convert this category to missing. "
                    "Leave event_observed blank when removing a category."
                ),
            }

            if worksheet.max_row >= 2:
                for field_name, instruction in instructions.items():
                    column_letter = get_column_letter(column_numbers[field_name])

                    validation = DataValidation(
                        type="list",
                        formula1='"True,False"',
                        allow_blank=True,
                        showDropDown=False,
                        showInputMessage=False,
                        showErrorMessage=True,
                        errorStyle="stop",
                        errorTitle="Invalid Boolean instruction",
                        error=("Select True, False, or leave the cell blank."),
                        promptTitle=field_name,
                        prompt=instruction,
                    )

                    worksheet.add_data_validation(validation)
                    validation.add(
                        f"{column_letter}2:" f"{column_letter}{worksheet.max_row}"
                    )

            workbook.save(filename=report_path)
        finally:
            workbook.close()

    def create_cat_meta_report(self, df, rename_mapping, profiles_path, report_path):
        self.save_cat_meta_report(
            df=df,
            rename_mapping=rename_mapping,
            profiles_path=profiles_path,
            report_path=report_path,
        )

        BaseExcel.format_cell_length(path=report_path)
        self._add_surv_cat_validation(report_path=report_path)
        BaseExcel.protect_cols(
            path=report_path,
            protected_col_names=[
                "col_name",
                "category",
                "count",
                "percentage",
                "category_encoding",
            ],
            password="statomix",
            lock=True,
        )

    @staticmethod
    def _get_surv_cat_edits(surv_cat_meta_df):
        """Validate complete mappings before retaining actionable rows."""

        frame = surv_cat_meta_df.copy()

        # Retain compatibility with callers supplying an old MultiIndex report.
        if isinstance(frame.index, pd.MultiIndex):
            if frame.index.nlevels != 2:
                raise ValueError(
                    "A survival-category report index must have two levels."
                )
            frame = frame.reset_index()

        if not frame.columns.is_unique:
            raise ValueError("SurvCatMeta contains duplicate column headers.")

        required_columns = {"col_name", "category", "remove"}
        missing_columns = required_columns - set(frame.columns)
        if missing_columns:
            raise ValueError(
                f"SurvCatMeta is missing columns: {sorted(missing_columns)!r}."
            )

        if "event_observed" in frame.columns:
            event_field = "event_observed"
        elif "rename_to" in frame.columns:
            event_field = "rename_to"
        else:
            raise ValueError("SurvCatMeta requires event_observed or legacy rename_to.")

        has_category_encoding = "category_encoding" in frame.columns

        if not has_category_encoding:
            # Legacy workbooks merged repeated column-name cells.
            frame["col_name"] = (
                frame["col_name"]
                .map(lambda value: pd.NA if is_blank_cell(value) else value)
                .ffill()
            )

        all_rows: dict[
            str,
            dict[str, tuple[object, bool | None, bool]],
        ] = defaultdict(dict)

        for row in frame.to_dict(orient="records"):
            if all(is_blank_cell(value) for value in row.values()):
                continue

            col_name = row["col_name"]
            if not isinstance(col_name, str) or not col_name.strip():
                raise ValueError("SurvCatMeta contains a missing column name.")

            if has_category_encoding:
                stored_encoding = row["category_encoding"]
                if is_blank_cell(stored_encoding):
                    raise ValueError(
                        f"Missing protected category_encoding in {col_name!r}."
                    )

                category = decode_category_scalar(stored_encoding)

                # Compare display text, not inferred Excel scalar types.
                if str(row["category"]) != str(category):
                    raise ValueError(
                        f"Protected category label changed in {col_name!r}: "
                        f"expected {str(category)!r}, "
                        f"found {row['category']!r}."
                    )
            else:
                category = row["category"]

            identity = encode_category_scalar(category)
            event_observed = parse_optional_event_observed(
                row[event_field],
                col_name=col_name,
                category=category,
            )
            remove = parse_remove_instruction(
                row["remove"],
                col_name=col_name,
                category=category,
            )

            if remove and event_observed is not None:
                raise ValueError(
                    f"Category {category!r} in {col_name!r} cannot "
                    "be both mapped and removed."
                )

            if identity in all_rows[col_name]:
                raise ValueError(
                    f"Duplicate category row in {col_name!r}: {category!r}."
                )

            all_rows[col_name][identity] = (
                category,
                event_observed,
                remove,
            )

        edits: dict[str, dict[str, SurvCatEdit]] = defaultdict(dict)

        for col_name, column_rows in all_rows.items():
            retained_rows = [row for row in column_rows.values() if not row[2]]
            has_mapping = any(
                event_observed is not None for _, event_observed, _ in retained_rows
            )

            if has_mapping:
                unmapped_categories = [
                    category
                    for category, event_observed, _ in retained_rows
                    if event_observed is None
                ]
                if unmapped_categories:
                    raise ValueError(
                        f"Survival event column {col_name!r} has a partial "
                        f"mapping. Unmapped categories: {unmapped_categories!r}."
                    )
            else:
                noncanonical_categories = [
                    category
                    for category, _, _ in retained_rows
                    if not is_canonical_event_value(category)
                ]
                if noncanonical_categories:
                    raise ValueError(
                        f"Survival event column {col_name!r} has labels "
                        "without event mappings: "
                        f"{noncanonical_categories!r}."
                    )

            for identity, row in column_rows.items():
                category, event_observed, remove = row
                if not remove and event_observed is None:
                    continue

                edits[col_name][identity] = SurvCatEdit(
                    col_name=col_name,
                    category=category,
                    category_encoding=identity,
                    event_observed=event_observed,
                    remove=remove,
                )

        return dict(edits)

    def get_surv_cat_meta_edit_schema(self, curated_meta_report):
        # dtype=object preserves Excel cell values; keep_default_na=False
        # prevents literal labels such as "NA" from becoming missing.
        report_df = curated_meta_report.parse(
            sheet_name="SurvCatMeta",
            dtype=object,
            keep_default_na=False,
        )
        edits = self._get_surv_cat_edits(report_df)
        return SurvCatMetaEditSchema(cat_edits=edits)
