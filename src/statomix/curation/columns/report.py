"""Column-profile artifacts and human-editable Excel reports."""

from collections import defaultdict
from collections.abc import Mapping
from dataclasses import dataclass, replace
from pathlib import Path
from types import MappingProxyType
from typing import ClassVar

import pandas as pd
from fileverse.formats.excel import BaseExcel
from openpyxl import load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.core.tabular import frame_from_rows
from statomix.logging import get_logger

from .audit import ColumnAudit, ColumnValueFrequency
from .profiler import ColProfile, ColProfiler
from .semantic_rules import DataTypes

logger = get_logger(name="col_report")

COLUMN_REPORT_FIELDS = (
    "col_name",
    "change_col_name",
    "inferred_datatype",
    "change_datatype",
    "remove",
    "source_dtype",
    "missing_n",
    "missing_pct",
    "unique_n",
    "num_conversion_pct",
    "numeric_n",
    "nonnumeric_n",
    "minimum",
    "q1",
    "median",
    "q3",
    "maximum",
)

AUDIT_FIELD_NAMES = (
    "source_dtype",
    "missing_n",
    "missing_pct",
    "unique_n",
    "num_conversion_pct",
    "numeric_n",
    "nonnumeric_n",
    "minimum",
    "q1",
    "median",
    "q3",
    "maximum",
)

COMMON_SHEET_CELL_MAP = {
    field_name: get_column_letter(index)
    for index, field_name in enumerate(
        COLUMN_REPORT_FIELDS,
        start=1,
    )
}

SHEET_CELL_MAP = {datatype.value: dict(COMMON_SHEET_CELL_MAP) for datatype in DataTypes}

EDITABLE_DATATYPE_SHEETS = frozenset(datatype.value for datatype in DataTypes)

READ_ONLY_AUDIT_SHEETS = frozenset(
    {
        "Value Counts",
        "Report Metadata",
    }
)

HIDDEN_SUPPORT_SHEETS = frozenset(
    {
        "__ValidationRanges__",
    }
)

PROTECTED_COL_NAMES = (
    "col_name",
    "inferred_datatype",
    *AUDIT_FIELD_NAMES,
)

# SHEET_CELL_MAP = {
#     DataTypes.IDENTIFIER.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#     },
#     DataTypes.NUMERICAL.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         # "units": 'C',
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#     },
#     DataTypes.CATEGORICAL.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#     },
#     DataTypes.SURVIVAL.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         # "units": 'C',
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#     },
#     DataTypes.DATETIME.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#         # "format": 'F',
#     },
#     DataTypes.FREE_TEXT.value: {
#         "col_name": "A",
#         "change_col_name": "B",
#         "inferred_datatype": "C",
#         "change_datatype": "D",
#         "remove": "E",
#     },
# }

# EDITABLE_COL_NAMES = {
#     "change_col_name",
#     "change_datatype",
#     "remove",
#     "format",
# }

# PROTECTED_COL_NAMES = [
#     "col_name",
#     "inferred_datatype",
# ]


@dataclass(frozen=True, slots=True, kw_only=True)
class ColEdit:
    """
    User modifications for a single column.
    """

    col_name: str
    remove: bool = False
    change_col_name: str | None = None
    change_datatype: DataTypes | None = None

    def to_dict(self) -> dict:
        """Converts the instance to a dictionary for saving."""
        return {
            "col_name": self.col_name,
            "remove": self.remove,
            "change_col_name": self.change_col_name,
            # We store the value of the Enum so it can be saved as a string/primitive
            "change_datatype": (
                self.change_datatype.value if self.change_datatype else None
            ),
        }

    @staticmethod
    def from_dict(data: dict) -> "ColEdit":
        """Creates an instance from a dictionary loaded from a file."""
        return ColEdit(
            col_name=data["col_name"],
            remove=bool(data.get("remove", False)),
            change_col_name=(
                data.get("change_col_name")
                if pd.notna(data.get("change_col_name"))
                else None
            ),
            # We convert the string value back into the DataTypes Enum
            change_datatype=(
                DataTypes(data["change_datatype"])
                if pd.notna(data.get("change_datatype"))
                else None
            ),
        )


@dataclass(frozen=True, slots=True, kw_only=True)
class ColEditSchema:
    """
    Contains only columns that have been modified by the user.
    """

    edits: Mapping[str, ColEdit]

    def __post_init__(self) -> None:
        object.__setattr__(self, "edits", MappingProxyType(dict(self.edits)))

    PARQUET_SCHEMA: ClassVar[dict[str, str]] = {
        "col_name": "object",
        "remove": "boolean",
        "change_col_name": "object",
        "change_datatype": "object",
    }

    @classmethod
    def empty(cls) -> "ColEditSchema":
        return cls(edits={})

    @property
    def is_empty(self) -> bool:
        return not self.edits

    def save(self, path: Path) -> None:
        """Save this edit schema as a parquet artifact."""
        rows = [edit.to_dict() for edit in self.edits.values()]

        df = frame_from_rows(
            rows=rows,
            schema=self.PARQUET_SCHEMA,
        )
        df.to_parquet(path=path, index=False)

    @classmethod
    def load(cls, path: Path) -> "ColEditSchema":
        """Load a parquet artifact and reconstruct the edit schema."""
        df = pd.read_parquet(path=path)

        edits: dict[str, ColEdit] = {}
        for _, row in df.iterrows():
            edit = ColEdit.from_dict(row.to_dict())
            edits[edit.col_name] = edit

        return cls(edits=edits)


class ColReport:
    def __init__(self):

        self.col_profiler = ColProfiler(cat_unique_thresh=4, num_conversion_thresh=95)

    def create_col_report(
        self,
        df: pd.DataFrame,
        report_path: Path,
        profiles_path: Path,
        replace: bool,
        rename_mapping=None,
        audit_profiles_path: Path | None = None,
        value_frequencies_path: Path | None = None,
        value_count_unique_threshold: int = 30,
        report_metadata: Mapping[str, object] | None = None,
    ) -> None:
        """Create the integrated editable and diagnostic report."""

        if report_path.suffix.lower() != ".xlsx":
            raise ValueError("report_path must have an .xlsx suffix")

        if (audit_profiles_path is None) != (value_frequencies_path is None):
            raise ValueError(
                "audit_profiles_path and "
                "value_frequencies_path must either both be "
                "provided or both be omitted"
            )

        if report_path.exists() and not replace:
            logger.warning(
                "Column report already exists at:\n"
                f"{report_path}\n"
                "Set replace=True to replace."
            )
            return

        col_profiles = self.load_col_profiles(path=profiles_path)

        column_audit = ColumnAudit.from_dataframe(
            df=df,
            col_profiles=col_profiles,
            value_count_unique_threshold=(value_count_unique_threshold),
        )

        if audit_profiles_path is not None and value_frequencies_path is not None:
            column_audit.save(
                profiles_path=audit_profiles_path,
                value_frequencies_path=(value_frequencies_path),
            )

        self._save_col_report(
            path=report_path,
            col_profiles=col_profiles,
            column_audit=column_audit,
            rename_mapping=rename_mapping,
            report_metadata=report_metadata,
        )

        BaseExcel.format_cell_length(path=report_path)
        self._add_validation_datatype(report_path=report_path)

        for datatype in DataTypes:
            self._add_validation_categories(
                df=df,
                report_path=report_path,
                datatype=datatype.value,
                rename_mapping=rename_mapping,
            )

        self._protect_cols(
            report_path=report_path,
            password="statomix",
            lock=True,
        )

    def create_col_profiles(self, df, path, replace):
        if path.exists() and not replace:
            logger.warning(
                f"Column profiles exists at:\n{path}\nSet replace=True to replace."
            )
            return

        col_profiles: dict[str, ColProfile] = {}
        for col_name in df.columns:
            col_series = df[col_name]

            col_profiles[col_name] = self.col_profiler.get_col_profile(
                col_name=col_name,
                col_series=col_series,
            )

            # col_profiles[col_name] = col_profile

        # rows = [profile.to_dict() for profile in col_profiles.values()]
        # pd.DataFrame(rows).to_parquet(profiles_path)
        self.save_col_profiles(col_profiles=col_profiles, path=path)

    def save_col_profiles(self, col_profiles, path):
        rows = [profile.to_dict() for profile in col_profiles.values()]
        profiles_df = frame_from_rows(
            rows=rows,
            schema=ColProfile.PARQUET_SCHEMA,
        )
        profiles_df.to_parquet(path=path, index=False)

    @staticmethod
    def get_curated_col_profiles(col_profiles, col_edit_schema: ColEditSchema):
        curated_profiles = dict(col_profiles)
        for col_name, col_edit in col_edit_schema.edits.items():
            if col_edit.remove and col_name in curated_profiles:
                del curated_profiles[col_name]
                continue

            profile = curated_profiles.get(col_name)
            if profile is None:
                continue
            if col_edit.change_col_name is not None:
                profile = replace(
                    profile,
                    col_name=col_edit.change_col_name,
                )

            if col_edit.change_datatype is not None:
                profile = replace(
                    profile,
                    col_type=col_edit.change_datatype,
                )
            curated_profiles[col_name] = profile

        return curated_profiles

    @staticmethod
    def load_col_profiles(
        path: Path,
    ) -> dict[str, ColProfile]:

        df = pd.read_parquet(path)

        col_profiles: dict[str, ColProfile] = {}

        for _, row in df.iterrows():
            profile = ColProfile.from_dict(row)
            col_profiles[profile.col_name] = profile

        return col_profiles

    # def _format_cell_length(self, report_path):
    #     workbook = load_workbook(filename=report_path)

    #     # Create a reusable alignment object (Memory efficient!)
    #     center_align = Alignment(horizontal="center", vertical="center")

    #     for worksheet in workbook.worksheets:
    #         # 1. Bold the header row
    #         for cell in worksheet[1]:
    #             cell.font = Font(bold=True)

    #         # 2. Loop through every column exactly once
    #         for column_cells in worksheet.columns:
    #             max_length = 0

    #             # Look at every single cell in this specific column
    #             for cell in column_cells:
    #                 # Apply the centering alignment
    #                 cell.alignment = center_align

    #                 # Calculate the text length for the width adjuster
    #                 if cell.value is not None:
    #                     max_length = max(max_length, len(str(cell.value)))

    #             # Apply the final adjusted width to the column letter
    #             adjusted_width = max_length + 4
    #             column_letter = get_column_letter(column_cells[0].column)
    #             worksheet.column_dimensions[column_letter].width = adjusted_width

    #     workbook.save(filename=report_path)

    def _get_validation_df(self):
        dropdown_options = [datatype.value for datatype in DataTypes]

        max_len = max(len(dropdown_options), 2)
        datatypes = dropdown_options + [""] * (max_len - len(dropdown_options))
        boolean = ["True", "False"] + [""] * (max_len - 2)

        validation_df = pd.DataFrame(data={"DataTypes": datatypes, "Booleans": boolean})

        return validation_df

    def _save_col_report(
        self,
        *,
        path: Path,
        col_profiles,
        column_audit: ColumnAudit,
        rename_mapping=None,
        report_metadata: Mapping[str, object] | None = None,
    ) -> None:
        """Render editable and diagnostic worksheets."""

        datatype_map = defaultdict(list)

        for profile in col_profiles.values():
            datatype_map[profile.col_type].append(profile.col_name)

        with pd.ExcelWriter(
            path=path,
            engine="openpyxl",
        ) as writer:
            for datatype, col_names in datatype_map.items():
                if datatype is None or not col_names:
                    continue

                rows = []

                for col_name in col_names:
                    profile = col_profiles[col_name]
                    audit_profile = column_audit.profiles[col_name]
                    audit_record = audit_profile.to_dict()

                    row_data = {field_name: "" for field_name in COLUMN_REPORT_FIELDS}

                    row_data["col_name"] = profile.col_name
                    row_data["inferred_datatype"] = profile.col_type.value

                    for field_name in AUDIT_FIELD_NAMES:
                        field_value = audit_record[field_name]
                        row_data[field_name] = (
                            "" if field_value is None else field_value
                        )

                    rows.append(row_data)

                datatype_df = pd.DataFrame(
                    data=rows,
                    columns=COLUMN_REPORT_FIELDS,
                )

                datatype_df.to_excel(
                    excel_writer=writer,
                    sheet_name=datatype.value[:31],
                    index=False,
                )

            frequency_rows = [
                frequency.to_dict() for frequency in column_audit.value_frequencies
            ]

            frequencies_df = frame_from_rows(
                rows=frequency_rows,
                schema=(ColumnValueFrequency.PARQUET_SCHEMA),
            )

            frequencies_df.to_excel(
                excel_writer=writer,
                sheet_name="Value Counts",
                index=False,
            )

            metadata = {
                "report_schema_version": 1,
                "source_column_count": len(column_audit.profiles),
                "value_count_unique_threshold": (
                    column_audit.value_count_unique_threshold
                ),
                "value_frequency_rule": (
                    "All inferred categorical columns and "
                    "all columns with unique_n less than or "
                    "equal to the configured threshold"
                ),
            }

            if report_metadata is not None:
                metadata.update(report_metadata)

            metadata_rows = [
                {
                    "property": property_name,
                    "value": ("" if property_value is None else str(property_value)),
                }
                for property_name, property_value in metadata.items()
            ]

            metadata_df = pd.DataFrame(
                data=metadata_rows,
                columns=["property", "value"],
            )

            metadata_df.to_excel(
                excel_writer=writer,
                sheet_name="Report Metadata",
                index=False,
            )

            validation_df = self._get_validation_df()

            validation_df.to_excel(
                excel_writer=writer,
                sheet_name="__ValidationRanges__",
                index=False,
            )

            writer.sheets["__ValidationRanges__"].sheet_state = "veryHidden"

    def _add_validation_datatype(self, report_path):
        workbook = load_workbook(filename=report_path)
        total_datatypes = len(DataTypes)

        for worksheet in workbook.worksheets:
            if worksheet.title not in EDITABLE_DATATYPE_SHEETS:
                continue
            max_row = worksheet.max_row
            if max_row < 2:
                continue
            validation_datatype = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!$A$2:$A${total_datatypes + 1}",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Datatype",
                error="You must select a valid datatype from the provided drop-down menu.",
            )

            worksheet.add_data_validation(validation_datatype)

            cell_coordinate_datatype = SHEET_CELL_MAP[worksheet.title][
                "change_datatype"
            ]
            validation_datatype.add(
                f"{cell_coordinate_datatype}2:{cell_coordinate_datatype}{max_row}"
            )

            validation_boolean = DataValidation(
                type="list",
                formula1="=__ValidationRanges__!$B$2:$B$3",
                allow_blank=True,
                showErrorMessage=True,
                errorStyle="stop",
                errorTitle="Invalid Input",
                error="Please choose only 'True' or 'False' from the dropdown list.",
            )
            worksheet.add_data_validation(validation_boolean)
            cell_coordinate_boolean = SHEET_CELL_MAP[worksheet.title]["remove"]
            validation_boolean.add(
                f"{cell_coordinate_boolean}2:{cell_coordinate_boolean}{max_row}"
            )

        workbook.save(filename=report_path)

    def _add_validation_categories(
        self, df, report_path, datatype, rename_mapping=None
    ):

        # mapping = rename_mapping or {}
        workbook = load_workbook(filename=report_path)

        if datatype not in workbook.sheetnames:
            return

        worksheet = workbook[datatype]
        val_sheet = workbook["__ValidationRanges__"]

        for cell in worksheet["A"][1:]:
            if not cell.value:
                continue

            # target_name = mapping.get(cell.value, cell.value)
            target_name = (
                rename_mapping.get(cell.value, cell.value)
                if rename_mapping
                else cell.value
            )
            categories = list(df[target_name].dropna().unique())
            categories.sort()

            # if len(categories) >= 20 or len(categories) == 0:
            categories = categories[:20]

            # Dynamically find the next available column on the validation sheet
            next_col_idx = val_sheet.max_column + 1
            next_col_letter = get_column_letter(next_col_idx)

            # Write the category options down this new column
            val_sheet.cell(row=1, column=next_col_idx, value=f"{cell.value}_cats")
            for idx, cat_val in enumerate(iterable=categories, start=2):
                val_sheet.cell(row=idx, column=next_col_idx, value=str(cat_val))

            # Informational reference dropdown on Column A (No error restriction popup)
            validation_category = DataValidation(
                type="list",
                formula1=f"=__ValidationRanges__!${next_col_letter}$2:${next_col_letter}${len(categories) + 1}",
                allow_blank=False,
                showErrorMessage=False,
            )

            worksheet.add_data_validation(validation_category)
            validation_category.add(cell.coordinate)

        workbook.save(filename=report_path)

    def _protect_cols(
        self,
        report_path: Path,
        lock: bool,
        password: str | None = None,
    ) -> None:
        """Protect identities, diagnostics, and audit worksheets."""

        workbook = load_workbook(filename=report_path)

        for worksheet in workbook.worksheets:
            sheet_name = worksheet.title

            if sheet_name in HIDDEN_SUPPORT_SHEETS:
                continue

            if sheet_name in READ_ONLY_AUDIT_SHEETS:
                for row in worksheet.iter_rows():
                    for cell in row:
                        cell.protection = Protection(locked=True)

                worksheet.auto_filter.ref = worksheet.dimensions
                worksheet.protection.sheet = lock
                worksheet.protection.autoFilter = True

                if password:
                    worksheet.protection.password = password

                continue

            if sheet_name not in EDITABLE_DATATYPE_SHEETS:
                raise RuntimeError(
                    "Unrecognized worksheet in column " f"report: {sheet_name!r}"
                )

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.protection = Protection(locked=False)

            for col_header in PROTECTED_COL_NAMES:
                col_letter = SHEET_CELL_MAP[sheet_name][col_header]

                for cell in worksheet[col_letter]:
                    cell.protection = Protection(locked=True)

            worksheet.auto_filter.ref = worksheet.dimensions
            worksheet.protection.sheet = lock
            worksheet.protection.autoFilter = True

            if password:
                worksheet.protection.password = password

            worksheet.protection.enableSelection = "unlockedCells"

        workbook.save(filename=report_path)

    def get_col_edit_schema(self, curated_col_report):
        rename_mapping = {}
        edits: dict[str, ColEdit] = {}

        for sheet_name in curated_col_report.sheet_names:
            if sheet_name not in EDITABLE_DATATYPE_SHEETS:
                continue

            datatype_df = curated_col_report.parse(sheet_name=sheet_name)

            for _, row in datatype_df.iterrows():
                col_name = row["col_name"]

                remove = False
                if pd.notna(row["remove"]):
                    remove = bool(row["remove"])

                change_col_name = None
                if pd.notna(row["change_col_name"]):
                    new_name = str(row["change_col_name"]).strip()
                    if new_name:
                        change_col_name = new_name
                        rename_mapping[new_name] = col_name

                change_datatype = None
                if pd.notna(row["change_datatype"]):
                    change_datatype = DataTypes(row["change_datatype"])

                if not (
                    remove or change_col_name is not None or change_datatype is not None
                ):
                    continue

                edits[col_name] = ColEdit(
                    col_name=col_name,
                    remove=remove,
                    change_col_name=change_col_name,
                    change_datatype=change_datatype,
                )

        return rename_mapping, ColEditSchema(edits=edits)
