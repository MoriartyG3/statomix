"""Format-aware comparison of two Statomix artifact trees."""

from __future__ import annotations

import argparse
import hashlib
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any


@dataclass(frozen=True, slots=True, kw_only=True)
class Difference:
    """One reproducible difference between reference and candidate output."""

    path: Path
    kind: str
    message: str


@dataclass(frozen=True, slots=True, kw_only=True)
class ComparisonReport:
    """Complete result of an artifact-tree comparison."""

    reference: Path
    candidate: Path
    compared_files: int
    differences: tuple[Difference, ...]

    @property
    def matches(self) -> bool:
        return not self.differences


def _sha256(*, path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _inventory(*, root: Path) -> set[Path]:
    return {path.relative_to(root) for path in root.rglob("*") if path.is_file()}


def _compare_parquet(*, reference: Path, candidate: Path) -> str | None:
    import pandas as pd
    from pandas.testing import assert_frame_equal

    left = pd.read_parquet(path=reference)
    right = pd.read_parquet(path=candidate)
    try:
        assert_frame_equal(
            left=left,
            right=right,
            check_dtype=True,
            check_exact=True,
            check_like=False,
            check_names=True,
        )
    except AssertionError as error:
        return str(error)
    return None


def _color_snapshot(color: Any) -> tuple[Any, ...] | None:
    if color is None:
        return None
    return (
        color.type,
        color.rgb,
        color.indexed,
        color.auto,
        color.theme,
        color.tint,
    )


def _cell_snapshot(cell: Any) -> tuple[Any, ...]:
    hyperlink = cell.hyperlink.target if cell.hyperlink is not None else None
    comment = (
        (cell.comment.text, cell.comment.author) if cell.comment is not None else None
    )
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    protection = cell.protection
    return (
        cell.coordinate,
        cell.value,
        cell.data_type,
        cell.number_format,
        hyperlink,
        comment,
        (
            font.name,
            font.sz,
            font.bold,
            font.italic,
            font.vertAlign,
            font.underline,
            font.strike,
            _color_snapshot(font.color),
        ),
        (
            fill.fill_type,
            _color_snapshot(fill.fgColor),
            _color_snapshot(fill.bgColor),
        ),
        str(cell.border),
        (
            alignment.horizontal,
            alignment.vertical,
            alignment.text_rotation,
            alignment.wrap_text,
            alignment.shrink_to_fit,
            alignment.indent,
        ),
        (protection.locked, protection.hidden),
    )


def _worksheet_snapshot(worksheet: Any) -> tuple[Any, ...]:
    cells = tuple(
        _cell_snapshot(cell)
        for row in worksheet.iter_rows()
        for cell in row
        if cell.value is not None or cell.has_style or cell.hyperlink or cell.comment
    )
    validations = tuple(
        sorted(
            (
                validation.type,
                validation.formula1,
                validation.formula2,
                str(validation.sqref),
                validation.allow_blank,
                validation.errorStyle,
                validation.errorTitle,
                validation.error,
            )
            for validation in worksheet.data_validations.dataValidation
        )
    )
    columns = tuple(
        sorted(
            (key, value.width, value.hidden)
            for key, value in worksheet.column_dimensions.items()
        )
    )
    rows = tuple(
        sorted(
            (key, value.height, value.hidden)
            for key, value in worksheet.row_dimensions.items()
        )
    )
    return (
        worksheet.title,
        worksheet.sheet_state,
        str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
        tuple(sorted(str(value) for value in worksheet.merged_cells.ranges)),
        cells,
        validations,
        columns,
        rows,
    )


def _compare_excel(*, reference: Path, candidate: Path) -> str | None:
    from openpyxl import load_workbook

    left = load_workbook(filename=reference, data_only=False)
    right = load_workbook(filename=candidate, data_only=False)
    if left.sheetnames != right.sheetnames:
        return f"sheet order differs: {left.sheetnames!r} != {right.sheetnames!r}"

    left_names = tuple(sorted(str(value) for value in left.defined_names.values()))
    right_names = tuple(sorted(str(value) for value in right.defined_names.values()))
    if left_names != right_names:
        return "defined names differ"

    for sheet_name in left.sheetnames:
        if _worksheet_snapshot(left[sheet_name]) != _worksheet_snapshot(
            right[sheet_name]
        ):
            return f"worksheet content or formatting differs: {sheet_name}"
    return None


def _compare_png(*, reference: Path, candidate: Path) -> str | None:
    import numpy as np
    from PIL import Image

    with Image.open(reference) as left_image, Image.open(candidate) as right_image:
        left = np.asarray(left_image.convert("RGBA"))
        right = np.asarray(right_image.convert("RGBA"))
    if left.shape != right.shape:
        return f"decoded image shapes differ: {left.shape} != {right.shape}"
    if not np.array_equal(left, right):
        differing = int(np.count_nonzero(left != right))
        return f"decoded pixels differ in {differing} channel values"
    return None


def _compare_json(*, reference: Path, candidate: Path) -> str | None:
    with reference.open(encoding="utf-8") as source:
        left = json.load(source)
    with candidate.open(encoding="utf-8") as source:
        right = json.load(source)
    return None if left == right else "JSON values differ"


def _compare_yaml(*, reference: Path, candidate: Path) -> str | None:
    try:
        import yaml
    except ModuleNotFoundError:
        return (
            None
            if _sha256(path=reference) == _sha256(path=candidate)
            else ("YAML bytes differ")
        )

    with reference.open(encoding="utf-8") as source:
        left = yaml.safe_load(source)
    with candidate.open(encoding="utf-8") as source:
        right = yaml.safe_load(source)
    return None if left == right else "YAML values differ"


def _compare_file(*, reference: Path, candidate: Path) -> str | None:
    suffix = reference.suffix.lower()
    if suffix == ".parquet":
        return _compare_parquet(reference=reference, candidate=candidate)
    if suffix == ".xlsx":
        return _compare_excel(reference=reference, candidate=candidate)
    if suffix == ".png":
        return _compare_png(reference=reference, candidate=candidate)
    if suffix == ".json":
        return _compare_json(reference=reference, candidate=candidate)
    if suffix in {".yaml", ".yml"}:
        return _compare_yaml(reference=reference, candidate=candidate)
    if _sha256(path=reference) != _sha256(path=candidate):
        return "file bytes differ"
    return None


def compare_artifact_trees(
    *,
    reference: Path,
    candidate: Path,
) -> ComparisonReport:
    """Compare complete artifact inventories and their format-aware content."""

    reference_root = Path(reference).resolve()
    candidate_root = Path(candidate).resolve()
    if not reference_root.is_dir():
        raise NotADirectoryError(reference_root)
    if not candidate_root.is_dir():
        raise NotADirectoryError(candidate_root)

    reference_files = _inventory(root=reference_root)
    candidate_files = _inventory(root=candidate_root)
    differences: list[Difference] = []

    for path in sorted(reference_files - candidate_files):
        differences.append(
            Difference(path=path, kind="missing", message="missing from candidate")
        )
    for path in sorted(candidate_files - reference_files):
        differences.append(
            Difference(path=path, kind="extra", message="only present in candidate")
        )

    shared = sorted(reference_files & candidate_files)
    for path in shared:
        try:
            message = _compare_file(
                reference=reference_root / path,
                candidate=candidate_root / path,
            )
        except Exception as error:
            message = f"comparison failed with {type(error).__name__}: {error}"
        if message is not None:
            differences.append(Difference(path=path, kind="content", message=message))

    return ComparisonReport(
        reference=reference_root,
        candidate=candidate_root,
        compared_files=len(shared),
        differences=tuple(differences),
    )


def main(argv: list[str] | None = None) -> int:
    """Run the artifact comparator as the ``statomix-compare`` command."""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("reference", type=Path)
    parser.add_argument("candidate", type=Path)
    arguments = parser.parse_args(argv)

    report = compare_artifact_trees(
        reference=arguments.reference,
        candidate=arguments.candidate,
    )
    if report.matches:
        print(f"MATCH: compared {report.compared_files} artifact(s).")
        return 0

    print(
        f"DIFFERENT: {len(report.differences)} difference(s) across "
        f"{report.compared_files} shared artifact(s)."
    )
    for difference in report.differences:
        print(f"- {difference.path}: {difference.kind}: {difference.message}")
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
