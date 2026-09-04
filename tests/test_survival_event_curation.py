from __future__ import annotations

from dataclasses import replace

import numpy as np
import pandas as pd
import pytest
from fileverse.formats.yaml import BaseYAML
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal, assert_series_equal

from statomix import Project
from statomix.analytics.datatypes.survival.data import prepare_survival_data
from statomix.curation.categorical import CatMetaEditSchema
from statomix.curation.columns import ColEditSchema, ColProfiler, DataTypes
from statomix.curation.inheritance import apply_inherited_category_edits
from statomix.curation.service import apply_curation_schemas
from statomix.curation.survival.events import (
    encode_category_scalar,
    normalize_survival_event_series,
    parse_optional_event_observed,
)
from statomix.curation.survival.profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
)
from statomix.curation.survival.report import (
    SurvCatEdit,
    SurvCatMetaEditSchema,
    SurvMetaReport,
    SurvPair,
    SurvPairs,
)
from statomix.storage.parquet_metadata import write_dataframe_with_category_ranks


def _report_frame(categories, mappings, removals, col_name="OS event"):
    return pd.DataFrame(
        {
            "col_name": [col_name] * len(categories),
            "category": [str(value) for value in categories],
            "event_observed": mappings,
            "remove": removals,
            "category_encoding": [
                encode_category_scalar(value) for value in categories
            ],
        }
    )


def _schema(categories, mappings, removals, col_name="OS event"):
    report_df = _report_frame(
        categories,
        mappings,
        removals,
        col_name=col_name,
    )
    return SurvCatMetaEditSchema(
        cat_edits=SurvMetaReport._get_surv_cat_edits(report_df)
    )


def _survival_profile(col_name, col_type):
    return SurvivalSemanticProfile(
        col_name=col_name,
        col_type=col_type,
        score=1.0,
        tokens=(),
        normalized_name=col_name.casefold(),
        matched_rules=(),
        all_scores={col_type: 1.0},
    )


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (True, True),
        (False, False),
        (np.bool_(True), True),
        (1, True),
        (0, False),
        (1.0, True),
        (0.0, False),
        ("True", True),
        (" false ", False),
        ("1", True),
        ("0", False),
        ("", None),
        ("   ", None),
        (None, None),
        (pd.NA, None),
        (float("nan"), None),
    ],
)
def test_instruction_parser_returns_bool_or_none(value, expected):
    result = parse_optional_event_observed(
        value,
        col_name="OS event",
        category="source category",
    )
    assert result is expected


@pytest.mark.parametrize(
    "value",
    [2, -1, 0.5, float("inf"), "Yes", "No", "1.0", object()],
)
def test_instruction_parser_rejects_ambiguous_input(value):
    with pytest.raises(ValueError, match="must be True/False"):
        parse_optional_event_observed(
            value,
            col_name="OS event",
            category="source category",
        )


def test_category_identity_distinguishes_text_number_and_boolean():
    assert encode_category_scalar(1) == encode_category_scalar(1.0)
    assert encode_category_scalar(1) != encode_category_scalar("1")
    assert encode_category_scalar(1) != encode_category_scalar(True)


def test_partial_report_mapping_is_rejected():
    with pytest.raises(ValueError, match="partial mapping"):
        _schema(
            ["Alive", "Dead"],
            [False, None],
            [False, False],
        )


def test_mapping_and_removal_cannot_both_be_set():
    with pytest.raises(ValueError, match="both mapped and removed"):
        _schema(["Unknown"], [False], [True])


def test_duplicate_report_rows_are_rejected():
    with pytest.raises(ValueError, match="Duplicate category row"):
        _schema(
            ["Dead", "Dead"],
            [True, True],
            [False, False],
        )


def test_canonical_numeric_categories_can_remain_unmapped():
    schema = _schema([0, 1], [None, None], [False, False])
    assert schema.is_empty


def test_text_zero_one_requires_explicit_mapping():
    with pytest.raises(ValueError, match="labels without event mappings"):
        _schema(["0", "1"], [None, None], [False, False])


def test_report_counts_do_not_merge_distinct_scalar_types():
    source = pd.DataFrame({"OS event": pd.Series([True, 1, "1"], dtype="object")})
    report_df = SurvMetaReport._get_surv_cat_meta_df(
        df=source,
        col_names=["OS event"],
        rename_mapping={},
    )

    assert len(report_df) == 3
    assert report_df["count"].tolist() == [1, 1, 1]
    assert report_df["category_encoding"].nunique() == 3


def test_mapping_rename_removal_and_missing_values():
    source = pd.DataFrame({"raw_status": ["Alive", "Dead", "Unknown", None]})
    original = source.copy(deep=True)
    schema = _schema(
        ["Alive", "Dead", "Unknown"],
        [False, True, None],
        [False, False, True],
    )

    curated = apply_curation_schemas(
        df=source,
        rename_mapping={"OS event": "raw_status"},
        col_edit_schema=ColEditSchema.empty(),
        cat_meta_edit_schema=CatMetaEditSchema.empty(),
        surv_cat_meta_edit_schema=schema,
        survival_event_columns=("OS event",),
    )

    expected = pd.Series(
        [False, True, pd.NA, pd.NA],
        name="OS event",
        dtype="boolean",
    )
    assert_series_equal(curated["OS event"], expected)
    assert_frame_equal(source, original)


def test_numeric_events_normalize_without_mapping_instructions():
    source = pd.DataFrame({"OS event": [0.0, 1.0, None]})

    curated = apply_curation_schemas(
        df=source,
        rename_mapping={},
        col_edit_schema=ColEditSchema.empty(),
        cat_meta_edit_schema=CatMetaEditSchema.empty(),
        surv_cat_meta_edit_schema=SurvCatMetaEditSchema.empty(),
        survival_event_columns=("OS event",),
    )

    assert_series_equal(
        curated["OS event"],
        pd.Series(
            [False, True, pd.NA],
            name="OS event",
            dtype="boolean",
        ),
    )


def test_unmapped_source_text_is_not_coerced_by_truthiness():
    source = pd.Series(["False", "True"], name="OS event")

    with pytest.raises(ValueError, match="noncanonical value"):
        normalize_survival_event_series(source)


def test_application_also_rejects_partial_programmatic_mapping():
    identity = encode_category_scalar(0)
    edit = SurvCatEdit(
        col_name="OS event",
        category=0,
        category_encoding=identity,
        event_observed=True,
        remove=False,
    )
    schema = SurvCatMetaEditSchema(cat_edits={"OS event": {identity: edit}})

    with pytest.raises(ValueError, match="unmapped category"):
        apply_curation_schemas(
            df=pd.DataFrame({"OS event": [0, 1]}),
            rename_mapping={},
            col_edit_schema=ColEditSchema.empty(),
            cat_meta_edit_schema=CatMetaEditSchema.empty(),
            surv_cat_meta_edit_schema=schema,
            survival_event_columns=("OS event",),
        )


def test_schema_round_trip_preserves_source_types_and_booleans(tmp_path):
    schema = _schema(
        [0, "1", "Unknown"],
        [False, True, None],
        [False, False, True],
    )
    path = tmp_path / "events.parquet"

    schema.save(path)
    loaded = SurvCatMetaEditSchema.load(path)

    numeric_edit = loaded.cat_edits["OS event"][encode_category_scalar(0)]
    text_edit = loaded.cat_edits["OS event"][encode_category_scalar("1")]

    assert isinstance(numeric_edit.category, int)
    assert numeric_edit.event_observed is False
    assert isinstance(text_edit.category, str)
    assert text_edit.event_observed is True
    assert loaded.edit_count == 3


def test_empty_schema_has_stable_parquet_columns(tmp_path):
    path = tmp_path / "empty.parquet"
    SurvCatMetaEditSchema.empty().save(path)

    stored = pd.read_parquet(path)
    assert list(stored.columns) == list(SurvCatMetaEditSchema.PARQUET_SCHEMA)
    assert stored.empty
    assert SurvCatMetaEditSchema.load(path).is_empty


def test_safe_legacy_schema_is_readable(tmp_path):
    path = tmp_path / "legacy.parquet"
    pd.DataFrame(
        {
            "col_name": ["OS event", "OS event"],
            "category": ["Alive", "Dead"],
            "rename_to": ["False", "True"],
            "remove": [False, False],
        }
    ).to_parquet(path, index=False)

    loaded = SurvCatMetaEditSchema.load(path)
    assert (
        loaded.cat_edits["OS event"][encode_category_scalar("Alive")].event_observed
        is False
    )
    assert (
        loaded.cat_edits["OS event"][encode_category_scalar("Dead")].event_observed
        is True
    )


def test_duplicate_persisted_schema_rows_are_rejected(tmp_path):
    schema = _schema(["Dead"], [True], [False])
    path = tmp_path / "duplicate.parquet"
    schema.save(path)
    stored = pd.read_parquet(path)
    pd.concat([stored, stored], ignore_index=True).to_parquet(
        path,
        index=False,
    )

    with pytest.raises(ValueError, match="Duplicate persisted event edit"):
        SurvCatMetaEditSchema.load(path)


def test_excel_dropdowns_and_literal_source_labels(tmp_path):
    reporter = SurvMetaReport()
    profile_path = tmp_path / "profiles.parquet"
    report_path = tmp_path / "events.xlsx"

    reporter.save_semantic_profiles(
        semantic_profiles={
            "OS event": _survival_profile(
                "OS event",
                SurvivalDataTypes.EVENT,
            )
        },
        path=profile_path,
    )

    source = pd.DataFrame(
        {
            "raw_status": pd.Series(
                ["Alive", "Dead", "NA", "1", "0", None],
                dtype="object",
            )
        }
    )
    reporter.create_cat_meta_report(
        df=source,
        rename_mapping={"OS event": "raw_status"},
        profiles_path=profile_path,
        report_path=report_path,
    )

    workbook = load_workbook(report_path)
    worksheet = workbook["SurvCatMeta"]
    headers = {cell.value: cell.column for cell in worksheet[1]}

    assert "event_observed" in headers
    assert "rename_to" not in headers
    assert "rank" not in headers

    encoding_letter = worksheet.cell(
        row=1,
        column=headers["category_encoding"],
    ).column_letter
    assert worksheet.column_dimensions[encoding_letter].hidden

    validations = worksheet.data_validations.dataValidation

    assert len(validations) == 2
    assert all(validation.showDropDown is False for validation in validations)
    assert all(validation.showInputMessage is False for validation in validations)

    # Check that each editable field has its own Boolean dropdown.
    for field_name in ("event_observed", "remove"):
        cell = worksheet.cell(
            row=2,
            column=headers[field_name],
        )
        matching_validations = [
            validation
            for validation in validations
            if cell.coordinate in validation.sqref
        ]

        assert len(matching_validations) == 1
        assert matching_validations[0].formula1 == '"True,False"'

    for row_number in range(2, worksheet.max_row + 1):
        category = worksheet.cell(
            row=row_number,
            column=headers["category"],
        ).value

        if category == "NA":
            worksheet.cell(
                row=row_number,
                column=headers["remove"],
            ).value = True
        else:
            worksheet.cell(
                row=row_number,
                column=headers["event_observed"],
            ).value = category in {"Dead", "1"}

    workbook.save(report_path)
    workbook.close()

    with pd.ExcelFile(report_path) as workbook_reader:
        schema = reporter.get_surv_cat_meta_edit_schema(workbook_reader)

    assert schema.cat_edits["OS event"][encode_category_scalar("NA")].remove
    assert schema.cat_edits["OS event"][encode_category_scalar("1")].category == "1"

    curated = apply_curation_schemas(
        df=source,
        rename_mapping={"OS event": "raw_status"},
        col_edit_schema=ColEditSchema.empty(),
        cat_meta_edit_schema=CatMetaEditSchema.empty(),
        surv_cat_meta_edit_schema=schema,
        survival_event_columns=("OS event",),
    )
    assert_series_equal(
        curated["OS event"],
        pd.Series(
            [False, True, pd.NA, True, False, pd.NA],
            dtype="boolean",
            name="OS event",
        ),
    )


def test_parent_mapping_allows_categories_absent_from_target():
    schema = _schema(
        ["Yes", "No"],
        [True, False],
        [False, False],
    )

    curated = apply_inherited_category_edits(
        target_df=pd.DataFrame({"OS event": ["Yes", "Yes"]}),
        source_cat_meta_edit_schema=CatMetaEditSchema.empty(),
        source_surv_cat_meta_edit_schema=schema,
        column_mapping={},
        changed_columns=["OS event"],
    )

    assert curated["OS event"].tolist() == [True, True]
    assert str(curated["OS event"].dtype) == "boolean"


def test_parquet_output_is_analyzer_compatible(tmp_path):
    frame = pd.DataFrame(
        {
            "time": [10.0, 12.0, 14.0],
            "event": pd.Series([True, False, pd.NA], dtype="boolean"),
        }
    )
    path = tmp_path / "df.parquet"

    write_dataframe_with_category_ranks(
        df=frame,
        path=path,
        category_ranks={},
    )
    reloaded = pd.read_parquet(path)
    prepared = prepare_survival_data(frame=reloaded)

    assert str(reloaded["event"].dtype) == "boolean"
    assert prepared.dropped_rows == 1
    assert prepared.frame["event"].tolist() == [True, False]


def test_cleaner_normalizes_declared_events_before_writing(
    tmp_path,
    monkeypatch,
):
    monkeypatch.chdir(tmp_path)
    source = pd.DataFrame(
        {
            "OS event": [0.0, 1.0, None],
            "OS months": [10.0, 12.0, 14.0],
        }
    )

    project = Project(
        project_name="event_normalization_test",
        project_dir=tmp_path / "projects",
    )
    dataset = project.add_dataset(
        df=source,
        dataset_name="source",
    )
    assert dataset is not None

    cleaner = dataset.cleaner
    bundle = cleaner._get_group_bundle(version=1, config_version=1)
    version_path = bundle["version"]["path"]
    config_path = bundle["config"]["path"]

    profiler = ColProfiler()
    col_profiles = {}
    for col_name in source.columns:
        profile = profiler.get_col_profile(
            col_name=col_name,
            col_series=source[col_name],
        )
        col_profiles[col_name] = replace(
            profile,
            col_type=DataTypes.SURVIVAL,
        )

    cleaner.col_report.save_col_profiles(
        col_profiles=col_profiles,
        path=version_path / "col_profiles_curated.parquet",
    )
    BaseYAML.save(
        data={},
        path=version_path / "rename_mapping.yaml",
        replace=True,
    )
    ColEditSchema.empty().save(version_path / "col_edit_schema.parquet")
    CatMetaEditSchema.empty().save(config_path / "cat_meta_edit_schema.parquet")
    SurvCatMetaEditSchema.empty().save(
        config_path / "surv_cat_meta_edit_schema.parquet"
    )

    event_profile = _survival_profile(
        "OS event",
        SurvivalDataTypes.EVENT,
    )
    time_profile = _survival_profile(
        "OS months",
        SurvivalDataTypes.TIME,
    )
    cleaner.surv_meta_report.save_semantic_profiles(
        semantic_profiles={
            "OS event": event_profile,
            "OS months": time_profile,
        },
        path=config_path / "surv_profiles_curated.parquet",
    )
    SurvPairs(
        pairs={
            "OS": SurvPair(
                surv_label="OS",
                event_profile=event_profile,
                time_profile=time_profile,
            )
        }
    ).save(config_path / "surv_pairs.parquet")

    cleaner.create_curated_data(version=1, config_version=1)

    curated = pd.read_parquet(config_path / "curated_data" / "df.parquet")
    assert_series_equal(
        curated["OS event"],
        pd.Series(
            [False, True, pd.NA],
            name="OS event",
            dtype="boolean",
        ),
    )
    assert_frame_equal(pd.read_parquet(cleaner.df_path), source)

    metadata = bundle["config"]["group"].attrs["meta"]
    assert metadata["survival_event_encoding"]["dtype"] == "boolean"
    assert metadata["survival_event_encoding"]["columns"] == ["OS event"]
