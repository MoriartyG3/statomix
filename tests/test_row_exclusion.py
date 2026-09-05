"""Regression tests for typed, provenance-preserving row exclusions."""

from __future__ import annotations

import json
from dataclasses import replace

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal

from statomix.core.artifacts import DatasetArtifactRef, canonical_json
from statomix.curation.columns import ColProfiler, DataTypes
from statomix.curation.survival.profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
)
from statomix.curation.survival.report import SurvPair, SurvPairs
from statomix.pipelines.transformer.transformer import Transformer
from statomix.reporting.excel.transformation import write_transformation_report
from statomix.storage.artifacts import publish_artifact
from statomix.storage.hashing import sha256_file
from statomix.transformation import MONTHS, ExcludeRows
from statomix.transformation.metadata import ArtifactData, initial_metadata
from statomix.transformation.operations import apply_operations
from statomix.transformation.specifications import operation_from_dict

PATIENT = "CAIB-T00004423OC"
REASON = (
    "Age at surgery recorded as zero; surgery-based PFS and LRC durations "
    "cannot be established from the available source."
)


def _survival_profile(name, role):
    return SurvivalSemanticProfile(
        col_name=name,
        col_type=role,
        score=1.0,
        tokens=(),
        normalized_name=name.casefold(),
        matched_rules=(),
        all_scores={role: 1.0},
    )


def _state():
    frame = pd.DataFrame(
        {
            "patientID": ["P1", PATIENT, "P3"],
            "OS Duration": pd.array(
                [12.0, 0.0, 36.0],
                dtype="Float64",
            ),
            "OS Event": pd.array(
                [False, True, True],
                dtype="boolean",
            ),
            "T Stage": ["T1", "T4", "T2"],
        },
        index=pd.Index([10, 44, 90], name="source_index"),
    )
    roles = {
        "patientID": DataTypes.IDENTIFIER,
        "OS Duration": DataTypes.SURVIVAL,
        "OS Event": DataTypes.SURVIVAL,
        "T Stage": DataTypes.CATEGORICAL,
    }
    profiler = ColProfiler()
    profiles = {
        name: replace(
            profiler.get_col_profile(col_name=name, col_series=frame[name]),
            col_type=role,
        )
        for name, role in roles.items()
    }
    pairs = SurvPairs(
        pairs={
            "OS": SurvPair(
                surv_label="OS",
                event_profile=_survival_profile(
                    "OS Event",
                    SurvivalDataTypes.EVENT,
                ),
                time_profile=_survival_profile(
                    "OS Duration",
                    SurvivalDataTypes.TIME,
                ),
            )
        }
    )
    metadata = initial_metadata(
        frame,
        profiles,
        pairs,
        units={"OS Duration": MONTHS},
        endpoint_definitions={"OS": "All-cause death; origin: surgery"},
    )
    ranks = {
        "schema_version": 1,
        "columns": {
            "T Stage": [
                {"category": {"type": "string", "value": "T1"}, "rank": 0},
                {"category": {"type": "string", "value": "T2"}, "rank": 1},
                {"category": {"type": "string", "value": "T4"}, "rank": 3},
            ]
        },
    }
    lineage = pd.DataFrame(
        {
            "output_row": range(3),
            "parent_artifact": ["months-artifact"] * 3,
            "parent_row": range(3),
            "source_dataset": ["n0_raw"] * 3,
        }
    )
    return ArtifactData(frame, profiles, pairs, metadata, ranks, lineage)


def test_exclusion_preserves_values_metadata_and_lineage():
    parent = _state()
    original = parent.copy()
    output, audit, exclusions = apply_operations(
        parent,
        [
            ExcludeRows(
                identifier="patientID",
                values=(PATIENT,),
                reason=REASON,
            )
        ],
    )

    assert output.df["patientID"].tolist() == ["P1", "P3"]
    assert output.df["OS Event"].tolist() == [False, True]
    assert output.df["OS Duration"].tolist() == [12.0, 36.0]
    assert output.df.index.tolist() == [10, 90]
    assert output.metadata == original.metadata
    assert output.ranks == original.ranks
    assert output.pairs.pairs["OS"].event_profile.col_name == "OS Event"
    assert output.pairs.pairs["OS"].time_profile.col_name == "OS Duration"
    assert output.lineage["output_row"].tolist() == [0, 1]
    assert output.lineage["parent_row"].tolist() == [0, 2]
    assert audit[0]["excluded_rows"] == 1
    assert exclusions[0]["parent_artifact"] == "months-artifact"
    assert exclusions[0]["parent_row"] == 1
    assert json.loads(exclusions[0]["identifier_value"]) == {
        "type": "string",
        "value": PATIENT,
    }
    assert_frame_equal(parent.df, original.df)


def test_specification_round_trip_preserves_identifier_types():
    operation = ExcludeRows(
        identifier="patientID",
        values=("P1", 2, 3.5, False),
        reason="Reviewed exclusions",
    )
    serialized = json.loads(canonical_json(operation.to_dict()))
    restored = operation_from_dict(serialized)
    assert restored.to_dict() == operation.to_dict()
    assert tuple(type(value) for value in restored.values) == (str, int, float, bool)


@pytest.mark.parametrize(
    ("values", "error"),
    [
        ((), ValueError),
        (("",), ValueError),
        (("P1", "P1"), ValueError),
        ((None,), TypeError),
    ],
)
def test_invalid_exclusion_specifications_are_rejected(values, error):
    with pytest.raises(error):
        ExcludeRows(identifier="patientID", values=values, reason="review")


def test_missing_requested_identifier_is_rejected():
    with pytest.raises(KeyError, match="absent"):
        apply_operations(
            _state(),
            [
                ExcludeRows(
                    identifier="patientID",
                    values=("UNKNOWN",),
                    reason="review",
                )
            ],
        )


def test_numeric_identifier_matching_is_type_preserving():
    state = _state()
    state.df["patientID"] = pd.Series(
        [1, 2, 3],
        index=state.df.index,
        dtype="int64",
    )
    state.profiles["patientID"] = replace(
        ColProfiler().get_col_profile(
            col_name="patientID",
            col_series=state.df["patientID"],
        ),
        col_type=DataTypes.IDENTIFIER,
    )
    with pytest.raises(KeyError, match="absent"):
        apply_operations(
            state,
            [
                ExcludeRows(
                    identifier="patientID",
                    values=(2.0,),
                    reason="wrong scalar type",
                )
            ],
        )


@pytest.mark.parametrize("failure", ["duplicate", "missing", "wrong_role", "all"])
def test_unsafe_identifier_selection_is_rejected(failure):
    state = _state()
    values = (PATIENT,)
    if failure == "duplicate":
        state.df.loc[90, "patientID"] = "P1"
    elif failure == "missing":
        state.df.loc[90, "patientID"] = pd.NA
    elif failure == "wrong_role":
        state.profiles["patientID"] = replace(
            state.profiles["patientID"],
            col_type=DataTypes.CATEGORICAL,
        )
    else:
        values = tuple(state.df["patientID"])

    with pytest.raises((TypeError, ValueError)):
        apply_operations(
            state,
            [
                ExcludeRows(
                    identifier="patientID",
                    values=values,
                    reason="review",
                )
            ],
        )


def test_transformer_marks_plans_with_exclusions_as_mixed_operations():
    captured = {}
    transformer = Transformer.__new__(Transformer)

    def capture(**kwargs):
        captured.update(kwargs)
        return kwargs

    transformer._create = capture
    transformer.create_data(
        source="parent",
        operations=[
            ExcludeRows(
                identifier="patientID",
                values=(PATIENT,),
                reason=REASON,
            )
        ],
        version=2,
        config_version=1,
        name="reviewed exclusion",
    )
    assert captured["specification"]["kind"] == "operations"
    assert captured["specification"]["operations"][0]["kind"] == "exclude_rows"


def test_report_contains_a_readable_excluded_rows_sheet(tmp_path):
    _, audit, exclusions = apply_operations(
        _state(),
        [ExcludeRows(identifier="patientID", values=(PATIENT,), reason=REASON)],
    )
    report = tmp_path / "audit.xlsx"
    write_transformation_report(
        path=report,
        audit=audit,
        exclusions=exclusions,
        parents=(),
        specification={"schema_version": 1, "kind": "operations"},
    )
    workbook = load_workbook(report, read_only=True)
    try:
        assert "Excluded Rows" in workbook.sheetnames
        sheet = workbook["Excluded Rows"]
        headers = [cell.value for cell in next(sheet.iter_rows())]
        assert "identifier_value" in headers
        assert "reason" in headers
    finally:
        workbook.close()


def test_published_artifact_checksums_machine_readable_exclusions(tmp_path):
    root = tmp_path.resolve()
    parent_dir = root / "parent"
    parent_dir.mkdir()
    files = {}
    for name in ("df", "col_profiles", "surv_pairs"):
        path = parent_dir / f"{name}.bin"
        path.write_bytes(name.encode())
        files[name] = {
            "path": path.relative_to(root).as_posix(),
            "sha256": sha256_file(path=path),
        }
    parent = DatasetArtifactRef(
        project_root=root,
        manifest_json=canonical_json(
            {
                "schema_version": 1,
                "status": "completed",
                "project": "study",
                "dataset": "n0_raw",
                "pipeline": "transformer",
                "version": 1,
                "config_version": 1,
                "files": files,
            }
        ),
    )
    state, audit, exclusions = apply_operations(
        _state(),
        [ExcludeRows(identifier="patientID", values=(PATIENT,), reason=REASON)],
    )
    reference = publish_artifact(
        project_root=root,
        destination=root / "child",
        state=state,
        identity={
            "project": "study",
            "dataset": "n0_raw",
            "pipeline": "transformer",
            "version": 2,
            "config_version": 1,
        },
        parents=(parent,),
        specification={"schema_version": 1, "kind": "operations"},
        audit=audit,
        exclusions=exclusions,
    )
    machine_audit = pd.read_parquet(reference.path("exclusions"))
    assert machine_audit["parent_row"].tolist() == [1]
    assert machine_audit["reason"].tolist() == [REASON]
    assert reference.manifest["files"]["exclusions"]["sha256"] == sha256_file(
        path=reference.path("exclusions")
    )
