"""Regression tests for two-parent keyed column updates."""

from __future__ import annotations

from dataclasses import replace

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal

from statomix.curation.columns import (
    ColProfiler,
    DataTypes,
)
from statomix.curation.survival.profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
)
from statomix.curation.survival.report import (
    SurvPair,
    SurvPairs,
)
from statomix.pipelines.transformer.transformer import (
    Transformer,
)
from statomix.reporting.excel.transformation import (
    write_transformation_report,
)
from statomix.transformation import (
    MONTHS,
    UpdateColumnsByKey,
)
from statomix.transformation.keyed_updates import (
    apply_keyed_update,
)
from statomix.transformation.metadata import (
    ArtifactData,
    initial_metadata,
)


def survival_profile(
    column_name,
    role,
):
    return SurvivalSemanticProfile(
        col_name=column_name,
        col_type=role,
        score=1.0,
        tokens=(),
        normalized_name=column_name.casefold(),
        matched_rules=(),
        all_scores={
            role: 1.0,
        },
    )


def make_profiles(
    dataframe,
    roles,
):
    profiler = ColProfiler()

    return {
        column_name: replace(
            profiler.get_col_profile(
                col_name=column_name,
                col_series=dataframe[column_name],
            ),
            col_type=role,
        )
        for column_name, role in roles.items()
    }


def make_base_state():
    dataframe = pd.DataFrame(
        {
            "Patient ID": [
                "P1",
                "P2",
            ],
            "OS Event": pd.array(
                [
                    False,
                    True,
                ],
                dtype="boolean",
            ),
            "OS Duration": pd.array(
                [
                    10.0,
                    20.0,
                ],
                dtype="Float64",
            ),
            "Grade": [
                "Low",
                "High",
            ],
        },
        index=pd.Index(
            [
                101,
                205,
            ],
            name="source_index",
        ),
    )

    roles = {
        "Patient ID": DataTypes.IDENTIFIER,
        "OS Event": DataTypes.SURVIVAL,
        "OS Duration": DataTypes.SURVIVAL,
        "Grade": DataTypes.CATEGORICAL,
    }

    profiles = make_profiles(
        dataframe,
        roles,
    )

    pairs = SurvPairs(
        pairs={
            "OS": SurvPair(
                surv_label="OS",
                event_profile=survival_profile(
                    "OS Event",
                    SurvivalDataTypes.EVENT,
                ),
                time_profile=survival_profile(
                    "OS Duration",
                    SurvivalDataTypes.TIME,
                ),
            )
        }
    )

    metadata = initial_metadata(
        dataframe,
        profiles,
        pairs,
        units={
            "OS Duration": MONTHS,
        },
        endpoint_definitions={
            "OS": ("All-cause death; time origin: " "randomisation."),
        },
    )

    ranks = {
        "schema_version": 1,
        "columns": {
            "Grade": [
                {
                    "category": {
                        "type": "string",
                        "value": "Low",
                    },
                    "rank": 0,
                },
                {
                    "category": {
                        "type": "string",
                        "value": "High",
                    },
                    "rank": 1,
                },
            ]
        },
    }

    lineage = pd.DataFrame(
        {
            "output_row": [
                0,
                1,
            ],
            "parent_artifact": [
                "base-artifact",
                "base-artifact",
            ],
            "parent_row": [
                0,
                1,
            ],
            "source_dataset": [
                "ocat_raw",
                "ocat_raw",
            ],
        }
    )

    return ArtifactData(
        dataframe,
        profiles,
        pairs,
        metadata,
        ranks,
        lineage,
    )


def make_update_state(
    *,
    identifiers=(
        "P2",
        "P1",
        "P3",
    ),
):
    event_by_patient = {
        "P1": True,
        "P2": False,
        "P3": True,
    }

    duration_by_patient = {
        "P1": 11.0,
        "P2": 22.0,
        "P3": 33.0,
    }

    dataframe = pd.DataFrame(
        {
            "patientID": list(identifiers),
            "OS Event Corrected": pd.array(
                [event_by_patient[patient] for patient in identifiers],
                dtype="boolean",
            ),
            "OS_months": pd.array(
                [duration_by_patient[patient] for patient in identifiers],
                dtype="Float64",
            ),
        }
    )

    roles = {
        "patientID": DataTypes.IDENTIFIER,
        "OS Event Corrected": (DataTypes.SURVIVAL),
        "OS_months": DataTypes.SURVIVAL,
    }

    profiles = make_profiles(
        dataframe,
        roles,
    )

    pairs = SurvPairs(
        pairs={
            "OS corrected": SurvPair(
                surv_label="OS corrected",
                event_profile=survival_profile(
                    "OS Event Corrected",
                    SurvivalDataTypes.EVENT,
                ),
                time_profile=survival_profile(
                    "OS_months",
                    SurvivalDataTypes.TIME,
                ),
            )
        }
    )

    metadata = initial_metadata(
        dataframe,
        profiles,
        pairs,
        units={
            "OS_months": MONTHS,
        },
        endpoint_definitions={
            "OS corrected": ("All-cause death; time origin: " "randomisation."),
        },
    )

    ranks = {
        "schema_version": 1,
        "columns": {},
    }

    lineage = pd.DataFrame(
        {
            "output_row": range(len(dataframe)),
            "parent_artifact": ["update-artifact"] * len(dataframe),
            "parent_row": range(len(dataframe)),
            "source_dataset": ["ocat_900_survival"] * len(dataframe),
        }
    )

    return ArtifactData(
        dataframe,
        profiles,
        pairs,
        metadata,
        ranks,
        lineage,
    )


def make_operation():
    return UpdateColumnsByKey(
        base_key="Patient ID",
        update_key="patientID",
        column_mapping=(
            (
                "OS Event",
                "OS Event Corrected",
            ),
            (
                "OS Duration",
                "OS_months",
            ),
        ),
        endpoint_mapping=(
            (
                "OS",
                "OS corrected",
            ),
        ),
        reason=("Use the reviewed survival source."),
    )


def test_keyed_update_uses_identity_not_row_position():
    base = make_base_state()
    updates = make_update_state()

    original_base = base.copy()
    original_updates = updates.copy()

    (
        output,
        audit,
        column_updates,
        unused_updates,
    ) = apply_keyed_update(
        base=base,
        updates=updates,
        operation=make_operation(),
    )

    assert output.df.index.tolist() == [
        101,
        205,
    ]

    assert output.df["Patient ID"].tolist() == [
        "P1",
        "P2",
    ]

    assert output.df["OS Event"].tolist() == [
        True,
        False,
    ]

    assert output.df["OS Duration"].tolist() == [
        11.0,
        22.0,
    ]

    assert output.df["Grade"].tolist() == [
        "Low",
        "High",
    ]

    assert output.ranks == (original_base.ranks)

    assert output.lineage.equals(original_base.lineage)

    assert audit[0]["matched_base_rows"] == 2

    assert audit[0]["unused_update_rows"] == 1

    assert audit[0]["audited_cells"] == 4

    assert len(column_updates) == 4

    assert len(unused_updates) == 1

    assert_frame_equal(
        base.df,
        original_base.df,
    )

    assert_frame_equal(
        updates.df,
        original_updates.df,
    )


def test_key_names_may_differ():
    base = make_base_state()
    updates = make_update_state()

    output, _, _, _ = apply_keyed_update(
        base=base,
        updates=updates,
        operation=make_operation(),
    )

    assert "Patient ID" in output.df.columns

    assert "patientID" not in output.df.columns


def test_missing_base_patient_is_rejected():
    base = make_base_state()

    updates = make_update_state(
        identifiers=(
            "P1",
            "P3",
        )
    )

    with pytest.raises(
        KeyError,
        match="does not cover every base patient",
    ):
        apply_keyed_update(
            base=base,
            updates=updates,
            operation=make_operation(),
        )


def test_duplicate_update_key_is_rejected():
    base = make_base_state()

    updates = make_update_state(
        identifiers=(
            "P1",
            "P2",
            "P2",
        )
    )

    with pytest.raises(
        ValueError,
        match="not unique",
    ):
        apply_keyed_update(
            base=base,
            updates=updates,
            operation=make_operation(),
        )


def test_incomplete_endpoint_mapping_is_rejected():
    operation = UpdateColumnsByKey(
        base_key="Patient ID",
        update_key="patientID",
        column_mapping=(
            (
                "OS Event",
                "OS Event Corrected",
            ),
        ),
        endpoint_mapping=(
            (
                "OS",
                "OS corrected",
            ),
        ),
        reason="Incomplete replacement.",
    )

    with pytest.raises(
        ValueError,
        match="requires duration mapping",
    ):
        apply_keyed_update(
            base=make_base_state(),
            updates=make_update_state(),
            operation=operation,
        )


def test_endpoint_definition_mismatch_is_rejected():
    updates = make_update_state()

    updates.metadata["endpoint_definitions"][
        "OS corrected"
    ] = "A different endpoint definition."

    with pytest.raises(
        ValueError,
        match="definitions differ",
    ):
        apply_keyed_update(
            base=make_base_state(),
            updates=updates,
            operation=make_operation(),
        )


def test_public_transformer_specification():
    captured = {}

    transformer = Transformer.__new__(Transformer)

    def capture(**kwargs):
        captured.update(kwargs)
        return kwargs

    transformer._create = capture

    transformer.create_keyed_update_data(
        base="base-artifact",
        updates="update-artifact",
        base_key="Patient ID",
        update_key="patientID",
        column_mapping={
            "OS Event": ("OS Event Corrected"),
            "OS Duration": "OS_months",
        },
        endpoint_mapping={
            "OS": "OS corrected",
        },
        version=1,
        config_version=1,
        reason="Reviewed replacement.",
        name="corrected_survival",
    )

    assert captured["parents"] == (
        "base-artifact",
        "update-artifact",
    )

    specification = captured["specification"]

    assert specification["kind"] == "keyed_update"

    assert specification["base_key"] == "Patient ID"

    assert specification["update_key"] == "patientID"


def test_excel_report_contains_update_audits(
    tmp_path,
):
    (
        _,
        audit,
        column_updates,
        unused_updates,
    ) = apply_keyed_update(
        base=make_base_state(),
        updates=make_update_state(),
        operation=make_operation(),
    )

    report_path = tmp_path / "audit.xlsx"

    write_transformation_report(
        path=report_path,
        audit=audit,
        parents=(),
        specification={
            "schema_version": 1,
            "kind": "keyed_update",
        },
        column_updates=column_updates,
        unused_updates=unused_updates,
    )

    workbook = load_workbook(
        report_path,
        read_only=True,
    )

    try:
        assert "Column Updates" in workbook.sheetnames

        assert "Unused Update Rows" in workbook.sheetnames
    finally:
        workbook.close()
