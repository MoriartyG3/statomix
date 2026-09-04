from __future__ import annotations

from dataclasses import replace

import pandas as pd
import pytest
from openpyxl import load_workbook
from pandas.testing import assert_frame_equal

from statomix import Project
from statomix.core.contracts import AnalyzerInputPaths
from statomix.curation.columns import ColProfiler, DataTypes
from statomix.curation.inheritance import _rebuild_survival_pairs
from statomix.curation.survival.declarations import (
    DECLARATION_FIELDS,
    EVENT_STRUCTURES,
    OBSERVATION_SCHEMES,
)
from statomix.curation.survival.profiler import (
    SurvivalDataTypes,
    SurvivalSemanticProfile,
)
from statomix.curation.survival.report import (
    SurvMetaReport,
    SurvPair,
    SurvPairs,
)
from statomix.pipelines.analyzer.group_analyzer import GroupAnalyzer


def _profile(name, role):
    return SurvivalSemanticProfile(
        col_name=name,
        col_type=role,
        score=1.0,
        tokens=(),
        normalized_name=name.casefold(),
        matched_rules=(),
        all_scores={role: 1.0},
    )


def _profiles():
    return {
        "status": _profile("status", SurvivalDataTypes.EVENT),
        "time": _profile("time", SurvivalDataTypes.TIME),
    }


def _pair(
    event_structure="single_event",
    observation_scheme="right_censored",
):
    profiles = _profiles()
    return SurvPair(
        surv_label="endpoint",
        event_profile=profiles["status"],
        time_profile=profiles["time"],
        event_structure=event_structure,
        observation_scheme=observation_scheme,
    )


def _frame(
    event_structure="single_event",
    observation_scheme="right_censored",
):
    return pd.DataFrame(
        {
            "col_name": ["status", "time"],
            "inferred_datatype": [
                SurvivalDataTypes.EVENT.value,
                SurvivalDataTypes.TIME.value,
            ],
            "change_datatype": ["", ""],
            "survival_label": ["endpoint", "endpoint"],
            "event_structure": [event_structure] * 2,
            "observation_scheme": [observation_scheme] * 2,
            "remove": ["", ""],
        }
    )


@pytest.mark.parametrize("event_structure", EVENT_STRUCTURES)
@pytest.mark.parametrize("observation_scheme", OBSERVATION_SCHEMES)
def test_declarations_round_trip(
    tmp_path,
    event_structure,
    observation_scheme,
):
    path = tmp_path / "pairs.parquet"
    original = _pair(event_structure, observation_scheme)
    SurvPairs(pairs={"endpoint": original}).save(path)

    loaded = SurvPairs.load(path)
    pair = loaded.pairs["endpoint"]

    assert pair.event_structure == event_structure
    assert pair.observation_scheme == observation_scheme

    if event_structure == "single_event" and observation_scheme == "right_censored":
        loaded.require_supported(operation="test")
    else:
        with pytest.raises(
            NotImplementedError,
            match="declared but not implemented",
        ):
            loaded.require_supported(operation="test")


def test_legacy_pairs_receive_defaults(tmp_path):
    row = _pair().to_dict()
    for field in DECLARATION_FIELDS:
        del row[field]

    path = tmp_path / "legacy.parquet"
    pd.DataFrame([row]).to_parquet(path, index=False)

    pair = SurvPairs.load(path).pairs["endpoint"]
    assert pair.event_structure == "single_event"
    assert pair.observation_scheme == "right_censored"


def test_empty_pairs_round_trip(tmp_path):
    path = tmp_path / "empty.parquet"
    SurvPairs.empty().save(path)

    assert SurvPairs.load(path).is_empty
    assert list(pd.read_parquet(path).columns) == list(SurvPairs.PARQUET_SCHEMA)


@pytest.mark.parametrize("field", DECLARATION_FIELDS)
@pytest.mark.parametrize("value", ["", None, "unknown"])
def test_invalid_explicit_declarations_are_rejected(field, value):
    row = _pair().to_dict()
    row[field] = value

    with pytest.raises(ValueError, match=field):
        SurvPair.from_dict(row)


def test_partial_declaration_schema_is_rejected():
    row = _pair().to_dict()
    del row["observation_scheme"]

    with pytest.raises(ValueError, match="missing fields"):
        SurvPair.from_dict(row)


def test_report_declarations_must_agree():
    frame = _frame()
    frame.loc[1, "event_structure"] = "competing_risks"

    with pytest.raises(ValueError, match="inconsistent"):
        SurvMetaReport.get_surv_pairs(frame, _profiles())


def test_nondefault_declaration_requires_endpoint_label():
    frame = _frame("competing_risks")
    frame["survival_label"] = ""

    with pytest.raises(ValueError, match="no survival_label"):
        SurvMetaReport.get_surv_pairs(frame, _profiles())


def test_legacy_report_receives_defaults():
    frame = _frame().drop(columns=list(DECLARATION_FIELDS))
    pairs = SurvMetaReport.get_surv_pairs(frame, _profiles())

    pairs.require_supported(operation="legacy report")


def test_pairing_uses_curated_roles():
    frame = _frame()
    frame["inferred_datatype"] = SurvivalDataTypes.TIME.value

    pairs = SurvMetaReport.get_surv_pairs(frame, _profiles())

    assert pairs.pairs["endpoint"].event_profile.col_name == "status"


def test_excel_declaration_dropdowns(tmp_path):
    path = tmp_path / "report.xlsx"

    SurvMetaReport().create_surv_report_from_profiles(
        col_names=["status", "time"],
        semantic_profiles=_profiles(),
        report_path=path,
        survival_labels={
            "status": "endpoint",
            "time": "endpoint",
        },
    )

    workbook = load_workbook(path)
    try:
        worksheet = workbook["SurvMeta"]
        columns = {cell.value: cell.column for cell in worksheet[1]}

        for field in DECLARATION_FIELDS:
            cell = worksheet.cell(2, columns[field])
            assert cell.protection.locked is False

            validations = [
                validation
                for validation in worksheet.data_validations.dataValidation
                if cell.coordinate in validation.sqref
            ]
            assert len(validations) == 1
            assert validations[0].showDropDown is False
            assert validations[0].showInputMessage is False

        options = workbook["__ValidationRanges__"]
        assert [
            options.cell(row, 4).value for row in range(2, len(EVENT_STRUCTURES) + 2)
        ] == list(EVENT_STRUCTURES)
        assert [
            options.cell(row, 5).value for row in range(2, len(OBSERVATION_SCHEMES) + 2)
        ] == list(OBSERVATION_SCHEMES)
    finally:
        workbook.close()


@pytest.fixture
def dataset_state(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)

    source = pd.DataFrame(
        {
            "status": ["censored", "cause_a", "cause_b"],
            "time": [10.0, 12.0, 14.0],
        }
    )
    project = Project(
        project_name="declaration_tests",
        project_dir=tmp_path / "projects",
    )
    dataset = project.add_dataset(
        df=source,
        dataset_name="source",
    )
    assert dataset is not None

    bundle = dataset.cleaner._get_group_bundle(
        version=1,
        config_version=1,
    )
    return dataset, bundle, source


@pytest.mark.parametrize(
    "method_name",
    [
        "create_surv_cat_meta_report",
        "create_surv_cat_meta_edit_schema",
        "create_curated_data",
    ],
)
def test_cleaner_rejects_unsupported_before_processing(
    dataset_state,
    method_name,
):
    dataset, bundle, source = dataset_state
    config_path = bundle["config"]["path"]

    SurvPairs(pairs={"endpoint": _pair("competing_risks")}).save(
        config_path / "surv_pairs.parquet"
    )

    with pytest.raises(
        NotImplementedError,
        match="declared but not implemented",
    ):
        getattr(dataset.cleaner, method_name)(
            version=1,
            config_version=1,
        )

    assert_frame_equal(
        pd.read_parquet(dataset.cleaner.df_path),
        source,
    )
    assert not (config_path / "curated_data" / "df.parquet").exists()


def test_cached_curated_files_cannot_bypass_guard(dataset_state):
    dataset, bundle, _ = dataset_state
    config_path = bundle["config"]["path"]

    SurvPairs(pairs={"endpoint": _pair("competing_risks")}).save(
        config_path / "surv_pairs.parquet"
    )

    curated_path = config_path / "curated_data"
    curated_path.mkdir(exist_ok=True)

    # Sentinel files must not be read or treated as reusable outputs.
    for filename in (
        "df.parquet",
        "surv_pairs.parquet",
        "col_profiles.parquet",
    ):
        (curated_path / filename).write_bytes(b"sentinel")

    with pytest.raises(NotImplementedError):
        dataset.cleaner.create_curated_data(
            version=1,
            config_version=1,
        )

    assert (curated_path / "df.parquet").read_bytes() == b"sentinel"


def test_analyzer_configuration_rejects_unsupported(dataset_state):
    dataset, bundle, _ = dataset_state

    SurvPairs(pairs={"endpoint": _pair("competing_risks")}).save(
        bundle["config"]["path"] / "surv_pairs.parquet"
    )

    with pytest.raises(NotImplementedError):
        dataset.configure_analyzer(version=1, config_version=1)


def test_group_analyzer_rejects_unsupported_pairs(tmp_path):
    pairs_path = tmp_path / "pairs.parquet"
    SurvPairs(pairs={"endpoint": _pair("competing_risks")}).save(pairs_path)

    analyzer = GroupAnalyzer(
        paths=AnalyzerInputPaths(
            df=tmp_path / "df.parquet",
            surv_pairs=pairs_path,
            col_profiles=tmp_path / "profiles.parquet",
        )
    )

    # Check both initial loading and the in-memory cached object.
    for _ in range(2):
        with pytest.raises(NotImplementedError):
            analyzer._get_surv_pairs()


def test_inheritance_rejects_unsupported_before_value_validation():
    with pytest.raises(NotImplementedError):
        _rebuild_survival_pairs(
            target_df=pd.DataFrame(),
            source_pairs=SurvPairs(pairs={"endpoint": _pair("competing_risks")}),
            survival_profiles={},
            complete_mapping={},
        )


def test_cleaner_persists_declaration_before_blocking(dataset_state):
    dataset, bundle, source = dataset_state
    cleaner = dataset.cleaner
    version_path = bundle["version"]["path"]
    config_path = bundle["config"]["path"]

    profiler = ColProfiler()
    col_profiles = {
        name: replace(
            profiler.get_col_profile(
                col_name=name,
                col_series=source[name],
            ),
            col_type=DataTypes.SURVIVAL,
        )
        for name in source.columns
    }
    cleaner.col_report.save_col_profiles(
        col_profiles=col_profiles,
        path=version_path / "col_profiles_curated.parquet",
    )
    cleaner.surv_meta_report.save_semantic_profiles(
        semantic_profiles=_profiles(),
        path=config_path / "surv_profiles.parquet",
    )

    report_path = config_path / "surv_meta_report_curated.xlsx"
    _frame("competing_risks").to_excel(
        report_path,
        sheet_name="SurvMeta",
        index=False,
    )

    cleaner.create_surv_meta_edit_schema(
        version=1,
        config_version=1,
    )

    persisted = SurvPairs.load(config_path / "surv_pairs.parquet")
    assert persisted.pairs["endpoint"].event_structure == "competing_risks"

    with pytest.raises(NotImplementedError):
        cleaner.create_surv_cat_meta_report(
            version=1,
            config_version=1,
        )
