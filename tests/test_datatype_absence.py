from __future__ import annotations

import pandas as pd
from openpyxl import load_workbook

from statomix.pipelines.analyzer.analysis_config import AnalysisConfig
from statomix.pipelines.analyzer.analyzer import Analyzer
from statomix.pipelines.analyzer.contracts import (
    CATEGORICAL_SUMMARY_SCHEMA,
    NORMALITY_DIAGNOSTIC_SCHEMA,
    NUMERICAL_SUMMARY_SCHEMA,
    SURVIVAL_DESCRIPTIVE_SCHEMA,
)
from statomix.pipelines.analyzer.group_analyzer import GroupAnalyzer
from statomix.pipelines.cleaner.col.col_profiler import ColProfile
from statomix.pipelines.cleaner.col.col_report import ColReport
from statomix.pipelines.cleaner.col.col_semantic_rules import DataTypes
from statomix.pipelines.cleaner.col.datatype_inventory import (
    DatatypeInventory,
)
from statomix.pipelines.cleaner.surv.surv_report import SurvPairs


def _profile(col_name: str, col_type: DataTypes) -> ColProfile:
    return ColProfile(
        col_name=col_name,
        col_type=col_type,
        missing_n=0,
        missing_pct=0.0,
        num_conversion_pct=0.0,
        unique_n=2,
        tokens=[col_name],
        normalized_name=col_name,
    )


def _group_analyzer(tmp_path, df, profiles):
    tmp_path.mkdir(parents=True, exist_ok=True)
    df_path = tmp_path / "df.parquet"
    profiles_path = tmp_path / "col_profiles.parquet"
    surv_pairs_path = tmp_path / "surv_pairs.parquet"

    df.to_parquet(path=df_path, index=False)
    ColReport().save_col_profiles(
        col_profiles=profiles,
        path=profiles_path,
    )
    SurvPairs.empty().save(path=surv_pairs_path)

    return GroupAnalyzer(
        paths={
            "df": df_path,
            "col_profiles": profiles_path,
            "surv_pairs": surv_pairs_path,
        }
    )


class _PathGroup:
    def __init__(self, path):
        self.path = path
        self.path.mkdir(parents=True, exist_ok=True)
        self.attrs = {}
        self._children = {}

    def require_group(self, name):
        if name not in self._children:
            self._children[name] = _PathGroup(self.path / name)
        return self._children[name]


def test_datatype_inventory_contains_every_datatype():
    profiles = {
        "group": _profile(
            col_name="group",
            col_type=DataTypes.CATEGORICAL,
        )
    }

    inventory = DatatypeInventory.from_profiles(profiles=profiles)

    assert set(inventory.columns_by_type) == set(DataTypes)
    assert inventory.columns(datatype=DataTypes.CATEGORICAL) == ("group",)
    assert inventory.columns(datatype=DataTypes.NUMERICAL) == ()
    assert inventory.columns(datatype=DataTypes.SURVIVAL) == ()


def test_no_numerical_columns_produce_typed_empty_results(tmp_path):
    analyzer = _group_analyzer(
        tmp_path=tmp_path,
        df=pd.DataFrame({"group": ["A", "B", "A"]}),
        profiles={
            "group": _profile(
                col_name="group",
                col_type=DataTypes.CATEGORICAL,
            )
        },
    )

    numerical_summary = analyzer.get_num_summary_df()
    normality_diagnostics = analyzer.get_normality_diagnostics_df()

    assert numerical_summary.empty
    assert numerical_summary.index.name == "name"
    assert list(numerical_summary.columns) == list(NUMERICAL_SUMMARY_SCHEMA)

    assert normality_diagnostics.empty
    assert normality_diagnostics.index.name == "name"
    assert list(normality_diagnostics.columns) == list(NORMALITY_DIAGNOSTIC_SCHEMA)


def test_no_categorical_or_survival_columns_remain_valid(tmp_path):
    analyzer = _group_analyzer(
        tmp_path=tmp_path,
        df=pd.DataFrame({"age": [41.0, 52.0, 63.0]}),
        profiles={
            "age": _profile(
                col_name="age",
                col_type=DataTypes.NUMERICAL,
            )
        },
    )

    categorical_summary = analyzer.get_cat_summary_df()
    datatype_map_df = analyzer._get_datatype_map_df()

    assert categorical_summary.empty
    assert categorical_summary.index.names == [
        "col_name",
        "category",
    ]
    assert list(categorical_summary.columns) == list(CATEGORICAL_SUMMARY_SCHEMA)
    assert list(datatype_map_df.columns) == [
        *(datatype.value for datatype in DataTypes),
        "Survival Labels",
    ]
    assert datatype_map_df[DataTypes.CATEGORICAL.value].dropna().empty
    assert datatype_map_df["Survival Labels"].dropna().empty
    assert analyzer._get_surv_pairs().is_empty


def test_no_analyzable_columns_still_create_a_summary_workbook(tmp_path):
    analyzer = _group_analyzer(
        tmp_path=tmp_path,
        df=pd.DataFrame({"patient_id": ["P1", "P2", "P3"]}),
        profiles={
            "patient_id": _profile(
                col_name="patient_id",
                col_type=DataTypes.IDENTIFIER,
            )
        },
    )
    summary_path = tmp_path / "summary.xlsx"

    analyzer.create_summary_report(path=summary_path)

    workbook = load_workbook(filename=summary_path)
    assert workbook.sheetnames == [
        "Numerical",
        "Normality Diagnostics",
        "Categorical",
    ]


def test_analysis_config_accepts_empty_option_lists(
    tmp_path,
    monkeypatch,
):
    datatype_map_df = pd.DataFrame(
        {datatype.value: pd.Series(dtype="object") for datatype in DataTypes}
    )
    datatype_map_df["Survival Labels"] = pd.Series(dtype="object")

    config_path = tmp_path / "analysis_config.xlsx"
    monkeypatch.setattr(
        "statomix.pipelines.analyzer.analysis_config.shutil.copy",
        lambda src, dst: None,
    )

    AnalysisConfig.create_analysis_config(
        path=config_path,
        datatype_map_df=datatype_map_df,
    )

    workbook = load_workbook(filename=config_path)
    assert "Datatype Map" in workbook.sheetnames
    assert "Survival - Multiclass" in workbook.sheetnames
    assert "Survival -Threshold MPV" in workbook.sheetnames

    for sheet_name in (
        "Survival - Multiclass",
        "Survival -Threshold MPV",
    ):
        worksheet = workbook[sheet_name]
        assert len(worksheet.data_validations.dataValidation) == 0


def test_survival_summary_accepts_zero_pairs(tmp_path, monkeypatch):
    group_analyzer = _group_analyzer(
        tmp_path=tmp_path / "analyzer_inputs",
        df=pd.DataFrame({"age": [41.0, 52.0, 63.0]}),
        profiles={
            "age": _profile(
                col_name="age",
                col_type=DataTypes.NUMERICAL,
            )
        },
    )
    config_group = _PathGroup(tmp_path / "analyzer_config")
    group_bundle = {
        "version": {"meta": {"version": 1}},
        "config": {
            "group": config_group,
            "meta": {"version": 1},
        },
    }

    analyzer = Analyzer.__new__(Analyzer)
    analyzer._get_group_bundle = lambda **kwargs: group_bundle
    analyzer._get_group_analyzer = lambda **kwargs: group_analyzer

    monkeypatch.setattr(
        "statomix.pipelines.analyzer.analyzer.BaseZARR.get_abs_path",
        lambda group: group.path,
    )

    analyzer._create_surv_summary_report(
        version=1,
        config_version=1,
    )

    surv_group = config_group.require_group("surv")
    descriptives_path = surv_group.path / "descriptives.xlsx"

    assert descriptives_path.exists()
    assert surv_group.attrs["meta"]["status"] == "not_applicable"
    assert surv_group.attrs["meta"]["reason"] == "no_survival_pairs"

    descriptives_df = pd.read_excel(
        io=descriptives_path,
        index_col=0,
    )
    assert descriptives_df.empty
    assert list(descriptives_df.columns) == list(SURVIVAL_DESCRIPTIVE_SCHEMA)
