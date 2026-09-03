from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from statomix import Project
from statomix.curation.columns import (
    ColReport,
    ColumnAudit,
)


def make_audit_df() -> pd.DataFrame:
    """Create categorical, mixed, event, and continuous columns."""

    return pd.DataFrame(
        {
            "patient_id": [f"P{index:03d}" for index in range(40)],
            "group": (["A", "B", None, "A"] * 10),
            "mixed_measure": (["1", "bad", "3", None] * 10),
            "event": ([0, 1] * 20),
            "duration_months": [float(index) for index in range(40)],
            "continuous": list(range(40)),
        }
    )


def create_direct_report(
    tmp_path: Path,
) -> tuple[
    pd.DataFrame,
    Path,
    Path,
    Path,
]:
    """Create one report without constructing a Project."""

    df = make_audit_df()
    report = ColReport()

    profiles_path = tmp_path / "col_profiles.parquet"
    audit_path = tmp_path / "col_audit.parquet"
    frequencies_path = tmp_path / "col_value_counts.parquet"
    report_path = tmp_path / "col_report.xlsx"

    report.create_col_profiles(
        df=df,
        path=profiles_path,
        replace=False,
    )

    report.create_col_report(
        df=df,
        report_path=report_path,
        profiles_path=profiles_path,
        audit_profiles_path=audit_path,
        value_frequencies_path=frequencies_path,
        value_count_unique_threshold=30,
        report_metadata={
            "dataset_name": "test_dataset",
            "source_row_count": len(df),
        },
        replace=False,
    )

    return (
        df,
        report_path,
        audit_path,
        frequencies_path,
    )


def test_audit_profiles_and_frequencies_are_persisted(
    tmp_path,
) -> None:
    (
        df,
        report_path,
        audit_path,
        frequencies_path,
    ) = create_direct_report(tmp_path=tmp_path)

    assert report_path.exists()
    assert audit_path.exists()
    assert frequencies_path.exists()

    audit = ColumnAudit.load(
        profiles_path=audit_path,
        value_frequencies_path=frequencies_path,
    )

    assert list(audit.profiles) == list(df.columns)

    group_profile = audit.profiles["group"]

    assert group_profile.source_dtype == "object"
    assert group_profile.missing_n == 10
    assert group_profile.missing_pct == 25.0
    assert group_profile.unique_n == 2
    assert group_profile.numeric_n == 0
    assert group_profile.nonnumeric_n == 30
    assert group_profile.exact_value_counts_included is True

    mixed_profile = audit.profiles["mixed_measure"]

    assert mixed_profile.numeric_n == 20
    assert mixed_profile.nonnumeric_n == 10
    assert mixed_profile.minimum == 1.0
    assert mixed_profile.median == 2.0
    assert mixed_profile.maximum == 3.0

    continuous_profile = audit.profiles["continuous"]

    assert continuous_profile.numeric_n == 40
    assert continuous_profile.nonnumeric_n == 0
    assert continuous_profile.minimum == 0.0
    assert continuous_profile.q1 == 9.75
    assert continuous_profile.median == 19.5
    assert continuous_profile.q3 == 29.25
    assert continuous_profile.maximum == 39.0
    assert continuous_profile.exact_value_counts_included is False

    group_frequencies = {
        frequency.value_display: frequency.count
        for frequency in audit.value_frequencies
        if frequency.col_name == "group"
    }

    assert group_frequencies == {
        "'A'": 20,
        "'B'": 10,
        "<MISSING>": 10,
    }

    continuous_frequencies = [
        frequency
        for frequency in audit.value_frequencies
        if frequency.col_name == "continuous"
    ]

    assert continuous_frequencies == []


def test_integrated_workbook_contains_protected_audit_fields(
    tmp_path,
) -> None:
    (
        _,
        report_path,
        _,
        _,
    ) = create_direct_report(tmp_path=tmp_path)

    workbook = load_workbook(report_path)

    assert "Value Counts" in workbook.sheetnames
    assert "Report Metadata" in workbook.sheetnames
    assert "__ValidationRanges__" in (workbook.sheetnames)

    validation_sheet = workbook["__ValidationRanges__"]
    assert validation_sheet.sheet_state == "veryHidden"

    categorical_sheet = workbook["Categorical"]

    headers = {cell.value: cell.column for cell in categorical_sheet[1]}

    assert "source_dtype" in headers
    assert "missing_n" in headers
    assert "missing_pct" in headers
    assert "unique_n" in headers
    assert "num_conversion_pct" in headers
    assert "minimum" in headers
    assert "median" in headers
    assert "maximum" in headers

    group_row = next(
        row_index
        for row_index in range(
            2,
            categorical_sheet.max_row + 1,
        )
        if categorical_sheet.cell(
            row=row_index,
            column=headers["col_name"],
        ).value
        == "group"
    )

    assert (
        categorical_sheet.cell(
            row=group_row,
            column=headers["change_col_name"],
        ).protection.locked
        is False
    )

    assert (
        categorical_sheet.cell(
            row=group_row,
            column=headers["source_dtype"],
        ).protection.locked
        is True
    )

    assert categorical_sheet.protection.sheet is True

    value_counts_sheet = workbook["Value Counts"]

    assert value_counts_sheet.protection.sheet is True
    assert value_counts_sheet["A2"].protection.locked is True

    workbook.close()


def test_audit_worksheets_are_not_parsed_as_edits(
    tmp_path,
) -> None:
    (
        _,
        report_path,
        _,
        _,
    ) = create_direct_report(tmp_path=tmp_path)

    workbook = load_workbook(report_path)
    categorical_sheet = workbook["Categorical"]

    headers = {cell.value: cell.column for cell in categorical_sheet[1]}

    group_row = next(
        row_index
        for row_index in range(
            2,
            categorical_sheet.max_row + 1,
        )
        if categorical_sheet.cell(
            row=row_index,
            column=headers["col_name"],
        ).value
        == "group"
    )

    categorical_sheet.cell(
        row=group_row,
        column=headers["change_col_name"],
        value="Study Group",
    )

    curated_path = tmp_path / "col_report_curated.xlsx"
    workbook.save(curated_path)
    workbook.close()

    rename_mapping, edit_schema = ColReport().get_col_edit_schema(
        curated_col_report=pd.ExcelFile(curated_path)
    )

    assert rename_mapping == {"Study Group": "group"}
    assert set(edit_schema.edits) == {"group"}
    assert edit_schema.edits["group"].change_col_name == "Study Group"


def test_cleaner_reuses_complete_artifacts_and_creates_new_version(
    tmp_path,
    monkeypatch,
) -> None:
    monkeypatch.chdir(tmp_path)

    project = Project(
        project_name="column_audit_test",
        project_dir=tmp_path / "projects",
    )

    dataset = project.add_dataset(
        df=make_audit_df(),
        dataset_name="cohort",
        display_label="Test cohort",
    )

    assert dataset is not None

    cleaner = dataset.cleaner

    cleaner.create_col_report(
        version=1,
        version_name="initial",
        config_version=1,
        config_name="initial",
        create_new=False,
    )

    version_one = cleaner._find_group_bundle(
        version=1,
        config_version=1,
    )
    version_one_path = Path(version_one["version"]["path"])

    artifact_names = (
        "col_report.xlsx",
        "col_profiles.parquet",
        "col_audit.parquet",
        "col_value_counts.parquet",
    )

    original_artifacts = {
        artifact_name: (version_one_path / artifact_name).read_bytes()
        for artifact_name in artifact_names
    }

    cleaner.create_col_report(
        version=1,
        config_version=1,
        create_new=False,
    )

    reused_artifacts = {
        artifact_name: (version_one_path / artifact_name).read_bytes()
        for artifact_name in artifact_names
    }

    assert reused_artifacts == original_artifacts

    cleaner.create_col_report(
        version=1,
        version_name="integrated_audit",
        config_version=1,
        config_name="initial",
        create_new=True,
    )

    version_two = cleaner._find_group_bundle(
        version=2,
        config_version=1,
    )
    version_two_path = Path(version_two["version"]["path"])

    for artifact_name in artifact_names:
        assert (version_two_path / artifact_name).exists()

    assert cleaner.meta["latest_version"] == 2

    # Version 1 remains byte-for-byte unchanged.
    for artifact_name, original_bytes in original_artifacts.items():
        assert (version_one_path / artifact_name).read_bytes() == original_bytes


def test_incomplete_existing_artifact_set_is_rejected(
    tmp_path,
    monkeypatch,
) -> None:
    monkeypatch.chdir(tmp_path)

    project = Project(
        project_name="partial_audit_test",
        project_dir=tmp_path / "projects",
    )

    dataset = project.add_dataset(
        df=make_audit_df(),
        dataset_name="cohort",
    )

    assert dataset is not None

    cleaner = dataset.cleaner

    cleaner.create_col_report(
        version=1,
        config_version=1,
        create_new=False,
    )

    version_one = cleaner._find_group_bundle(
        version=1,
        config_version=1,
    )
    version_one_path = Path(version_one["version"]["path"])

    (version_one_path / "col_value_counts.parquet").unlink()

    with pytest.raises(
        RuntimeError,
        match=("Incomplete column-report artifact set"),
    ):
        cleaner.create_col_report(
            version=1,
            config_version=1,
            create_new=False,
        )


def test_invalid_frequency_threshold_is_rejected(
    tmp_path,
) -> None:
    df = make_audit_df()
    report = ColReport()

    profiles_path = tmp_path / "col_profiles.parquet"

    report.create_col_profiles(
        df=df,
        path=profiles_path,
        replace=False,
    )

    profiles = report.load_col_profiles(path=profiles_path)

    with pytest.raises(
        ValueError,
        match=("value_count_unique_threshold " "must be at least 1"),
    ):
        ColumnAudit.from_dataframe(
            df=df,
            col_profiles=profiles,
            value_count_unique_threshold=0,
        )
