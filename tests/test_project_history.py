"""Regression tests for read-only project-history reports."""

from __future__ import annotations

import hashlib
import json
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from statomix import Project
from statomix.history import (
    HistoryEdge,
    HistoryNode,
    ProjectHistory,
    discover_project_history,
)
from statomix.history.discovery import ProjectHistoryDiscovery
from statomix.history.validation import validate_graph
from statomix.transformation import MONTHS


def _hash_tree(root: Path) -> dict[str, str]:
    return {
        path.relative_to(root).as_posix(): hashlib.sha256(path.read_bytes()).hexdigest()
        for path in sorted(root.rglob("*"))
        if path.is_file()
    }


@pytest.fixture
def project(tmp_path, monkeypatch):
    monkeypatch.chdir(tmp_path)
    project = Project(
        project_name="history_test",
        project_dir=tmp_path / "projects",
    )
    project.add_dataset(
        df=pd.DataFrame(
            {
                "Patient ID": ["P1", "P2", "P3"],
                "OS Event": [0, 1, 0],
                "OS Duration": [12.0, 24.0, 36.0],
            }
        ),
        dataset_name="analysis_cohort",
        display_label="Analysis cohort",
    )
    return project


def _reference_artifact(project):
    dataset = project.datasets["analysis_cohort"]
    project.set_dataset_role(
        dataset_name="analysis_cohort",
        dataset_role="reference",
        reason="Create a typed reference artifact for the history test.",
    )
    artifact = dataset.create_reference_artifact(
        version=1,
        config_version=1,
        identifier="Patient ID",
        column_mapping={
            "Patient ID": "Patient ID",
            "OS Event": "OS Event",
            "OS Duration": "OS Duration",
        },
        event_columns={"OS Event": {0: False, 1: True}},
        duration_units={"OS Duration": MONTHS},
        endpoints={
            "OS": {
                "event": "OS Event",
                "duration": "OS Duration",
                "definition": "Overall survival from randomisation.",
            }
        },
        reason="Reviewed external survival reference.",
        name="reviewed_survival",
    )
    project.set_dataset_role(
        dataset_name="analysis_cohort",
        dataset_role="analysis",
        reason="Restore analysis eligibility after publishing the test artifact.",
    )
    return dataset, artifact


def test_history_identity_is_deterministic():
    node = HistoryNode(
        node_id="source:a",
        node_type="source",
        label="A",
        dataset="a",
        display_label="A",
        dataset_role="analysis",
        pipeline="source",
        status="completed",
    )
    first = ProjectHistory(project_name="P", nodes=(node,), edges=())
    second = ProjectHistory(project_name="P", nodes=(node,), edges=())

    assert first.history_id == second.history_id
    assert first.to_dict() == second.to_dict()


def test_graph_validation_reports_dangling_edges_and_cycles():
    nodes = tuple(
        HistoryNode(
            node_id=name,
            node_type="transformer",
            label=name,
            dataset="d",
            display_label="D",
            dataset_role="analysis",
            pipeline="transformer",
        )
        for name in ("a", "b")
    )
    edges = (
        HistoryEdge(source="a", target="b", relationship="derived_from"),
        HistoryEdge(source="b", target="a", relationship="derived_from"),
        HistoryEdge(source="missing", target="a", relationship="derived_from"),
    )

    warnings = validate_graph(nodes=nodes, edges=edges)

    assert {warning.code for warning in warnings} == {
        "dangling_edge",
        "provenance_cycle",
    }


def test_source_only_report_is_external_atomic_and_read_only(project, tmp_path):
    project_root = (project.project_dir / project.project_name).resolve()
    before = _hash_tree(project_root)

    report = project.create_history_report(
        output_dir=tmp_path / "history_reports",
        verify_checksums=True,
    )

    after = _hash_tree(project_root)
    assert after == before
    assert report.directory.parent == (tmp_path / "history_reports").resolve()
    assert report.directory.name.endswith(report.history_id[:12])
    assert report.html_path.is_file()
    assert report.svg_path.is_file()
    assert report.json_path.is_file()
    assert report.audit_path.is_file()

    payload = json.loads(report.json_path.read_text(encoding="utf-8"))
    assert payload["history_id"] == report.history_id
    assert len(payload["nodes"]) == 1
    assert payload["nodes"][0]["node_type"] == "source"
    assert payload["nodes"][0]["rows"] == 3
    assert payload["nodes"][0]["columns"] == 3

    html = report.html_path.read_text(encoding="utf-8")
    assert "dataset-filter" in html
    assert "pipeline-filter" in html
    assert "Show report nodes" in html
    assert "No warnings detected" in html

    svg = report.svg_path.read_text(encoding="utf-8")
    assert "artifact lineage" in svg
    assert "Analysis cohort" in svg

    workbook = load_workbook(report.audit_path, read_only=True)
    try:
        assert {"Nodes", "Edges", "Warnings", "Node Attributes"}.issubset(
            workbook.sheetnames
        )
    finally:
        workbook.close()

    reused = project.create_history_report(
        output_dir=tmp_path / "history_reports",
        verify_checksums=True,
    )
    assert reused.directory == report.directory
    assert _hash_tree(project_root) == before


def test_report_rejects_output_inside_project_store(project):
    project_root = project.project_dir / project.project_name

    with pytest.raises(ValueError, match="outside the Statomix project store"):
        project.create_history_report(
            output_dir=project_root / "history_reports",
        )


def test_reference_artifact_and_analyzer_binding_are_connected(
    project,
    tmp_path,
):
    _, artifact = _reference_artifact(project)
    analysis = project.datasets["analysis_cohort"]
    analysis.configure_analyzer_from_artifact(
        source=artifact,
        version=1,
        config_version=1,
        survival_evaluation={
            "OS": {
                "unit": MONTHS,
                "time_points": [12.0, 24.0],
            }
        },
    )
    bundle = analysis.analyzer._find_group_bundle(
        version=1,
        config_version=1,
    )
    (bundle["config"]["path"] / "summary.xlsx").write_bytes(b"summary")

    history = discover_project_history(
        project=project,
        verify_checksums=True,
        include_files=True,
    )

    artifact_node_id = f"artifact:{artifact.artifact_id}"
    analyzer_node_id = "analyzer:analysis_cohort:v1:c1"
    report_node_id = f"report:{analyzer_node_id}:summary"
    node_ids = {node.node_id for node in history.nodes}
    edge_keys = {
        (edge.source, edge.target, edge.relationship) for edge in history.edges
    }

    assert artifact_node_id in node_ids
    assert analyzer_node_id in node_ids
    assert report_node_id in node_ids
    assert (
        "source:analysis_cohort",
        artifact_node_id,
        "reference_from",
    ) in edge_keys
    assert (
        artifact_node_id,
        analyzer_node_id,
        "analyzed_from",
    ) in edge_keys
    assert (
        analyzer_node_id,
        report_node_id,
        "reported_by",
    ) in edge_keys

    artifact_node = next(
        node for node in history.nodes if node.node_id == artifact_node_id
    )
    assert artifact_node.status == "completed"
    assert artifact_node.rows == 3
    assert artifact_node.columns == 3
    assert artifact_node.attributes["checksum_status"] == "verified"
    assert artifact_node.attributes["survival_endpoints"] == {
        "OS": {
            "event": "OS Event",
            "duration": "OS Duration",
        }
    }
    assert "files" in artifact_node.attributes

    report = project.create_history_report(
        output_dir=tmp_path / "reports",
        include_files=True,
    )
    workbook = load_workbook(report.audit_path, read_only=True)
    try:
        assert "Files" in workbook.sheetnames
    finally:
        workbook.close()


def test_checksum_mismatch_is_visible(project):
    _, artifact = _reference_artifact(project)
    audit_path = artifact.path("audit")
    audit_path.write_bytes(audit_path.read_bytes() + b"modified")

    history = discover_project_history(
        project=project,
        verify_checksums=True,
    )

    artifact_node = next(
        node
        for node in history.nodes
        if node.node_id == f"artifact:{artifact.artifact_id}"
    )
    assert artifact_node.status == "invalid"
    assert artifact_node.attributes["checksum_status"] == "failed"
    assert "artifact_checksum_mismatch" in {
        warning.code for warning in history.warnings
    }


@pytest.mark.parametrize(
    ("pipeline", "kind", "index", "expected"),
    [
        ("transformer", "keyed_update", 0, "base_parent"),
        ("transformer", "keyed_update", 1, "update_parent"),
        ("transformer", "concatenate", 2, "concatenated_from"),
        ("transformer", "columns", 0, "transformed_from"),
        ("reference", "reference", 0, "derived_from"),
    ],
)
def test_parent_relationships_are_typed(pipeline, kind, index, expected):
    relationship = ProjectHistoryDiscovery._parent_relationship(
        pipeline=pipeline,
        specification={"kind": kind},
        parent_index=index,
    )
    assert relationship == expected
