import pandas as pd

from pathlib import Path

from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from statomix.semantic_rules import DataTypes
from statomix.col_profiler import ColProfiler, ColProfile


class ColReport:
    def __init__(self, df: pd.DataFrame, df_name: str):
        self.df = df
        self.df_name = df_name

        self.col_profiler = ColProfiler(cat_unique_thresh=4, num_conversion_thresh=95)

    def create_col_report_default(self, report_path: Path):
        assert report_path.suffix == ".xlsx", "report_path should be a .xlsx path."

        self._create_col_profiles()
        self._create_col_report(report_path=report_path)
        self._format_cell_length(report_path=report_path)
        self._add_validation_datatype(report_path=report_path)
        self._add_validation_categories(
            report_path=report_path, datatype=DataTypes.CATEGORICAL.value
        )
        self._add_validation_categories(
            report_path=report_path, datatype=DataTypes.SURVIVAL.value
        )

    def _create_col_profiles(self):

        col_profiles: dict[str, ColProfile] = {}
        for col_name in self.df.columns:
            col_series = self.df[col_name]

            col_profile = self.col_profiler.get_col_profile(
                col_name=col_name,
                col_series=col_series,
            )

            col_profiles[col_name] = col_profile

        self.col_profiles = col_profiles

    def _get_col_names_by_type(self, datatype):
        return [
            profile.col_name
            for profile in self.col_profiles.values()
            if profile.col_type == datatype
        ]

    def _create_col_report(self, report_path):
        """
        Creates a raw column report without any formatting.
        """
        # file_path = self.report_dir/"col_report.xlsx"

        sheet_map = {
            datatype: [
                profile.col_name
                for profile in self.col_profiles.values()
                if profile.col_type == datatype
            ]
            for datatype in DataTypes
        }

        profiled_cols_n = sum(len(col_names) for col_names in sheet_map.values())

        assert profiled_cols_n == len(self.df.columns)

        # Create the .xlsx file
        with pd.ExcelWriter(path=report_path, engine="openpyxl") as writer:
            for datatype, col_names in sheet_map.items():
                rows = []

                for col_name in col_names:
                    profile = self.col_profiles[col_name]

                    if datatype == DataTypes.DATETIME:
                        rows.append(
                            {
                                "col_name": profile.col_name,
                                "change_col_name": "",  # profile.col_name,
                                "inferred_datatype": profile.col_type.value,
                                "change_datatype": "",  # profile.col_type.value,
                                "format": "",
                                "remove": "",
                            }
                        )
                    else:
                        rows.append(
                            {
                                "col_name": profile.col_name,
                                "change_col_name": "",  # profile.col_name,
                                "inferred_datatype": profile.col_type.value,
                                "change_datatype": "",  # profile.col_type.value,
                                "remove": "",
                            }
                        )

                df_sheet = pd.DataFrame(data=rows)

                sheet_name = datatype.value[
                    :31
                ]  # Excel worksheet names have a 31-character limit.

                df_sheet.to_excel(
                    excel_writer=writer,
                    sheet_name=sheet_name,
                    index=False,
                )

    def _format_cell_length(self, report_path):
        workbook = load_workbook(filename=report_path)

        for worksheet in workbook.worksheets:

            for cell in worksheet[1]:
                cell.font = Font(bold=True)

            for column_cells in worksheet.columns:
                max_length = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )

                adjusted_width = max_length + 4

                column_letter = get_column_letter(column_cells[0].column)

                worksheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(filename=report_path)

    def _add_validation_datatype(self, report_path):
        workbook = load_workbook(filename=report_path)

        dropdown_options = [datatype.value for datatype in DataTypes]

        dropdown_string = ",".join(dropdown_options)

        for worksheet in workbook.worksheets:

            # Add Dropdown
            dropdown_validation = DataValidation(
                type="list",
                formula1=f'"{dropdown_string}"',
                allow_blank=True,
            )

            worksheet.add_data_validation(dropdown_validation)

            max_row = worksheet.max_row
            if max_row >= 2:
                dropdown_validation.add(f"D2:D{max_row}")

        workbook.save(filename=report_path)

    def _add_validation_categories(self, report_path, datatype):

        workbook = load_workbook(filename=report_path)

        worksheet = workbook[datatype]

        for cell in worksheet["A"][1:]:
            categories = list(self.df[cell.value].unique())

            if len(categories) >= 20:
                continue

            dropdown_string = ",".join(map(str, categories))

            dropdown_validation = DataValidation(
                type="list",
                formula1=f'"{dropdown_string}"',
                allow_blank=True,
            )

            worksheet.add_data_validation(dropdown_validation)

            dropdown_validation.add(cell.coordinate)

        workbook.save(filename=report_path)
